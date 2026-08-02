"""プロジェクトタスク管理ルート。

管理職・マスタは全タスクを閲覧・編集できる。
一般ユーザーは自分に割り当てられたタスクのみ参照可能（編集不可）。
"""
from __future__ import annotations

import io
import json
import logging
from datetime import date, timedelta

import openpyxl
from flask import (
    Blueprint, abort, flash, jsonify, redirect, render_template,
    request, send_file, session, url_for,
)
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..auth_helpers import can_access_user, is_master, is_privileged
from ..models import (
    PROJECT_TASK_STATUSES,
    add_project_task,
    delete_project_task,
    delete_routine_task,
    get_accessible_users,
    get_accessible_users_for_dashboard,
    get_all_categories,
    get_all_project_tasks,
    get_all_subcategories,
    get_all_users,
    get_project_task_by_id,
    get_routine_schedules,
    get_subcategory_category_id,
    get_task_master,
    get_user_by_id,
    get_task_overview_summary,
    get_task_progress_summary,
    import_brabio_excel,
    save_routine_task,
    update_project_task,
)

logger = logging.getLogger(__name__)


def _resolve_category(cat_id: str, subcat_id: str) -> tuple[int | None, int | None]:
    """フォームの大区分/中区分IDから、保存用の (大区分ID, 中区分ID) を決定する。

    中区分が選択されている場合は、その親を大区分として優先採用する
    （中区分は必ず大区分に属するため、大区分の登録漏れ・不整合を防ぐ）。

    Args:
        cat_id: フォームの category_id 文字列。
        subcat_id: フォームの subcategory_id 文字列。

    Returns:
        tuple[int | None, int | None]: (大区分ID, 中区分ID)。
    """
    category_id = int(cat_id) if cat_id else None
    subcategory_id = int(subcat_id) if subcat_id else None
    if subcategory_id:
        parent = get_subcategory_category_id(subcategory_id)
        if parent is not None:
            category_id = parent
    return category_id, subcategory_id

project_tasks_bp = Blueprint(
    "project_tasks_bp", __name__, url_prefix="/project-tasks",
)


def _parse_member_ids(raw_values: list[str]) -> list[str]:
    """イベント関係者のフォーム値を整数ID文字列のリスト（重複除去・順序維持）に整形する。

    ``<select multiple>``（各値が個別要素）と、👤ボタン方式のカンマ区切り単一値の
    どちらの送信形式にも対応するため、各要素をさらにカンマで分割してから検証する。

    Args:
        raw_values: request.form.getlist で取得した生の値リスト。

    Returns:
        list[str]: 実在チェック前の整数ID文字列（重複除去・入力順維持）。
    """
    seen: set[str] = set()
    result: list[str] = []
    for raw in raw_values:
        for part in str(raw or "").split(","):
            part = part.strip()
            if not part:
                continue
            try:
                sid = str(int(part))
            except (TypeError, ValueError):
                continue
            if sid not in seen:
                seen.add(sid)
                result.append(sid)
    return result


def _user_can_edit_event(task: dict, user_id: int) -> bool:
    """ログインユーザーが当該イベントの関係者（=編集可能）かどうかを判定する。

    event_member_ids（カンマ区切り）または assigned_to / assigned_to_2 に
    ユーザーIDが含まれていれば編集可能とみなす。

    Args:
        task: project_task の辞書。
        user_id: ログインユーザーID。

    Returns:
        bool: 編集可能なら True。
    """
    members = {
        s.strip() for s in (task.get("event_member_ids") or "").split(",") if s.strip()
    }
    return (
        str(user_id) in members
        or task.get("assigned_to") == user_id
        or task.get("assigned_to_2") == user_id
    )


def _can_touch_gantt_task(task: dict) -> bool:
    """ガントチャート画面（gantt/update-dates 等）でログインユーザーが対象タスクを
    操作できるか判定する。

    マスタは常に可。それ以外は、タスクの担当者（assigned_to/assigned_to_2）に
    自分が含まれるか、または管理職・所属長が can_access_user で担当者にアクセス
    可能な場合のみ許可する。

    Args:
        task: project_task の辞書。

    Returns:
        bool: 操作可能なら True。
    """
    login_role = session.get("user_role", "")
    login_id = int(session["user_id"])
    if is_master(login_role):
        return True
    related_ids = {task.get("assigned_to"), task.get("assigned_to_2")}
    related_ids.discard(None)
    if login_id in related_ids:
        return True
    if not is_privileged(login_role):
        return False
    login_user_dict = {
        "id": login_id, "role": login_role, "dept": session.get("user_dept", ""),
    }
    for uid in related_ids:
        u = get_user_by_id(uid)
        if u and can_access_user(login_user_dict, u):
            return True
    return False


def _resolve_routine_target_user_id(login_id: int) -> int:
    """定例作業の登録・削除における代理操作対象ユーザーIDを解決する。

    フォームの target_user_id が指定されていれば、所属長・管理職・
    システム管理者のスコープ内（get_accessible_users）かを確認し、範囲内
    ならその ID を、範囲外・未指定・一般ユーザーならログイン本人を返す。

    Args:
        login_id: ログインユーザーのID。

    Returns:
        int: 代理操作対象ユーザーID。
    """
    login_role = session.get("user_role", "")
    form_user_id = request.form.get("target_user_id", "").strip()
    if not form_user_id or not is_privileged(login_role):
        return login_id
    try:
        candidate = int(form_user_id)
    except ValueError:
        return login_id
    login_dept = session.get("user_dept", "")
    accessible = get_accessible_users(login_id, login_role, login_dept)
    if any(u["id"] == candidate for u in accessible):
        return candidate
    return login_id


@project_tasks_bp.before_request
def _check_login() -> object | None:
    """未ログインならログイン画面へリダイレクトする。"""
    if not session.get("user_id"):
        return redirect(url_for("auth.login"))
    return None


@project_tasks_bp.route("/")
def task_list() -> str:
    """プロジェクトタスク一覧画面を表示する。

    管理職・マスタ: 全タスク表示（編集可）
    一般ユーザー: 自分のタスクのみ表示（参照のみ）

    Returns:
        str: レンダリング済みHTML
    """
    return _render_task_page(None)


@project_tasks_bp.route("/routine")
def routine_page() -> str:
    """定例作業の専用設定画面を表示する（上部ナビの「定例作業」）。"""
    return _render_task_page("routine")


@project_tasks_bp.route("/events")
def events_page() -> str:
    """イベントの専用設定画面を表示する（上部ナビの「イベント」）。"""
    return _render_task_page("events")


def _render_task_page(only_tab: str | None) -> str:
    """タスク管理／定例作業／イベントの各画面を描画する共通処理。

    Args:
        only_tab: None=タスク一覧, 'routine'=定例作業, 'events'=イベント。
    """
    login_role = session.get("user_role", "")
    login_id = int(session["user_id"])
    login_dept = session.get("user_dept", "")
    # 記憶トークンで自動ログインした古いセッションは user_dept が空の場合があるため、
    # 空ならDBから補完してセッションにも保存する（関係者候補の部署絞りが機能しなくなるのを防ぐ）。
    if not login_dept:
        _u = get_user_by_id(login_id)
        if _u:
            login_dept = _u.get("dept", "") or ""
            session["user_dept"] = login_dept
    privileged = is_privileged(login_role)
    is_events = only_tab == "events"
    # 実効所属：システム管理者が所属切替中なら active_dept、通常は自分の所属。
    # これにより「所属を切り替えるとその所属のメンバーで表示される」を実現する。
    from ..auth_helpers import (
        get_active_scope_dept, get_effective_depts, is_dept_head, is_system_admin,
    )
    from ..models import get_user_affiliations, get_all_users_by_depts
    scope_dept = get_active_scope_dept(
        login_role, login_dept, session.get("active_dept")
    )

    def _scoped_users() -> list[dict]:
        """実効所属に属するユーザー一覧を返す（役職別の絞り込みを一元化）。

        - システム管理者: scope_dept があればその所属、無ければ全ユーザー。
        - 所属長: 担当所属（active_dept で1所属に絞られる場合あり）に属するユーザー。
          担当所属が無ければ空（他所属を覗かせない）。
        - 管理職・一般: 自分の主所属のみ。
        """
        if is_dept_head(login_role):
            depts = get_effective_depts(
                {"role": login_role, "dept": login_dept},
                affiliations=get_user_affiliations(login_id),
                active_dept=session.get("active_dept"),
            )
            return get_all_users_by_depts(depts or [])
        if is_system_admin(login_role) and not scope_dept:
            return get_all_users()  # 全所属表示
        return get_all_users(dept_filter=scope_dept)

    # マスタ権限: ユーザー切替対応（デフォルトは全員）
    target_user_id: int = login_id
    selectable_users: list[dict] = []
    if is_events:
        # イベント画面：所属長・管理職・マスタは配下（実効所属）メンバー全員または
        # 個別メンバーを切替閲覧できる（既定＝全員）。一般ユーザーは自分自身のみで
        # 切替プルダウン自体を出さない（selectable_users を空のままにする）。
        if privileged:
            selectable_users = _scoped_users()
            allowed_ids = {u["id"] for u in selectable_users} | {login_id}
            req_uid = request.args.get("user_id", "").strip()
            if req_uid:
                try:
                    cand = int(req_uid)
                except ValueError:
                    cand = 0
                # 権限外ユーザーのタスクを URL 直打ちで覗けないよう、許可範囲に照合する。
                target_user_id = cand if (cand in allowed_ids or cand == 0) else 0
            else:
                target_user_id = 0  # 既定は全員（配下・実効所属メンバー）
            if target_user_id == 0:
                member_ids = [u["id"] for u in selectable_users]
                tasks = get_all_project_tasks(user_ids=member_ids)
            else:
                tasks = get_all_project_tasks(assigned_to=target_user_id)
        else:
            target_user_id = login_id  # 一般ユーザーは自分の関係イベントのみ
            tasks = get_all_project_tasks(assigned_to=target_user_id)
    elif privileged:
        # 管理職・所属長・システム管理者：get_accessible_users のスコープ内メンバーを
        # 切替閲覧できる（既定＝全員）。権限外ユーザーの指定は許可範囲に照合してフォールバック。
        selectable_users = get_accessible_users(login_id, login_role, login_dept)
        allowed_ids = {u["id"] for u in selectable_users} | {login_id}
        req_uid = request.args.get("user_id", "").strip()
        if req_uid is not None and req_uid != "":
            try:
                cand = int(req_uid)
            except ValueError:
                cand = 0
            target_user_id = cand if (cand in allowed_ids or cand == 0) else 0
        elif session.get("task_selected_user_id") is not None:
            target_user_id = int(session["task_selected_user_id"])
            if target_user_id not in allowed_ids and target_user_id != 0:
                target_user_id = 0
        else:
            target_user_id = 0  # デフォルトは全員
        session["task_selected_user_id"] = target_user_id
        # 「全員」選択時は全メンバーのタスクを表示
        if target_user_id == 0:
            member_ids = [u["id"] for u in selectable_users]
            tasks = get_all_project_tasks(user_ids=member_ids)
        else:
            tasks = get_all_project_tasks(assigned_to=target_user_id)
    else:
        tasks = get_all_project_tasks(assigned_to=login_id)

    categories = get_all_categories()
    subcategories = get_all_subcategories()

    # 関係者選択用のユーザーリスト（users=全員: 既存関係者の名前表示に使用）
    users = get_all_users()
    # イベント関係者の「選択候補」は実効所属に限定する（所属長は担当所属のみ・他所属を出さない）。
    assign_users = _scoped_users()

    # JS の人物選択ポップ用に安全な形へ整形（tojson で一括出力し、手書きループの構文事故を防ぐ）。
    def _disp(u: dict) -> str:
        return (u.get("last_name") or u.get("name") or "").strip()
    assign_users_js = [{"id": u["id"], "name": _disp(u)} for u in assign_users]
    all_name_js = {str(u["id"]): _disp(u) for u in users}

    # イベント画面: 各イベントに can_edit（ログインユーザーが関係者かどうか）を付与する。
    # 自分が関係者（event_member_ids / assigned_to / assigned_to_2）に含まれるイベントのみ編集可。
    if is_events:
        for t in tasks:
            member_ids = {
                s.strip() for s in (t.get("event_member_ids") or "").split(",") if s.strip()
            }
            related = (
                str(login_id) in member_ids
                or t.get("assigned_to") == login_id
                or t.get("assigned_to_2") == login_id
            )
            t["can_edit"] = related

    # 定例スケジュール（作業名は自由入力のため、作業マスタからの候補取得は不要）
    # 「切り替え」中は代理操作対象（target_user_id）の定例を表示する。
    # イベントタブは target_user_id=0（全員）の場合があるため、その時は本人の定例を表示する。
    routine_user_id = target_user_id if target_user_id else login_id
    routine_schedules = get_routine_schedules(routine_user_id)
    used_rows = {r["row_number"] for r in routine_schedules}

    return render_template(
        "project_tasks.html",
        tasks=tasks,
        categories=categories,
        subcategories=subcategories,
        statuses=PROJECT_TASK_STATUSES,
        privileged=privileged,
        users=users,
        assign_users=assign_users,
        assign_users_js=assign_users_js,
        all_name_js=all_name_js,
        routine_schedules=routine_schedules,
        used_rows=used_rows,
        selectable_users=selectable_users,
        target_user_id=target_user_id,
        is_master=is_master(login_role),
        only_tab=only_tab,
    )


@project_tasks_bp.route("/add", methods=["POST"])
def add_task() -> object:
    """プロジェクトタスクを追加する。

    イベントは全ユーザー、通常タスクは管理職・マスタのみ。

    Returns:
        object: 一覧画面へのリダイレクト
    """
    is_event_add = request.form.get("is_event") == "1"
    if request.form.get("csrf_token") != session.get("csrf_token"):
        abort(400)

    cat_id = request.form.get("category_id", "")
    subcat_id = request.form.get("subcategory_id", "")
    task_name = request.form.get("task_name", "").strip()
    description = request.form.get("description", "").strip()
    assigned_to_str = request.form.get("assigned_to", "").strip()
    start_date = request.form.get("start_date", "").strip()
    end_date = request.form.get("end_date", "").strip()
    # イベントは開催日のみのため end_date が空の場合は start_date を使用
    if is_event_add and not end_date:
        end_date = start_date
    status = request.form.get("status", "未着手")
    progress_str = request.form.get("progress", "0").strip()
    delay_str = request.form.get("delay_days", "0").strip()

    if not task_name or not start_date or not end_date:
        flash("タスク名・開始日・終了日は必須です。", "warning")
        return redirect(url_for("project_tasks_bp.task_list"))

    if status not in PROJECT_TASK_STATUSES:
        status = "未着手"

    try:
        progress = max(0, min(int(progress_str), 100))
    except ValueError:
        progress = 0
    try:
        delay_days = max(0, int(delay_str))
    except ValueError:
        delay_days = 0
    # 担当者1: 未選択ならログインユーザーを自動セット
    try:
        assigned_to = int(assigned_to_str) if assigned_to_str else int(session["user_id"])
    except ValueError:
        assigned_to = int(session["user_id"])
    assigned_to_2_str = request.form.get("assigned_to_2", "").strip()
    try:
        assigned_to_2 = int(assigned_to_2_str) if assigned_to_2_str else None
    except ValueError:
        assigned_to_2 = None

    is_milestone = 1 if request.form.get("is_milestone") == "1" else 0
    is_event = 1 if is_event_add else 0
    event_start_time = request.form.get("event_start_time", "").strip() if is_event_add else ""
    event_end_time = request.form.get("event_end_time", "").strip() if is_event_add else ""
    planned_hours_str = request.form.get("planned_hours", "0").strip()
    try:
        planned_hours = max(0.0, float(planned_hours_str))
    except ValueError:
        planned_hours = 0.0
    # 担当者ごとの予定反映フラグ（チェックボックス）。未送信は OFF として扱う。
    import_to_schedule_1 = 1 if request.form.get("import_to_schedule_1") == "1" else 0
    import_to_schedule_2 = 1 if request.form.get("import_to_schedule_2") == "1" else 0

    # イベントの場合は参加者を複数受け取り、event_member_ids に保存。
    # 先頭2人は assigned_to / assigned_to_2 にも反映して既存機能との互換を保つ。
    event_member_ids: str = ""
    if is_event_add:
        unique_ids = _parse_member_ids(request.form.getlist("event_members"))
        event_member_ids = ",".join(unique_ids)
        if unique_ids:
            try:
                assigned_to = int(unique_ids[0])
            except ValueError:
                pass
            if len(unique_ids) >= 2:
                try:
                    assigned_to_2 = int(unique_ids[1])
                except ValueError:
                    assigned_to_2 = None
            else:
                assigned_to_2 = None

    category_id_val, subcategory_id_val = _resolve_category(cat_id, subcat_id)
    add_project_task(
        category_id=category_id_val,
        subcategory_id=subcategory_id_val,
        task_name=task_name,
        description=description,
        start_date=start_date,
        end_date=end_date,
        status=status,
        progress=progress,
        delay_days=delay_days,
        created_by=int(session["user_id"]),
        updated_by=session.get("user_name", ""),
        assigned_to=assigned_to,
        assigned_to_2=assigned_to_2,
        is_milestone=is_milestone,
        is_event=is_event,
        event_start_time=event_start_time,
        event_end_time=event_end_time,
        planned_hours=planned_hours,
        import_to_schedule_1=import_to_schedule_1,
        import_to_schedule_2=import_to_schedule_2,
        event_member_ids=event_member_ids,
    )
    flash("タスクを追加しました。", "success")

    # ガントチャート画面から追加された場合は、絞り込み状態を引き継いでガントチャートに戻る
    return_to: str = request.form.get("return_to", "").strip()
    if return_to == "gantt":
        from urllib.parse import quote as _quote
        params: list[str] = []
        for src_key, url_key in (
            ("return_subcat_filter", "subcat_filter"),
            ("return_period_start", "period_start"),
            ("return_period_end", "period_end"),
        ):
            v = request.form.get(src_key, "").strip()
            if v:
                params.append(f"{url_key}={_quote(v)}")
        ruid = request.form.get("return_user_id", "").strip()
        if ruid and ruid != "0":
            params.append(f"user_id={_quote(ruid)}")
        redir = url_for("project_tasks_bp.gantt")
        if params:
            redir += "?" + "&".join(params)
        return redirect(redir)

    # イベント追加時はイベント専用画面へ戻る
    # （カレンダーの表示月はクライアント側の状態のため、直前に見ていた月を
    #   隠しフィールド経由で受け取り、リダイレクト先URLに引き継いで当月に
    #   戻ってしまわないようにする）。
    if is_event == 1:
        events_redir = url_for("project_tasks_bp.events_page") + "?add_open=1&tab=events"
        return_year = request.form.get("return_year", "").strip()
        return_month = request.form.get("return_month", "").strip()
        if return_year.isdigit() and return_month.isdigit():
            events_redir += f"&year={return_year}&month={return_month}"
        return redirect(events_redir)

    return redirect(url_for("project_tasks_bp.task_list") + "?add_open=1")


@project_tasks_bp.route("/update/<int:task_id>", methods=["POST"])
def update_task(task_id: int) -> object:
    """プロジェクトタスクを更新する（管理職・マスタのみ）。

    Args:
        task_id: タスクID

    Returns:
        object: 一覧画面へのリダイレクト
    """
    if request.form.get("csrf_token") != session.get("csrf_token"):
        abort(400)

    if not is_privileged(session.get("user_role", "")):
        abort(403)

    existing = get_project_task_by_id(task_id)
    if not existing:
        abort(404)
    if not _can_touch_gantt_task(existing):
        abort(403)

    cat_id = request.form.get("category_id", "")
    subcat_id = request.form.get("subcategory_id", "")
    task_name = request.form.get("task_name", "").strip()
    description = request.form.get("description", "").strip()
    assigned_to_str = request.form.get("assigned_to", "").strip()
    start_date = request.form.get("start_date", "").strip()
    end_date = request.form.get("end_date", "").strip()
    status = request.form.get("status", "未着手")
    progress_str = request.form.get("progress", "0").strip()
    delay_str = request.form.get("delay_days", "0").strip()

    if not task_name or not start_date or not end_date:
        flash("タスク名・開始日・終了日は必須です。", "warning")
        return redirect(url_for("project_tasks_bp.task_list"))

    if status not in PROJECT_TASK_STATUSES:
        status = "未着手"

    try:
        progress = max(0, min(int(progress_str), 100))
    except ValueError:
        progress = 0
    try:
        delay_days = max(0, int(delay_str))
    except ValueError:
        delay_days = 0
    try:
        assigned_to = int(assigned_to_str) if assigned_to_str else None
    except ValueError:
        assigned_to = None
    assigned_to_2_str = request.form.get("assigned_to_2", "").strip()
    try:
        assigned_to_2 = int(assigned_to_2_str) if assigned_to_2_str else None
    except ValueError:
        assigned_to_2 = None

    is_milestone = 1 if request.form.get("is_milestone") == "1" else 0
    is_event = 1 if request.form.get("is_event") == "1" else 0
    event_start_time = request.form.get("event_start_time", "").strip()
    event_end_time = request.form.get("event_end_time", "").strip()
    planned_hours_str = request.form.get("planned_hours", "0").strip()
    try:
        planned_hours = max(0.0, float(planned_hours_str))
    except ValueError:
        planned_hours = 0.0
    # 担当者ごとの予定反映フラグ
    import_to_schedule_1 = 1 if request.form.get("import_to_schedule_1") == "1" else 0
    import_to_schedule_2 = 1 if request.form.get("import_to_schedule_2") == "1" else 0

    category_id_val, subcategory_id_val = _resolve_category(cat_id, subcat_id)
    update_project_task(
        task_id=task_id,
        category_id=category_id_val,
        subcategory_id=subcategory_id_val,
        task_name=task_name,
        description=description,
        start_date=start_date,
        end_date=end_date,
        status=status,
        progress=progress,
        delay_days=delay_days,
        updated_by=session.get("user_name", ""),
        assigned_to=assigned_to,
        assigned_to_2=assigned_to_2,
        is_milestone=is_milestone,
        is_event=is_event,
        event_start_time=event_start_time,
        event_end_time=event_end_time,
        planned_hours=planned_hours,
        import_to_schedule_1=import_to_schedule_1,
        import_to_schedule_2=import_to_schedule_2,
    )
    flash("タスクを更新しました。", "success")
    return redirect(url_for("project_tasks_bp.task_list"))


@project_tasks_bp.route("/bulk-update", methods=["POST"])
def bulk_update_tasks() -> object:
    """プロジェクトタスクを一括更新する（管理職・マスタのみ）。

    Returns:
        object: 一覧画面へのリダイレクト
    """
    if request.form.get("csrf_token") != session.get("csrf_token"):
        abort(400)

    if not is_privileged(session.get("user_role", "")):
        abort(403)

    login_id = int(session["user_id"])
    task_ids_raw = request.form.getlist("task_id")
    updated_count = 0
    deleted_count = 0

    for raw_id in task_ids_raw:
        try:
            task_id = int(raw_id)
        except ValueError:
            continue

        existing = get_project_task_by_id(task_id)
        if not existing:
            continue

        # イベントは関係者（編集可能）以外による更新・削除を拒否する。
        # 画面上は閲覧のみ行としてフォームに含めないが、直接POSTへの防御も行う。
        if existing.get("is_event") and not _user_can_edit_event(existing, login_id):
            continue
        # 通常タスクは、担当者本人か can_access_user が許可する範囲のみ操作可能。
        if not existing.get("is_event") and not _can_touch_gantt_task(existing):
            continue

        # 削除チェックボックスが ON の場合は削除して次へ
        if request.form.get(f"delete_{task_id}"):
            delete_project_task(task_id)
            deleted_count += 1
            continue

        sfx = f"_{task_id}"
        task_name = request.form.get(f"task_name{sfx}", "").strip()
        description = request.form.get(f"description{sfx}", "").strip()
        start_date = request.form.get(f"start_date{sfx}", "").strip()
        end_date = request.form.get(f"end_date{sfx}", "").strip()
        status = request.form.get(f"status{sfx}", "未着手")
        progress_str = request.form.get(f"progress{sfx}", "0").strip()
        delay_str = request.form.get(f"delay_days{sfx}", "0").strip()
        assigned_to_str = request.form.get(f"assigned_to{sfx}", "").strip()
        assigned_to_2_str = request.form.get(f"assigned_to_2{sfx}", "").strip()
        cat_id = request.form.get(f"category_id{sfx}", "")
        subcat_id = request.form.get(f"subcategory_id{sfx}", "")

        if not task_name or not start_date or not end_date:
            continue

        if status not in PROJECT_TASK_STATUSES:
            status = "未着手"
        try:
            progress = max(0, min(int(progress_str), 100))
        except ValueError:
            progress = 0
        try:
            delay_days = max(0, int(delay_str))
        except ValueError:
            delay_days = 0
        try:
            assigned_to = int(assigned_to_str) if assigned_to_str else None
        except ValueError:
            assigned_to = None
        try:
            assigned_to_2 = int(assigned_to_2_str) if assigned_to_2_str else None
        except ValueError:
            assigned_to_2 = None

        is_milestone = 1 if request.form.get(f"is_milestone{sfx}") == "1" else 0
        is_event = 1 if request.form.get(f"is_event{sfx}") == "1" else 0
        event_start_time = request.form.get(f"event_start_time{sfx}", "").strip()
        event_end_time = request.form.get(f"event_end_time{sfx}", "").strip()
        planned_hours_str = request.form.get(f"planned_hours{sfx}", "0").strip()
        try:
            planned_hours_val = max(0.0, float(planned_hours_str))
        except ValueError:
            planned_hours_val = 0.0
        # 担当者ごとの予定反映フラグ
        import_to_schedule_1 = 1 if request.form.get(f"import_to_schedule_1{sfx}") == "1" else 0
        import_to_schedule_2 = 1 if request.form.get(f"import_to_schedule_2{sfx}") == "1" else 0

        # イベントの場合は参加者を複数受け取り、event_member_ids に保存
        event_member_ids: str = ""
        if is_event == 1:
            unique_ids = _parse_member_ids(request.form.getlist(f"event_members{sfx}"))
            event_member_ids = ",".join(unique_ids)
            # 先頭2人を assigned_to / assigned_to_2 にも反映（互換性保持）
            if unique_ids:
                try:
                    assigned_to = int(unique_ids[0])
                except ValueError:
                    pass
                if len(unique_ids) >= 2:
                    try:
                        assigned_to_2 = int(unique_ids[1])
                    except ValueError:
                        assigned_to_2 = None
                else:
                    assigned_to_2 = None

        category_id_val, subcategory_id_val = _resolve_category(cat_id, subcat_id)
        update_project_task(
            task_id=task_id,
            category_id=category_id_val,
            subcategory_id=subcategory_id_val,
            task_name=task_name,
            description=description,
            start_date=start_date,
            end_date=end_date,
            status=status,
            progress=progress,
            delay_days=delay_days,
            updated_by=session.get("user_name", ""),
            assigned_to=assigned_to,
            assigned_to_2=assigned_to_2,
            is_milestone=is_milestone,
            is_event=is_event,
            event_start_time=event_start_time,
            event_end_time=event_end_time,
            planned_hours=planned_hours_val,
            import_to_schedule_1=import_to_schedule_1,
            import_to_schedule_2=import_to_schedule_2,
            event_member_ids=event_member_ids,
        )
        updated_count += 1

    msgs = []
    if updated_count:
        msgs.append(f"{updated_count}件更新")
    if deleted_count:
        msgs.append(f"{deleted_count}件削除")
    flash("、".join(msgs) + "しました。" if msgs else "変更はありませんでした。", "success")
    subcat_f = request.form.get("subcat_filter", "")
    from_tab = request.form.get("from_tab", "").strip()
    # イベント一括更新はイベント専用画面へ戻る
    if from_tab == "events":
        redir = url_for("project_tasks_bp.events_page")
    else:
        redir = url_for("project_tasks_bp.task_list")
        if subcat_f:
            redir += f"?subcat_filter={subcat_f}"
    return redirect(redir)


@project_tasks_bp.route("/import-brabio", methods=["POST"])
def import_brabio() -> object:
    """ブラビオExcelからタスクをインポートする（管理職・マスタのみ）。

    data/ ディレクトリ内のExcelファイルを読み込み、
    タスク管理にインポートする。

    Returns:
        object: 一覧画面へのリダイレクト
    """
    import pathlib

    if request.form.get("csrf_token") != session.get("csrf_token"):
        abort(400)

    # アップロードファイル or デフォルトファイル
    uploaded = request.files.get("brabio_file")
    if uploaded and uploaded.filename:
        # アップロードされたファイルを一時保存
        import tempfile
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        uploaded.save(tmp.name)
        file_path = tmp.name
    else:
        # data/ または reports/ ディレクトリのデフォルトファイルを検索
        project_root = pathlib.Path(__file__).resolve().parent.parent.parent
        candidates = [
            project_root / "reports" / "商品開発業務.xlsx",
            project_root / "data" / "ブラビオ商品開発業務.xlsx",
        ]
        # reports/ 内の .xlsx を追加検索
        reports_dir = project_root / "reports"
        if reports_dir.is_dir():
            for f in reports_dir.glob("*.xlsx"):
                if f not in candidates:
                    candidates.append(f)
        file_path = None
        for cand in candidates:
            if cand.exists():
                file_path = str(cand)
                break
        if file_path is None:
            flash("インポートファイルが見つかりません", "warning")
            return redirect(url_for("project_tasks_bp.task_list"))

    result = import_brabio_excel(
        file_path=file_path,
        created_by=int(session["user_id"]),
        updated_by=session.get("user_name", ""),
    )

    parts = []
    if result["imported"]:
        parts.append(f"{result['imported']}件取込")
    if result.get("updated"):
        parts.append(f"{result['updated']}件更新")
    msg = f"インポート完了: {' / '.join(parts) if parts else '変更なし'}"
    if result["errors"]:
        msg += f" / {len(result['errors'])}件エラー"
    flash(msg, "success" if (result["imported"] > 0 or result.get("updated", 0) > 0) else "info")
    return redirect(url_for("project_tasks_bp.task_list"))


@project_tasks_bp.route("/delete/<int:task_id>", methods=["POST"])
def delete_task(task_id: int) -> object:
    """プロジェクトタスクを削除する（管理職・マスタのみ）。

    Args:
        task_id: タスクID

    Returns:
        object: 一覧画面へのリダイレクト
    """
    if request.form.get("csrf_token") != session.get("csrf_token"):
        abort(400)

    if not is_privileged(session.get("user_role", "")):
        abort(403)

    existing = get_project_task_by_id(task_id)
    if not existing:
        abort(404)
    if not _can_touch_gantt_task(existing):
        abort(403)

    delete_project_task(task_id)
    flash("タスクを削除しました。", "success")
    return redirect(url_for("project_tasks_bp.task_list"))


@project_tasks_bp.route("/routine/save", methods=["POST"])
def save_routine() -> object:
    """定例スケジュールを登録する。

    「切り替え」中は、フォームの target_user_id が示す代理操作対象に登録する
    （所属長・管理職・システム管理者は get_accessible_users のスコープ内のみ許可）。

    Returns:
        object: タスク管理画面へのリダイレクト
    """
    if request.form.get("csrf_token") != session.get("csrf_token"):
        abort(400)
    login_id = int(session["user_id"])
    user_id = _resolve_routine_target_user_id(login_id)
    task_name = request.form.get("task_name", "").strip()
    subcategory_name = request.form.get("subcategory_name", "").strip()
    period = request.form.get("period", "").strip().upper()
    default_hours_str = request.form.get("default_hours", "0").strip()

    redirect_url = url_for("project_tasks_bp.routine_page") + "?add_open=1&tab=routine"

    if not task_name:
        flash("作業名を選択してください。", "warning")
        return redirect(redirect_url)

    # 区分（AM=1〜3 / PM=6〜8）から空いている行番号を自動割当する。各区分最大3件。
    if period not in ("AM", "PM"):
        flash("区分（午前／午後）を選択してください。", "warning")
        return redirect(redirect_url)
    used = {r["row_number"] for r in get_routine_schedules(user_id)}
    candidates = [1, 2, 3] if period == "AM" else [6, 7, 8]
    row_number = next((n for n in candidates if n not in used), None)
    if row_number is None:
        label = "午前" if period == "AM" else "午後"
        flash(f"{label}の定例は最大3件までです。既存を削除してから追加してください。", "warning")
        return redirect(redirect_url)
    try:
        default_hours = max(0.0, float(default_hours_str))
    except ValueError:
        default_hours = 0.0

    # 曜日フラグ（チェックボックスから取得）
    days_list = []
    for di in range(5):
        days_list.append("1" if request.form.get(f"day_{di}") else "0")
    days = ",".join(days_list)

    # 詰め方向（空きスロットへの割当を上から/下からのどちらにするか）
    fill_direction = request.form.get("fill_direction", "top").strip()
    if fill_direction not in ("top", "bottom"):
        fill_direction = "top"

    ok = save_routine_task(
        user_id, task_name, subcategory_name, default_hours, row_number, days, fill_direction
    )
    flash("定例スケジュールを登録しました。" if ok else "登録に失敗しました（行番号重複の可能性）。",
          "success" if ok else "warning")
    return redirect(redirect_url)


@project_tasks_bp.route("/routine/delete/<int:routine_id>", methods=["POST"])
def delete_routine(routine_id: int) -> object:
    """定例スケジュールを削除する。

    「切り替え」中は、フォームの target_user_id が示す代理操作対象から削除する
    （所属長・管理職・システム管理者は get_accessible_users のスコープ内のみ許可）。

    Args:
        routine_id: 定例スケジュールID

    Returns:
        object: タスク管理画面へのリダイレクト
    """
    if request.form.get("csrf_token") != session.get("csrf_token"):
        abort(400)
    login_id = int(session["user_id"])
    user_id = _resolve_routine_target_user_id(login_id)
    delete_routine_task(routine_id, user_id)
    flash("定例スケジュールを削除しました。", "success")
    # 削除後は定例作業画面をそのまま維持する
    return redirect(url_for("project_tasks_bp.routine_page"))


# -- ステータス→色マッピング（ダッシュボード用） --
_STATUS_COLOR_MAP: dict[str, str] = {
    "未着手": "#9ca3af",
    "着手": "#60a5fa",
    "順調": "#34d399",
    "遅れ": "#f87171",
    "完了": "#10b981",
    "停止": "#d1d5db",
}


def _build_chart_json(summary: dict) -> dict:
    """サマリー情報からグラフ描画用のJSON構造を構築する。

    タスク別進捗のラベルは「親タスク名　子タスク名」を1行にまとめる。同じ親が
    続く間は親名を省略し、同じ文字数分を全角スペースで埋めて子タスク名の位置を
    揃える（親は複数行にまたがっても表示は1回だけ）。

    Args:
        summary: get_task_progress_summary() の戻り値。

    Returns:
        dict: ステータス別集計とタスク別進捗を含むグラフ用辞書。
    """
    status_breakdown: dict[str, int] = summary["status_breakdown"]
    status_labels: list[str] = list(status_breakdown.keys())
    status_counts: list[int] = list(status_breakdown.values())
    status_colors: list[str] = [
        _STATUS_COLOR_MAP.get(s, "#9ca3af") for s in status_labels
    ]

    # 親タスク名の解決用（対象ユーザーの担当タスクだけでは親情報が欠けるため全件から引く）。
    parent_name_by_id: dict[int, str] = {
        t["id"]: t["task_name"] for t in get_all_project_tasks()
    }

    task_names: list[str] = []
    task_progresses: list[float | None] = []
    task_colors: list[str] = []
    task_statuses: list[str] = []

    # タスク別進捗のラベルは「親タスク名　子タスク名」を1行にまとめる。
    # 同じ親が続く間は親名を省略し、親名と同じ文字数分を全角スペースで
    # 埋めて子タスク名の位置を揃える（親は複数行にまたがっても見出しは1回だけ）。
    seen_parent_ids: set[int] = set()
    parent_indent_by_id: dict[int, str] = {}
    for task in summary["tasks"]:
        parent_id = task.get("parent_task_id")
        task_name = task.get("task_name", "")
        if parent_id and parent_id in parent_name_by_id:
            parent_name = parent_name_by_id[parent_id]
            if parent_id not in seen_parent_ids:
                seen_parent_ids.add(parent_id)
                parent_indent_by_id[parent_id] = "　" * len(parent_name)
                label = f"{parent_name}　{task_name}"
            else:
                label = f"{parent_indent_by_id[parent_id]}　{task_name}"
        else:
            label = task_name

        task_names.append(label)
        task_progresses.append(float(task.get("progress", 0) or 0))
        status: str = task.get("status", "未着手")
        task_statuses.append(status)
        task_colors.append(_STATUS_COLOR_MAP.get(status, "#9ca3af"))

    return {
        "status_labels": status_labels,
        "status_counts": status_counts,
        "status_colors": status_colors,
        "task_names": task_names,
        "task_progresses": task_progresses,
        "task_colors": task_colors,
        "task_statuses": task_statuses,
    }


def _resolve_dashboard_target(
    login_id: int, login_role: str, login_dept: str,
) -> tuple[int, str, list[dict]]:
    """ダッシュボード表示対象のユーザーIDと名前、選択可能ユーザーリストを返す。

    Args:
        login_id: ログインユーザーID。
        login_role: ログインユーザーの役職。
        login_dept: ログインユーザーの部署。

    Returns:
        tuple: (target_user_id, target_user_name, selectable_users)

    Raises:
        Werkzeug 403: 権限外のユーザーを指定した場合。
    """
    selectable_users: list[dict] = get_accessible_users_for_dashboard(
        login_id, login_role, login_dept,
    )
    privileged: bool = is_privileged(login_role)

    # クエリパラメータからユーザーIDを取得（なければセッションの選択ユーザーを維持）
    raw_user_id: str | None = request.args.get("user_id")
    if raw_user_id is not None:
        try:
            target_user_id: int = int(raw_user_id)
        except (ValueError, TypeError):
            abort(400)
        session["selected_user_id"] = target_user_id
    elif privileged and session.get("selected_user_id"):
        target_user_id = int(session["selected_user_id"])
    else:
        target_user_id = login_id

    # 権限チェック（自分自身は常に許可）
    if target_user_id != login_id:
        if not privileged:
            abort(403)
        else:
            accessible_ids: set[int] = {u["id"] for u in selectable_users}
            if target_user_id not in accessible_ids:
                abort(403)

    # 対象ユーザー名を解決
    target_user_name: str = session.get("user_name", "")
    for u in selectable_users:
        if u["id"] == target_user_id:
            target_user_name = u["name"]
            break

    return target_user_id, target_user_name, selectable_users


@project_tasks_bp.route("/overview")
def task_overview() -> str:
    """管理者・マスタ向けタスク全体俯瞰ダッシュボードを表示する。

    全タスクのステータス別円グラフ、カテゴリ別進捗バーチャート、
    担当者別集計を表示する。管理職・マスタのみアクセス可。

    Returns:
        str: レンダリング済みHTML
    """
    login_role: str = session.get("user_role", "")
    if not is_privileged(login_role):
        abort(403)

    summary: dict = get_task_overview_summary()
    chart_json: dict = _build_overview_chart_json(summary)

    return render_template(
        "project_tasks_overview.html",
        summary=summary,
        chart_json=chart_json,
    )


def _build_overview_chart_json(summary: dict) -> dict:
    """全体俯瞰ダッシュボード用のグラフJSON構造を構築する。

    Args:
        summary: get_task_overview_summary() の戻り値。

    Returns:
        dict: ステータス別・カテゴリ別・担当者別のグラフ描画データ。
    """
    status_breakdown: dict[str, int] = summary["status_breakdown"]
    status_labels: list[str] = list(status_breakdown.keys())
    status_counts: list[int] = list(status_breakdown.values())
    status_colors: list[str] = [
        _STATUS_COLOR_MAP.get(s, "#9ca3af") for s in status_labels
    ]

    # カテゴリ別
    cat_names: list[str] = [c["name"] for c in summary["category_summary"]]
    cat_avg_progress: list[float] = [c["avg_progress"] for c in summary["category_summary"]]
    cat_counts: list[int] = [c["count"] for c in summary["category_summary"]]
    cat_completed: list[int] = [c["completed"] for c in summary["category_summary"]]
    cat_delayed: list[int] = [c["delayed"] for c in summary["category_summary"]]

    # 担当者別
    user_names: list[str] = [u["name"] for u in summary["user_summary"]]
    user_avg_progress: list[float] = [u["avg_progress"] for u in summary["user_summary"]]
    user_counts: list[int] = [u["count"] for u in summary["user_summary"]]
    user_completed: list[int] = [u["completed"] for u in summary["user_summary"]]
    user_delayed: list[int] = [u["delayed"] for u in summary["user_summary"]]

    return {
        "status_labels": status_labels,
        "status_counts": status_counts,
        "status_colors": status_colors,
        "cat_names": cat_names,
        "cat_avg_progress": cat_avg_progress,
        "cat_counts": cat_counts,
        "cat_completed": cat_completed,
        "cat_delayed": cat_delayed,
        "user_names": user_names,
        "user_avg_progress": user_avg_progress,
        "user_counts": user_counts,
        "user_completed": user_completed,
        "user_delayed": user_delayed,
    }


@project_tasks_bp.route("/dashboard")
def progress_dashboard() -> str:
    """進捗ダッシュボード画面を表示する。

    セッションの権限に応じて、自分または他ユーザーのタスク進捗を
    グラフ付きで閲覧できる。

    Returns:
        str: レンダリングされたHTMLテンプレート。
    """
    login_role: str = session.get("user_role", "")
    login_id: int = int(session["user_id"])
    login_dept: str = session.get("user_dept", "")
    privileged: bool = is_privileged(login_role)

    target_user_id, target_user_name, selectable_users = _resolve_dashboard_target(
        login_id, login_role, login_dept,
    )

    summary: dict = get_task_progress_summary(target_user_id)
    chart_json: dict = _build_chart_json(summary)

    # 管理職・マスタは全体ステータスも同時表示
    overview_summary: dict | None = None
    overview_chart_json: dict | None = None
    if privileged:
        overview_summary = get_task_overview_summary()
        overview_chart_json = _build_overview_chart_json(overview_summary)

    return render_template(
        "project_tasks_dashboard.html",
        summary=summary,
        chart_json=chart_json,
        privileged=privileged,
        selectable_users=selectable_users,
        selected_user_id=target_user_id,
        selected_user_name=target_user_name,
        overview_summary=overview_summary,
        overview_chart_json=overview_chart_json,
    )


@project_tasks_bp.route("/dashboard/api")
def progress_dashboard_api() -> tuple:
    """ユーザー切替時にダッシュボードデータをJSONで返すAPI。

    クエリパラメータ user_id で対象ユーザーを指定する。
    権限チェックは progress_dashboard() と同一ロジック。

    Returns:
        tuple: (Response, status_code) JSON形式のレスポンス。
    """
    login_role: str = session.get("user_role", "")
    login_id: int = int(session["user_id"])
    login_dept: str = session.get("user_dept", "")
    privileged: bool = is_privileged(login_role)

    target_user_id, target_user_name, selectable_users = _resolve_dashboard_target(
        login_id, login_role, login_dept,
    )

    summary: dict = get_task_progress_summary(target_user_id)
    chart_json: dict = _build_chart_json(summary)

    return jsonify({
        "summary": summary,
        "chart_json": chart_json,
        "privileged": privileged,
        "selectable_users": selectable_users,
        "selected_user_id": target_user_id,
        "selected_user_name": target_user_name,
    }), 200


@project_tasks_bp.route("/gantt/update-dates/<int:task_id>", methods=["POST"])
def gantt_update_dates(task_id: int) -> tuple:
    """ガントチャートのドラッグ操作で開始日・終了日を更新するAPI。

    管理職・マスタのみ使用可能。JSON形式で start_date / end_date を受け取る。

    Args:
        task_id: 更新対象のタスクID。

    Returns:
        tuple: (Response, status_code) JSON形式のレスポンス。
    """
    # JSON APIのCSRFチェック（X-CSRF-Token ヘッダー）
    csrf = request.headers.get("X-CSRF-Token", "")
    if csrf != session.get("csrf_token"):
        abort(400)

    # 管理職・マスタのみ使用可能（docstring通りの権限制限）。
    if not is_privileged(session.get("user_role", "")):
        return jsonify({"error": "権限がありません"}), 403

    existing = get_project_task_by_id(task_id)
    if not existing:
        abort(404)
    if not _can_touch_gantt_task(existing):
        return jsonify({"error": "権限がありません"}), 403

    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "JSONデータが必要です"}), 400

    start_date: str = data.get("start_date", "").strip()
    end_date: str = data.get("end_date", "").strip()

    if not start_date or not end_date:
        return jsonify({"error": "開始日・終了日は必須です"}), 400

    from datetime import datetime
    try:
        new_start = datetime.strptime(start_date, "%Y-%m-%d").date()
        new_end = datetime.strptime(end_date, "%Y-%m-%d").date()
    except ValueError:
        return jsonify({"error": "日付形式が不正です（YYYY-MM-DD）"}), 400

    # 元の日付との差分を計算（後続タスク連動用）
    try:
        old_end = date.fromisoformat(existing["end_date"])
    except (ValueError, TypeError):
        old_end = new_end
    shift_days: int = (new_end - old_end).days

    update_project_task(
        task_id=task_id,
        category_id=existing.get("category_id"),
        subcategory_id=existing.get("subcategory_id"),
        task_name=existing["task_name"],
        description=existing.get("description", ""),
        start_date=start_date,
        end_date=end_date,
        status=existing["status"],
        progress=existing.get("progress", 0),
        delay_days=existing.get("delay_days", 0),
        updated_by=session.get("user_name", ""),
        assigned_to=existing.get("assigned_to"),
        assigned_to_2=existing.get("assigned_to_2"),
        is_milestone=existing.get("is_milestone", 0),
    )

    # 後続タスク連動: cascade_ids が指定されていれば同じ日数分ずらす
    cascaded: list[dict] = []
    cascade_ids: list[int] = data.get("cascade_ids", [])
    if shift_days != 0 and cascade_ids:
        for cid in cascade_ids:
            ct = get_project_task_by_id(int(cid))
            if not ct:
                continue
            if not _can_touch_gantt_task(ct):
                continue
            try:
                cs = date.fromisoformat(ct["start_date"])
                ce = date.fromisoformat(ct["end_date"])
            except (ValueError, TypeError):
                continue
            ns = (cs + timedelta(days=shift_days)).isoformat()
            ne = (ce + timedelta(days=shift_days)).isoformat()
            update_project_task(
                task_id=int(cid),
                category_id=ct.get("category_id"),
                subcategory_id=ct.get("subcategory_id"),
                task_name=ct["task_name"],
                description=ct.get("description", ""),
                start_date=ns, end_date=ne,
                status=ct["status"],
                progress=ct.get("progress", 0),
                delay_days=ct.get("delay_days", 0),
                updated_by=session.get("user_name", ""),
                assigned_to=ct.get("assigned_to"),
                assigned_to_2=ct.get("assigned_to_2"),
                is_milestone=ct.get("is_milestone", 0),
            )
            cascaded.append({"id": int(cid), "start_date": ns, "end_date": ne})

    return jsonify({"ok": True, "start_date": start_date, "end_date": end_date, "cascaded": cascaded}), 200


@project_tasks_bp.route("/gantt/update-fields/<int:task_id>", methods=["POST"])
def gantt_update_fields(task_id: int) -> tuple:
    """ガントチャートから状態・進捗・遅延日を更新するAPI。

    全ユーザー使用可能。JSON形式で status / progress / delay_days を受け取る。

    Args:
        task_id: 更新対象のタスクID。

    Returns:
        tuple: (Response, status_code) JSON形式のレスポンス。
    """
    csrf = request.headers.get("X-CSRF-Token", "")
    if csrf != session.get("csrf_token"):
        abort(400)

    existing = get_project_task_by_id(task_id)
    if not existing:
        abort(404)
    if not _can_touch_gantt_task(existing):
        return jsonify({"error": "権限がありません"}), 403

    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "JSONデータが必要です"}), 400

    status: str = data.get("status", existing["status"])
    if status not in PROJECT_TASK_STATUSES:
        status = existing["status"]

    try:
        progress: int = max(0, min(int(data.get("progress", existing.get("progress", 0))), 100))
    except (ValueError, TypeError):
        progress = existing.get("progress", 0) or 0

    try:
        delay_days: int = max(0, int(data.get("delay_days", existing.get("delay_days", 0))))
    except (ValueError, TypeError):
        delay_days = existing.get("delay_days", 0) or 0

    update_project_task(
        task_id=task_id,
        category_id=existing.get("category_id"),
        subcategory_id=existing.get("subcategory_id"),
        task_name=existing["task_name"],
        description=existing.get("description", ""),
        start_date=existing["start_date"],
        end_date=existing["end_date"],
        status=status,
        progress=progress,
        delay_days=delay_days,
        updated_by=session.get("user_name", ""),
        assigned_to=existing.get("assigned_to"),
        assigned_to_2=existing.get("assigned_to_2"),
        is_milestone=existing.get("is_milestone", 0),
    )

    return jsonify({
        "ok": True, "status": status,
        "progress": progress, "delay_days": delay_days,
    }), 200


@project_tasks_bp.route("/gantt/reorder", methods=["POST"])
def gantt_reorder() -> tuple:
    """ガントチャート上でタスクの表示順を入れ替えるAPI。

    JSON: { "order": [id1, id2, ...] } — 対象タスクの表示順を更新する。
    対象タスクが持つ既存 display_order 値の集合を、新しい順序で各 ID に再割当することで、
    対象外のタスクとの相対順序を壊さずに並び替えを実現する。

    Returns:
        tuple: (Response, status_code)
    """
    import logging as _logging
    _logger = _logging.getLogger(__name__)

    try:
        csrf = request.headers.get("X-CSRF-Token", "")
        if csrf != session.get("csrf_token"):
            return jsonify({"error": "CSRFトークン不一致"}), 400

        data = request.get_json(silent=True)
        if not data or "order" not in data:
            return jsonify({"error": "orderが必要です"}), 400

        order_list = data["order"]
        if not order_list:
            return jsonify({"ok": True, "updated": 0}), 200

        try:
            ids: list[int] = [int(x) for x in order_list]
        except (TypeError, ValueError):
            return jsonify({"error": "不正なID形式"}), 400

        for tid in ids:
            t = get_project_task_by_id(tid)
            if t and not _can_touch_gantt_task(t):
                return jsonify({"error": "権限がありません"}), 403

        from ..database import get_db
        db = get_db()
        placeholders = ",".join("?" * len(ids))
        rows = db.execute(
            f"SELECT id, display_order FROM project_task WHERE id IN ({placeholders})",
            tuple(ids),
        ).fetchall()
        if not rows:
            return jsonify({"error": "対象タスクが見つかりません"}), 400

        existing_orders: list[int] = sorted(int(r["display_order"] or 0) for r in rows)
        updated_by: str = session.get("user_name", "") or ""
        updated = 0
        for idx, task_id in enumerate(ids):
            if idx >= len(existing_orders):
                break
            db.execute(
                "UPDATE project_task SET display_order = ?, "
                "  updated_at=datetime('now','localtime'), updated_by=? "
                "WHERE id = ?",
                (existing_orders[idx], updated_by, task_id),
            )
            updated += 1
        db.commit()
        return jsonify({"ok": True, "updated": updated}), 200
    except Exception as e:
        # 例外詳細をログに残し、クライアントにもメッセージを返してデバッグしやすくする
        _logger.exception("gantt_reorder で例外発生")
        return jsonify({"error": f"内部エラー: {type(e).__name__}: {str(e)}"}), 500


@project_tasks_bp.route("/gantt")
def gantt() -> str:
    """ガントチャート画面を表示する。

    全権限で閲覧可能。一般ユーザーは自分のタスクのみ表示。

    Returns:
        str: レンダリング済みHTML
    """
    login_role = session.get("user_role", "")
    login_id = int(session["user_id"])
    login_dept = session.get("user_dept", "")
    privileged = is_privileged(login_role)

    # タスク管理画面から引き継がれる担当者絞り込み（マスタ・管理職のみ有効）
    target_user_id_str: str = request.args.get("user_id", "").strip()
    target_user_id_int: int | None = None
    if target_user_id_str and privileged:
        try:
            tid = int(target_user_id_str)
            if tid > 0:
                target_user_id_int = tid
        except ValueError:
            target_user_id_int = None

    # タスク一覧と同じスコープで取得（管理職・所属長・システム管理者は
    # get_accessible_users のスコープ内メンバーのタスクのみ表示する）。
    if target_user_id_int is not None:
        # 担当者絞り込みあり: 指定ユーザーが担当のタスクのみ
        tasks = get_all_project_tasks(assigned_to=target_user_id_int)
    elif privileged:
        accessible = get_accessible_users(login_id, login_role, login_dept)
        accessible_ids = [u["id"] for u in accessible]
        tasks = get_all_project_tasks(user_ids=accessible_ids)
    else:
        tasks = get_all_project_tasks(assigned_to=login_id)

    # テンプレートに渡すJSON用データ
    gantt_data = []
    for t in tasks:
        # 担当者表示（姓のみ、2名対応）
        names = []
        ln1 = t.get("assigned_last_name") or t.get("assigned_name") or ""
        ln2 = t.get("assigned_last_name_2") or t.get("assigned_name_2") or ""
        if ln1:
            names.append(ln1)
        if ln2:
            names.append(ln2)
        is_event = t.get("is_event", 0)
        is_ms = 1 if t.get("is_milestone", 0) else 0
        # イベント（MSなし）は完全除外
        if is_event and not is_ms:
            continue
        # マイルストーン（イベント+MS または 通常タスク+MS）はイベント行のみ
        if is_ms:
            gantt_data.append({
                "id": t["id"], "name": t["task_name"], "assigned": "",
                "category": "", "subcategory": "",
                "start": t["start_date"], "end": t["end_date"],
                "progress": 0, "status": t["status"],
                "delay_days": 0, "is_milestone": 1,
                "milestone_only": True,
            })
            continue
        gantt_data.append({
            "id": t["id"],
            "name": t["task_name"],
            "assigned": "・".join(names),
            "category": t.get("category_name") or "",
            "subcategory": t.get("subcategory_name") or "",
            "start": t["start_date"],
            "end": t["end_date"],
            "progress": t.get("progress", 0),
            "status": t["status"],
            "delay_days": t.get("delay_days", 0) or 0,
            "is_milestone": 0,
        })

    # タスク管理画面から引き継いだフィルタ初期値（クエリパラメータ）
    initial_subcat: str = request.args.get("subcat_filter", "").strip()
    initial_period_start: str = request.args.get("period_start", "").strip()
    initial_period_end: str = request.args.get("period_end", "").strip()

    # ガントチャート画面のプルダウン用：表示対象タスクの中区分名（重複なし・ソート済）
    subcat_options: list[str] = sorted({
        (t.get("subcategory_name") or "")
        for t in tasks
        if not t.get("is_event", 0) and (t.get("subcategory_name") or "")
    })

    # 管理職・所属長・システム管理者は、画面上で担当者を切り替えられるよう選択肢を渡す
    # （新ガント画面 planner.py と同じ方針：is_master 限定ではなく privileged 全体に対応）。
    is_master_flag: bool = is_master(login_role)
    selectable_users: list[dict] = []
    if privileged:
        selectable_users = get_accessible_users(login_id, login_role, login_dept)

    # ガントチャート画面でのタスク追加モーダル用データ
    add_categories = get_all_categories()
    add_subcategories = get_all_subcategories()
    add_users = get_all_users()

    return render_template(
        "project_tasks_gantt.html",
        gantt_json=json.dumps(gantt_data, ensure_ascii=False),
        privileged=privileged,
        csrf_token=session.get("csrf_token", ""),
        subcat_options=subcat_options,
        initial_subcat=initial_subcat,
        initial_period_start=initial_period_start,
        initial_period_end=initial_period_end,
        initial_target_user_id=target_user_id_int or 0,
        is_master_flag=is_master_flag,
        selectable_users=selectable_users,
        add_categories=add_categories,
        add_subcategories=add_subcategories,
        add_users=add_users,
        statuses=PROJECT_TASK_STATUSES,
    )


# ---------------------------------------------------------------------------
# ガントチャート Excel エクスポート
# ---------------------------------------------------------------------------

# ステータスに対応するExcelセル色（RRGGBB）
_STATUS_FILL: dict[str, str] = {
    "未着手": "D1D5DB",
    "着手":   "93C5FD",
    "順調":   "6EE7B7",
    "遅れ":   "FCA5A5",
    "完了":   "1E40AF",
    "停止":   "E5E7EB",
}

# 進捗バーの色
_PROGRESS_FILL: dict[str, str] = {
    "遅れ": "EF4444",
    "_default": "059669",
}


def _get_monday(d: date) -> date:
    """指定日が属する週の月曜日を返す。"""
    return d - timedelta(days=d.weekday())


def _build_gantt_excel(
    tasks: list[dict],
    start_date: date,
    display_days: int,
    show_completed: bool = False,
    login_dept: str = "",
) -> openpyxl.Workbook:
    """ガントチャート付きExcelワークブックを生成する。

    左側に大項目・タスク名・担当・状態・進捗を配置し、
    右側に日付ごとのセル塗りつぶしでガントバーを描画する。

    Args:
        tasks: タスク一覧（get_all_project_tasks の戻り値）
        start_date: 表示開始日
        display_days: 表示日数
        login_dept: ヘッダーに表示するログインユーザーの所属名（空なら所属名なし）

    Returns:
        openpyxl.Workbook: 生成済みワークブック
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ガントチャート"

    # -- スタイル定義 --
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    header_fill = PatternFill("solid", fgColor="334155")
    header_font = Font(bold=True, color="FFFFFF", size=9, name="游ゴシック")
    cat_fill = PatternFill("solid", fgColor="E2E8F0")
    cat_font = Font(bold=True, size=10, name="游ゴシック")
    body_font = Font(size=9, name="游ゴシック")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    sun_fill = PatternFill("solid", fgColor="FEF2F2")
    sat_fill = PatternFill("solid", fgColor="EFF6FF")
    today_fill = PatternFill("solid", fgColor="FEE2E2")
    milestone_fill = PatternFill("solid", fgColor="C4B5FD")

    # 固定列: A=大項目, B=タスク名, C=担当, D=状態, E=進捗%
    fixed_cols = 5
    col_widths = {"A": 14, "B": 24, "C": 10, "D": 8, "E": 7}

    # -- ヘッダー行1: 固定列名 + 日付 --
    headers = ["大項目", "タスク名", "担当", "状態", "進捗%"]
    weekday_ja = ["月", "火", "水", "木", "金", "土", "日"]

    # 行1: 日付ヘッダー (月/日)
    # 行2: 曜日ヘッダー
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        # 行2もヘッダー背景
        cell2 = ws.cell(2, ci, "")
        cell2.fill = header_fill
        cell2.border = thin_border

    today_d = date.today()
    today_side = Side(style="medium", color="EF4444")
    today_border_top = Border(left=today_side, right=today_side, top=today_side,
                              bottom=Side(style="thin", color="CCCCCC"))
    today_border_mid = Border(left=today_side, right=today_side,
                              top=Side(style="thin", color="CCCCCC"),
                              bottom=Side(style="thin", color="CCCCCC"))
    today_border_bot = Border(left=today_side, right=today_side,
                              top=Side(style="thin", color="CCCCCC"), bottom=today_side)
    today_fill_header = PatternFill("solid", fgColor="FEE2E2")
    today_fill_body = PatternFill("solid", fgColor="FEF2F2")
    today_col_idx: int = -1  # 今日の列インデックス（後で罫線適用用）

    for di in range(display_days):
        dt = start_date + timedelta(days=di)
        col = fixed_cols + di + 1
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 3.5
        is_today = (dt == today_d)

        # 行1: 月/日
        cell1 = ws.cell(1, col, f"{dt.month}/{dt.day}")
        cell1.font = Font(size=7, name="游ゴシック", bold=True,
                          color="EF4444" if dt.weekday() == 6 else
                          "3B82F6" if dt.weekday() == 5 else "FFFFFF")
        cell1.fill = today_fill_header if is_today else header_fill
        cell1.alignment = center_align
        cell1.border = today_border_top if is_today else thin_border
        if is_today:
            today_col_idx = col

        # 行2: 曜日
        cell2 = ws.cell(2, col, weekday_ja[dt.weekday()])
        cell2.font = Font(size=7, name="游ゴシック",
                          color="EF4444" if dt.weekday() == 6 else
                          "3B82F6" if dt.weekday() == 5 else "333333")
        cell2.fill = today_fill_header if is_today else PatternFill("solid", fgColor="F1F5F9")
        cell2.alignment = center_align
        cell2.border = today_border_mid if is_today else thin_border

    # 列幅
    for letter, w in col_widths.items():
        ws.column_dimensions[letter].width = w

    # -- イベント(MS)とタスクを分離 --
    ms_tasks: list[dict] = []
    normal_tasks: list[dict] = []
    completed_fill = PatternFill("solid", fgColor="F0F0F0")
    completed_font = Font(size=9, name="游ゴシック", color="999999")
    for t in tasks:
        is_event = t.get("is_event", 0)
        is_ms = t.get("is_milestone", 0)
        if is_event and not is_ms:
            continue
        if is_event and is_ms:
            ms_tasks.append(t)
        elif is_ms:
            ms_tasks.append(t)
        else:
            if not show_completed and t.get("status") == "完了":
                continue
            normal_tasks.append(t)

    # -- イベント行（行3）: マイルストーンを日付セルに表示 --
    row = 3
    top_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    if ms_tasks:
        ms_font = Font(size=7, name="游ゴシック", color="6D28D9", bold=True)
        ms_fill_bg = PatternFill("solid", fgColor="F5F3FF")
        ws.row_dimensions[row].height = 90
        # イベント行ヘッダー
        ws.cell(row, 1, "📅 イベント").font = Font(size=8, name="游ゴシック", bold=True, color="6D28D9")
        ws.cell(row, 1).fill = ms_fill_bg
        ws.cell(row, 1).alignment = top_align
        ws.cell(row, 1).border = thin_border
        for ci in range(2, fixed_cols + 1):
            ws.cell(row, ci).fill = ms_fill_bg
            ws.cell(row, ci).border = thin_border
        # 日付セルにMS名を記入
        for di in range(display_days):
            dt = start_date + timedelta(days=di)
            col = fixed_cols + di + 1
            cell = ws.cell(row, col)
            is_today = (dt == today_d)
            cell.fill = today_fill_body if is_today else ms_fill_bg
            cell.border = today_border_mid if is_today else thin_border
            # この日に該当するMS
            ms_names: list[str] = []
            for mt in ms_tasks:
                try:
                    ms_d = date.fromisoformat(mt["start_date"])
                except (ValueError, TypeError):
                    continue
                if ms_d == dt:
                    ms_names.append(mt["task_name"])
            if ms_names:
                cell.value = "◆" + "／".join(ms_names)
                cell.font = ms_font
                cell.alignment = top_align
                cell.fill = milestone_fill
        row += 1

    # -- 大区分→中区分→タスクの3階層にグループ化（画面のガントチャート構造に合わせる） --
    cat_order: list[str] = []
    cat_map: dict[str, dict] = {}
    # cat_map[cat] = {"subcat_order": [...], "subcat_map": {subcat: [task...]}, "total": int}
    for t in normal_tasks:
        cat = t.get("category_name") or "（未分類）"
        sub = t.get("subcategory_name") or "（その他）"
        if cat not in cat_map:
            cat_map[cat] = {"subcat_order": [], "subcat_map": {}, "total": 0}
            cat_order.append(cat)
        cm = cat_map[cat]
        if sub not in cm["subcat_map"]:
            cm["subcat_map"][sub] = []
            cm["subcat_order"].append(sub)
        cm["subcat_map"][sub].append(t)
        cm["total"] += 1

    # 中区分行スタイル（大区分よりやや薄く）
    subcat_fill = PatternFill("solid", fgColor="F1F5F9")
    subcat_font = Font(bold=True, size=9, name="游ゴシック", color="475569")

    for cat in cat_order:
        cm = cat_map[cat]

        # 大区分行
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=fixed_cols + display_days)
        cell = ws.cell(row, 1, f"▼ {cat}（{cm['total']}件）")
        cell.fill = cat_fill
        cell.font = cat_font
        cell.alignment = left_align
        cell.border = thin_border
        row += 1

        for sub in cm["subcat_order"]:
            sub_tasks = cm["subcat_map"][sub]

            # 中区分行
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=fixed_cols + display_days)
            scell = ws.cell(row, 1, f"　└ {sub}（{len(sub_tasks)}件）")
            scell.fill = subcat_fill
            scell.font = subcat_font
            scell.alignment = left_align
            scell.border = thin_border
            row += 1

            # タスク行（中区分内）
            for t in sub_tasks:
                # 担当者名
                names = []
                ln1 = t.get("assigned_last_name") or t.get("assigned_name") or ""
                ln2 = t.get("assigned_last_name_2") or t.get("assigned_name_2") or ""
                if ln1:
                    names.append(ln1)
                if ln2:
                    names.append(ln2)
                assigned = "・".join(names)

                progress = t.get("progress", 0) or 0
                status = t.get("status", "")
                is_completed = (status == "完了")
                row_font = completed_font if is_completed else body_font
                row_fill = completed_fill if is_completed else None

                # 固定列
                ws.cell(row, 1, cat).font = row_font
                ws.cell(row, 1).alignment = left_align
                ws.cell(row, 1).border = thin_border
                if row_fill: ws.cell(row, 1).fill = row_fill

                ws.cell(row, 2, t["task_name"]).font = row_font
                ws.cell(row, 2).alignment = left_align
                ws.cell(row, 2).border = thin_border
                if row_fill: ws.cell(row, 2).fill = row_fill

                ws.cell(row, 3, assigned).font = row_font
                ws.cell(row, 3).alignment = center_align
                ws.cell(row, 3).border = thin_border
                if row_fill: ws.cell(row, 3).fill = row_fill

                status_cell = ws.cell(row, 4, status)
                status_cell.font = row_font
                status_cell.alignment = center_align
                status_cell.border = thin_border
                if row_fill: status_cell.fill = row_fill

                prog_cell = ws.cell(row, 5, f"{progress}%")
                prog_cell.font = row_font
                prog_cell.alignment = center_align
                prog_cell.border = thin_border
                if row_fill: prog_cell.fill = row_fill

                # ガントバー描画
                try:
                    t_start = date.fromisoformat(t["start_date"])
                    t_end = date.fromisoformat(t["end_date"])
                except (ValueError, TypeError, KeyError):
                    row += 1
                    continue

                for di in range(display_days):
                    dt = start_date + timedelta(days=di)
                    col = fixed_cols + di + 1
                    cell = ws.cell(row, col)
                    is_today = (dt == today_d)
                    cell.border = today_border_mid if is_today else thin_border

                    # 土日背景
                    if dt.weekday() == 5:
                        cell.fill = sat_fill
                    elif dt.weekday() == 6:
                        cell.fill = sun_fill

                    # 今日ハイライト
                    if is_today:
                        cell.fill = today_fill_body

                    # タスク期間内
                    if t_start <= dt <= t_end:
                        if status == "完了":
                            cell.fill = PatternFill("solid", fgColor="1E40AF")
                        elif status == "停止":
                            cell.fill = PatternFill("solid", fgColor="E5E7EB")
                        else:
                            # 進捗バー計算
                            total_days = (t_end - t_start).days + 1
                            day_idx = (dt - t_start).days
                            prog_days = round(total_days * progress / 100)

                            if day_idx < prog_days:
                                # 進捗済み部分
                                fill_color = _PROGRESS_FILL.get(
                                    status, _PROGRESS_FILL["_default"]
                                )
                                cell.fill = PatternFill("solid", fgColor=fill_color)
                            else:
                                # 予定部分
                                cell.fill = PatternFill("solid", fgColor="93C5FD")

                row += 1

    # 今日の列の最終行に下罫線を付ける
    if today_col_idx > 0:
        last_data_row = row - 1
        c = ws.cell(last_data_row, today_col_idx)
        c.border = today_border_bot

    # ヘッダー+イベント行を固定（イベント行がない場合は行3から）
    freeze_row = 4 if ms_tasks else 3
    ws.freeze_panes = f"F{freeze_row}"

    # -- 凡例行（色セル＋ラベル＋空白列の3列間隔） --
    row += 1
    legend_font = Font(size=8, name="游ゴシック")
    legend_items = [
        ("93C5FD", "予定期間"),
        ("34D399", "進捗（実績）"),
        ("FCA5A5", "遅延"),
        ("1E40AF", "完了"),
        ("E5E7EB", "停止"),
        ("C4B5FD", "◆ マイルストーン"),
    ]
    ws.cell(row, 1, "【凡例】").font = Font(size=8, name="游ゴシック", bold=True)
    ws.cell(row, 1).border = thin_border
    # 3列×2行のレイアウト（1行に3項目）
    items_per_row = 3
    for li, (color, label) in enumerate(legend_items):
        r_offset = li // items_per_row
        c_offset = li % items_per_row
        c = fixed_cols + c_offset * 3 + 1
        r = row + r_offset
        cell_color = ws.cell(r, c)
        cell_color.fill = PatternFill("solid", fgColor=color)
        cell_color.border = thin_border
        cell_label = ws.cell(r, c + 1, label)
        cell_label.font = legend_font
        cell_label.border = thin_border

    # -- 印刷設定: 横1枚に収まるように --
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # 縦は自動
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    # ヘッダー: 左に所属名+業務進捗、右に出力日付
    ws.oddHeader.left.text = f"{login_dept}業務進捗" if login_dept else "業務進捗"
    ws.oddHeader.left.size = 10
    ws.oddHeader.right.text = f"出力日：{today_d.strftime('%Y/%m/%d')}"
    ws.oddHeader.right.size = 9

    return wb


def _build_flat_excel_for_migration(tasks: list[dict]) -> "openpyxl.Workbook":
    """全プロジェクトタスクを移行用フラットExcelとして出力する。

    新ガントチャート画面（planner.py）のインポート機能で読み込む前提の
    フォーマット。親子ツリー（parent_task_id）・メンバー（member_ids）を
    そのまま列として書き出す。

    Args:
        tasks: get_all_project_tasks() の全件（絞り込みなし）。

    Returns:
        openpyxl.Workbook: 生成済みワークブック。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "タスク移行"

    header_fill = PatternFill("solid", fgColor="334155")
    header_font = Font(bold=True, color="FFFFFF", size=10, name="游ゴシック")
    body_font = Font(size=10, name="游ゴシック")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    headers = ["id", "parent_task_id", "task_name", "担当者1", "担当者2",
               "開始日", "終了日", "状態", "進捗%", "メンバー"]
    col_widths = [6, 14, 32, 10, 10, 12, 12, 8, 8, 20]

    user_name_map: dict[int, str] = {
        u["id"]: (u.get("last_name") or u.get("name") or "").strip() for u in get_all_users()
    }

    def _member_names(raw: str) -> str:
        ids = [int(s) for s in (raw or "").split(",") if s.strip().isdigit()]
        return ",".join(user_name_map.get(i, "") for i in ids if user_name_map.get(i))

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(ci)].width = col_widths[ci - 1]

    for ri, t in enumerate(tasks, 2):
        values = [
            t["id"],
            t.get("parent_task_id") or "",
            t.get("task_name", ""),
            user_name_map.get(t.get("assigned_to"), ""),
            user_name_map.get(t.get("assigned_to_2"), ""),
            t.get("start_date", ""),
            t.get("end_date", ""),
            t.get("status", ""),
            t.get("progress", 0),
            _member_names(t.get("member_ids", "")),
        ]
        for ci, v in enumerate(values, 1):
            cell = ws.cell(ri, ci, v)
            cell.font = body_font
            cell.border = thin_border
            cell.alignment = center_align if ci in (1, 2, 9) else left_align

    return wb


def _build_gantt_excel_by_tree(
    visible_roots: list[dict],
    children: dict[int, list[dict]],
    start_date: date,
    display_days: int,
    login_dept: str = "",
) -> "openpyxl.Workbook":
    """親子ツリー（ガントチャート／親=大項目・子=タスク）のExcelを生成する。

    大区分/中区分ではなく、ガントチャート画面と同じ親タスク（大項目）に
    紐づく子タスクの構造でグルーピングして出力する。

    Args:
        visible_roots: 表示対象のルート（親）タスク一覧。
        children: 親タスクID → 子タスク一覧のマップ。
        start_date: 表示開始日。
        display_days: 表示日数。
        login_dept: ヘッダーに表示するログインユーザーの所属名（空なら所属名なし）。

    Returns:
        openpyxl.Workbook: 生成済みワークブック。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ガントチャート"

    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC"),
    )
    header_fill = PatternFill("solid", fgColor="334155")
    header_font = Font(bold=True, color="FFFFFF", size=9, name="游ゴシック")
    cat_fill = PatternFill("solid", fgColor="E2E8F0")
    cat_font = Font(bold=True, size=10, name="游ゴシック")
    body_font = Font(size=9, name="游ゴシック")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    sun_fill = PatternFill("solid", fgColor="FEF2F2")
    sat_fill = PatternFill("solid", fgColor="EFF6FF")
    today_fill_body = PatternFill("solid", fgColor="FEF2F2")
    today_fill_header = PatternFill("solid", fgColor="FEE2E2")
    completed_fill = PatternFill("solid", fgColor="F0F0F0")
    completed_font = Font(size=9, name="游ゴシック", color="999999")

    fixed_cols = 5
    col_widths = {"A": 18, "B": 26, "C": 10, "D": 8, "E": 7}
    headers = ["大項目", "タスク名", "担当", "状態", "進捗%"]
    weekday_ja = ["月", "火", "水", "木", "金", "土", "日"]
    today_d = date.today()
    today_side = Side(style="medium", color="EF4444")
    today_border_top = Border(left=today_side, right=today_side, top=today_side,
                              bottom=Side(style="thin", color="CCCCCC"))
    today_border_mid = Border(left=today_side, right=today_side,
                              top=Side(style="thin", color="CCCCCC"),
                              bottom=Side(style="thin", color="CCCCCC"))
    today_border_bot = Border(left=today_side, right=today_side,
                              top=Side(style="thin", color="CCCCCC"), bottom=today_side)
    today_col_idx = -1

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = center_align; cell.border = thin_border
        ws.cell(2, ci, "").fill = header_fill
        ws.cell(2, ci).border = thin_border

    for di in range(display_days):
        dt = start_date + timedelta(days=di)
        col = fixed_cols + di + 1
        ws.column_dimensions[get_column_letter(col)].width = 3.5
        is_today = (dt == today_d)
        c1 = ws.cell(1, col, f"{dt.month}/{dt.day}")
        c1.font = Font(size=7, name="游ゴシック", bold=True,
                       color="EF4444" if dt.weekday() == 6 else "3B82F6" if dt.weekday() == 5 else "FFFFFF")
        c1.fill = today_fill_header if is_today else header_fill
        c1.alignment = center_align
        c1.border = today_border_top if is_today else thin_border
        if is_today:
            today_col_idx = col
        c2 = ws.cell(2, col, weekday_ja[dt.weekday()])
        c2.font = Font(size=7, name="游ゴシック",
                       color="EF4444" if dt.weekday() == 6 else "3B82F6" if dt.weekday() == 5 else "333333")
        c2.fill = today_fill_header if is_today else PatternFill("solid", fgColor="F1F5F9")
        c2.alignment = center_align
        c2.border = today_border_mid if is_today else thin_border

    for letter, w in col_widths.items():
        ws.column_dimensions[letter].width = w

    def _sortkey(t: dict) -> tuple:
        return (t.get("display_order") or 0, (t.get("start_date") or "9999-99-99"), t["id"])

    def _assigned_names(t: dict) -> str:
        names = []
        ln1 = t.get("assigned_last_name") or t.get("assigned_name") or ""
        ln2 = t.get("assigned_last_name_2") or t.get("assigned_name_2") or ""
        if ln1:
            names.append(ln1)
        if ln2:
            names.append(ln2)
        return "・".join(names)

    row = 3

    def _write_task_row(t: dict, root_name: str, indent: int) -> None:
        nonlocal row
        status = t.get("status", "")
        progress = t.get("progress", 0) or 0
        is_completed = (status == "完了")
        row_font = completed_font if is_completed else body_font
        row_fill = completed_fill if is_completed else None

        ws.cell(row, 1, root_name).font = row_font
        ws.cell(row, 1).alignment = left_align
        ws.cell(row, 1).border = thin_border
        if row_fill: ws.cell(row, 1).fill = row_fill

        name_cell = ws.cell(row, 2, "　" * indent + t.get("task_name", ""))
        name_cell.font = row_font; name_cell.alignment = left_align; name_cell.border = thin_border
        if row_fill: name_cell.fill = row_fill

        ws.cell(row, 3, _assigned_names(t)).font = row_font
        ws.cell(row, 3).alignment = center_align
        ws.cell(row, 3).border = thin_border
        if row_fill: ws.cell(row, 3).fill = row_fill

        ws.cell(row, 4, status).font = row_font
        ws.cell(row, 4).alignment = center_align
        ws.cell(row, 4).border = thin_border
        if row_fill: ws.cell(row, 4).fill = row_fill

        ws.cell(row, 5, f"{progress}%").font = row_font
        ws.cell(row, 5).alignment = center_align
        ws.cell(row, 5).border = thin_border
        if row_fill: ws.cell(row, 5).fill = row_fill

        try:
            t_start = date.fromisoformat(t["start_date"])
            t_end = date.fromisoformat(t["end_date"])
        except (ValueError, TypeError, KeyError):
            row += 1
            return

        for di in range(display_days):
            dt = start_date + timedelta(days=di)
            col = fixed_cols + di + 1
            cell = ws.cell(row, col)
            is_today = (dt == today_d)
            cell.border = today_border_mid if is_today else thin_border
            if dt.weekday() == 5:
                cell.fill = sat_fill
            elif dt.weekday() == 6:
                cell.fill = sun_fill
            if is_today:
                cell.fill = today_fill_body
            if t_start <= dt <= t_end:
                if status == "完了":
                    cell.fill = PatternFill("solid", fgColor="1E40AF")
                elif status == "停止":
                    cell.fill = PatternFill("solid", fgColor="E5E7EB")
                else:
                    total_days = (t_end - t_start).days + 1
                    day_idx = (dt - t_start).days
                    prog_days = round(total_days * progress / 100)
                    if day_idx < prog_days:
                        fill_color = _PROGRESS_FILL.get(status, _PROGRESS_FILL["_default"])
                        cell.fill = PatternFill("solid", fgColor=fill_color)
                    else:
                        cell.fill = PatternFill("solid", fgColor="93C5FD")
        row += 1

    def _count_descendants(t: dict) -> int:
        total = 0
        for c in children.get(t["id"], []):
            total += 1 + _count_descendants(c)
        return total

    def _write_descendants(t: dict, root_name: str, indent: int) -> None:
        for c in sorted(children.get(t["id"], []), key=_sortkey):
            _write_task_row(c, root_name, indent)
            _write_descendants(c, root_name, indent + 1)

    for r in sorted(visible_roots, key=_sortkey):
        desc_count = _count_descendants(r)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=fixed_cols + display_days)
        cell = ws.cell(row, 1, f"▼ {r.get('task_name', '')}（{desc_count}件）")
        cell.fill = cat_fill; cell.font = cat_font; cell.alignment = left_align; cell.border = thin_border
        row += 1
        if desc_count == 0:
            # 子がない場合は親タスク自身を1行として出力する（期間データを失わないため）。
            _write_task_row(r, r.get("task_name", ""), 0)
        else:
            _write_descendants(r, r.get("task_name", ""), 0)

    if today_col_idx > 0:
        last_data_row = row - 1
        ws.cell(last_data_row, today_col_idx).border = today_border_bot

    ws.freeze_panes = "F3"

    row += 1
    legend_font = Font(size=8, name="游ゴシック")
    legend_items = [
        ("93C5FD", "予定期間"), ("34D399", "進捗（実績）"), ("FCA5A5", "遅延"),
        ("1E40AF", "完了"), ("E5E7EB", "停止"),
    ]
    ws.cell(row, 1, "【凡例】").font = Font(size=8, name="游ゴシック", bold=True)
    ws.cell(row, 1).border = thin_border
    items_per_row = 3
    for li, (color, label) in enumerate(legend_items):
        r_offset = li // items_per_row
        c_offset = li % items_per_row
        c = fixed_cols + c_offset * 3 + 1
        r = row + r_offset
        ws.cell(r, c).fill = PatternFill("solid", fgColor=color)
        ws.cell(r, c).border = thin_border
        lbl = ws.cell(r, c + 1, label)
        lbl.font = legend_font; lbl.border = thin_border

    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.oddHeader.left.text = f"{login_dept}業務進捗" if login_dept else "業務進捗"
    ws.oddHeader.left.size = 10
    ws.oddHeader.right.text = f"出力日：{today_d.strftime('%Y/%m/%d')}"
    ws.oddHeader.right.size = 9

    return wb


def _build_gantt_pdf_by_tree(
    visible_roots: list[dict],
    children: dict[int, list[dict]],
    start_date: date,
    display_days: int,
    login_dept: str = "",
) -> "io.BytesIO":
    """親子ツリーのガントチャートをPDF（A4横）で生成する。

    _build_gantt_excel_by_tree と同じ構成（固定列＝大項目/タスク名/担当/状態/進捗% ＋
    日付グリッド＋進捗バー）を reportlab で描画する。Excelをそのまま紙にした見た目。

    Args:
        visible_roots: 表示対象のルート（親）タスク一覧。
        children: 親タスクID → 子タスク一覧のマップ。
        start_date: 表示開始日。
        display_days: 表示日数。
        login_dept: ヘッダーに表示するログインユーザーの所属名（空なら所属名なし）。

    Returns:
        io.BytesIO: 生成済みPDFのバイナリ。
    """
    import io as _io
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.pdfgen import canvas as _canvas
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont

    # 日本語CIDフォント（追加ファイル不要）
    try:
        pdfmetrics.registerFont(UnicodeCIDFont("HeiseiKakuGo-W5"))
        jp_font = "HeiseiKakuGo-W5"
    except Exception:
        jp_font = "Helvetica"

    def _rgb(hexcode: str):
        return (int(hexcode[0:2], 16) / 255, int(hexcode[2:4], 16) / 255, int(hexcode[4:6], 16) / 255)

    page_w, page_h = landscape(A4)
    margin = 14
    today_d = date.today()
    weekday_ja = ["月", "火", "水", "木", "金", "土", "日"]

    # 列幅（pt）。固定列＋日付列。Excelの相対幅に合わせる。
    fixed = [("大項目", 78), ("タスク名", 150), ("担当", 54), ("状態", 40), ("進捗%", 32)]
    fixed_w = sum(w for _, w in fixed)
    grid_w = page_w - margin * 2 - fixed_w
    day_w = grid_w / display_days
    row_h = 15
    header_h = 26

    # 行データを平坦化（親カテゴリ行＋タスク行）
    def _sortkey(t: dict) -> tuple:
        return (t.get("display_order") or 0, (t.get("start_date") or "9999-99-99"), t["id"])

    def _assigned(t: dict) -> str:
        n = []
        a1 = t.get("assigned_last_name") or t.get("assigned_name") or ""
        a2 = t.get("assigned_last_name_2") or t.get("assigned_name_2") or ""
        if a1: n.append(a1)
        if a2: n.append(a2)
        return "・".join(n)

    def _count_desc(t: dict) -> int:
        return sum(1 + _count_desc(c) for c in children.get(t["id"], []))

    rows: list[dict] = []  # {type:'cat'|'task', ...}
    def _emit_desc(t: dict, root_name: str, indent: int) -> None:
        for c in sorted(children.get(t["id"], []), key=_sortkey):
            rows.append({"type": "task", "t": c, "root": root_name, "indent": indent})
            _emit_desc(c, root_name, indent + 1)
    for r in sorted(visible_roots, key=_sortkey):
        dc = _count_desc(r)
        rows.append({"type": "cat", "name": f"▼ {r.get('task_name','')}（{dc}件）"})
        if dc == 0:
            rows.append({"type": "task", "t": r, "root": r.get("task_name", ""), "indent": 0})
        else:
            _emit_desc(r, r.get("task_name", ""), 0)

    buf = _io.BytesIO()
    c = _canvas.Canvas(buf, pagesize=landscape(A4))
    rows_per_page = int((page_h - margin * 2 - header_h) // row_h)

    def _draw_header(y_top: float) -> None:
        x = margin
        c.setFont(jp_font, 7)
        # 固定列ヘッダー
        c.setFillColorRGB(*_rgb("334155"))
        c.rect(margin, y_top - header_h, fixed_w, header_h, fill=1, stroke=0)
        c.setFillColorRGB(1, 1, 1)
        for label, w in fixed:
            c.drawCentredString(x + w / 2, y_top - header_h / 2 - 3, label)
            x += w
        # 日付列ヘッダー（月/日＋曜日）
        for di in range(display_days):
            dt = start_date + timedelta(days=di)
            cx = margin + fixed_w + di * day_w
            is_today = (dt == today_d)
            c.setFillColorRGB(*(_rgb("FEE2E2") if is_today else _rgb("334155")))
            c.rect(cx, y_top - header_h, day_w, header_h, fill=1, stroke=0)
            wd = dt.weekday()
            col = "EF4444" if wd == 6 else ("3B82F6" if wd == 5 else ("333333" if is_today else "FFFFFF"))
            c.setFillColorRGB(*_rgb(col))
            c.setFont(jp_font, 5.5)
            c.drawCentredString(cx + day_w / 2, y_top - 10, f"{dt.month}/{dt.day}")
            c.drawCentredString(cx + day_w / 2, y_top - 20, weekday_ja[wd])
        # 枠線
        c.setStrokeColorRGB(*_rgb("CCCCCC")); c.setLineWidth(0.4)
        c.rect(margin, y_top - header_h, page_w - margin * 2, header_h, fill=0, stroke=1)

    def _draw_row(item: dict, y: float) -> None:
        if item["type"] == "cat":
            c.setFillColorRGB(*_rgb("E2E8F0"))
            c.rect(margin, y, page_w - margin * 2, row_h, fill=1, stroke=0)
            c.setFillColorRGB(0.1, 0.13, 0.2); c.setFont(jp_font, 8)
            c.drawString(margin + 3, y + 4, item["name"][:60])
            c.setStrokeColorRGB(*_rgb("CCCCCC")); c.setLineWidth(0.4)
            c.rect(margin, y, page_w - margin * 2, row_h, fill=0, stroke=1)
            return
        t = item["t"]
        status = t.get("status", ""); progress = t.get("progress", 0) or 0
        completed = (status == "完了")
        # 固定列テキスト
        c.setFont(jp_font, 7)
        c.setFillColorRGB(*(_rgb("999999") if completed else (0.1, 0.13, 0.2)))
        vals = [item["root"], "  " * item["indent"] + t.get("task_name", ""),
                _assigned(t), status, f"{progress}%"]
        x = margin
        for (label, w), v in zip(fixed, vals):
            if label in ("担当", "状態", "進捗%"):
                c.drawCentredString(x + w / 2, y + 4, str(v)[:10])
            else:
                c.drawString(x + 2, y + 4, str(v)[:24])
            x += w
        # 日付セル（土日・当日・バー）
        try:
            ts = date.fromisoformat(t["start_date"]); te = date.fromisoformat(t["end_date"])
        except (ValueError, TypeError, KeyError):
            ts = te = None
        for di in range(display_days):
            dt = start_date + timedelta(days=di)
            cx = margin + fixed_w + di * day_w
            wd = dt.weekday()
            bg = None
            if wd == 5: bg = "EFF6FF"
            elif wd == 6: bg = "FEF2F2"
            if dt == today_d: bg = "FEF2F2"
            if bg:
                c.setFillColorRGB(*_rgb(bg)); c.rect(cx, y, day_w, row_h, fill=1, stroke=0)
            if ts and te and ts <= dt <= te:
                if status == "完了":
                    fill = "1E40AF"
                elif status == "停止":
                    fill = "E5E7EB"
                else:
                    total = (te - ts).days + 1
                    idx = (dt - ts).days
                    prog_days = round(total * progress / 100)
                    fill = _PROGRESS_FILL.get(status, _PROGRESS_FILL["_default"]) if idx < prog_days else "93C5FD"
                c.setFillColorRGB(*_rgb(fill))
                c.rect(cx + 0.5, y + 2, day_w - 1, row_h - 4, fill=1, stroke=0)
        # 罫線（行下）
        c.setStrokeColorRGB(*_rgb("E5E7EB")); c.setLineWidth(0.3)
        c.line(margin, y, page_w - margin * 2 + margin, y)

    idx = 0
    while idx < len(rows) or idx == 0:
        y_top = page_h - margin
        # ヘッダー（左：タイトル、右：出力日）
        c.setFont(jp_font, 9); c.setFillColorRGB(0.2, 0.2, 0.2)
        c.drawString(margin, y_top - 2, f"{login_dept}業務進捗" if login_dept else "業務進捗")
        c.drawRightString(page_w - margin, y_top - 2, f"出力日：{today_d.strftime('%Y/%m/%d')}")
        y_top -= 12
        _draw_header(y_top)
        y = y_top - header_h - row_h
        drawn = 0
        while idx < len(rows) and drawn < rows_per_page:
            _draw_row(rows[idx], y)
            y -= row_h; idx += 1; drawn += 1
        # 縦の列区切り線（固定列境界）
        c.setStrokeColorRGB(*_rgb("CCCCCC")); c.setLineWidth(0.4)
        gx = margin
        for _, w in fixed:
            c.line(gx, y_top - header_h, gx, y + row_h)
            gx += w
        c.line(margin + fixed_w, y_top - header_h, margin + fixed_w, y + row_h)
        c.showPage()
        if idx >= len(rows):
            break

    c.save()
    buf.seek(0)
    return buf


@project_tasks_bp.route("/export-migration")
def export_migration() -> object:
    """全タスクを移行用フラットExcelとして出力する（システム管理者のみ）。

    新ガントチャート画面（planner.py）へのデータ移行のために、
    全プロジェクトタスクを親子ツリー情報付きのフラット一覧で出力する。

    Returns:
        object: Excelファイルのダウンロードレスポンス。権限がなければ403。
    """
    if not is_master(session.get("user_role", "")):
        abort(403)

    tasks = get_all_project_tasks()
    wb = _build_flat_excel_for_migration(tasks)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    today_str = date.today().isoformat()
    return send_file(
        buf, as_attachment=True, download_name=f"タスク移行_{today_str}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@project_tasks_bp.route("/gantt/export")
def export_gantt() -> object:
    """ガントチャートをExcelファイルとしてエクスポートする。

    権限別の出力範囲:
      - ユーザー: 自分が担当のタスクのみ
      - 管理職: 自部署メンバーが担当のタスク
      - マスタ: 全タスク

    クエリパラメータ:
      - start: 表示開始日（YYYY-MM-DD）。省略時は前週月曜日。
      - days: 表示日数。省略時は28。

    Returns:
        object: Excelファイルのダウンロードレスポンス
    """
    login_role: str = session.get("user_role", "")
    login_id: int = int(session["user_id"])
    login_dept: str = session.get("user_dept", "")
    privileged: bool = is_privileged(login_role)

    # 表示期間
    start_param: str = request.args.get("start", "")
    days_param: str = request.args.get("days", "28")

    try:
        display_days = max(7, min(90, int(days_param)))
    except ValueError:
        display_days = 28

    if start_param:
        try:
            start_d = date.fromisoformat(start_param)
        except ValueError:
            start_d = _get_monday(date.today()) - timedelta(days=7)
    else:
        start_d = _get_monday(date.today()) - timedelta(days=7)

    # 担当者絞り込み（マスタ・管理職のみ有効）
    target_user_id_str: str = request.args.get("user_id", "").strip()
    target_user_id_int: int | None = None
    if target_user_id_str and privileged:
        try:
            tid = int(target_user_id_str)
            if tid > 0:
                target_user_id_int = tid
        except ValueError:
            target_user_id_int = None

    # 権限別タスク取得
    if target_user_id_int is not None:
        # 担当者絞り込みあり: 指定ユーザーが担当のタスクのみ
        tasks = get_all_project_tasks(assigned_to=target_user_id_int)
    elif is_master(login_role):
        # マスタ: 全タスク
        tasks = get_all_project_tasks()
    elif privileged:
        # 管理職: 自部署メンバーのタスク
        dept_users = get_all_users(dept_filter=login_dept)
        dept_user_ids: set[int] = {u["id"] for u in dept_users}
        all_tasks = get_all_project_tasks()
        tasks = [
            t for t in all_tasks
            if (t.get("assigned_to") in dept_user_ids
                or t.get("assigned_to_2") in dept_user_ids)
        ]
    else:
        # 一般ユーザー: 自分のタスクのみ
        tasks = get_all_project_tasks(assigned_to=login_id)

    show_comp = request.args.get("show_completed") == "1"

    # 中区分絞り込み（ガントチャート画面から引き継ぐ）
    subcat_f: str = request.args.get("subcat_filter", "").strip()
    if subcat_f:
        tasks = [t for t in tasks if (t.get("subcategory_name") or "") == subcat_f]

    # 期間絞り込み（タスクの期間と指定範囲が重なるもの）
    period_start: str = request.args.get("period_start", "").strip()
    period_end: str = request.args.get("period_end", "").strip()
    if period_start or period_end:
        tasks = [
            t for t in tasks
            if (
                (not period_end or (t.get("start_date") or "") <= period_end)
                and (not period_start or (t.get("end_date") or "") >= period_start)
            )
        ]

    wb = _build_gantt_excel(tasks, start_d, display_days, show_completed=show_comp, login_dept=login_dept)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"業務進捗_{date.today().isoformat()}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
