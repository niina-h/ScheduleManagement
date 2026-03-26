"""週間予定のExcelエクスポートルートを提供するBlueprintモジュール。"""
from __future__ import annotations

import io
import logging
from datetime import date, timedelta
from pathlib import Path

import openpyxl
from flask import Blueprint, abort, flash, redirect, request, send_file, session, url_for

logger = logging.getLogger(__name__)
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..models import (
    get_accessible_users,
    get_all_users,
    get_daily_comment,
    get_daily_result,
    get_user_by_id,
    get_weekly_leave,
    get_weekly_schedule,
    save_weekly_schedule,
)
from ..auth_helpers import is_privileged, is_master, can_access_user

export_bp = Blueprint("export_bp", __name__, url_prefix="/export")


def _get_monday(d: date) -> date:
    """指定日が属する週の月曜日を返す。

    Args:
        d (date): 基準日。

    Returns:
        date: その週の月曜日。
    """
    return d - timedelta(days=d.weekday())


def _get_current_week_start() -> str:
    """今週の月曜日をISO形式文字列で返す。

    Returns:
        str: 今週月曜日の 'YYYY-MM-DD' 形式文字列。
    """
    return _get_monday(date.today()).isoformat()


def _append_schedule_sheet(
    wb: openpyxl.Workbook,
    user: dict,
    week_start: str,
    sheet_label: str,
    schedule: dict,
) -> None:
    """既存のWorkbookに週間予定の1シートを追加する。

    午前セクション（水色背景）と午後セクション（黄色背景）に分けて
    各日・各枠の作業名と時間を出力し、小計行・1日合計行で合計を表示する。
    小計が基本勤務時間と一致しない場合は赤字で警告する。

    Args:
        wb: 追加先のWorkbook。
        user: ユーザー情報（name, std_hours を含む）。
        week_start: 週開始日（'YYYY-MM-DD' 形式）。
        sheet_label: シート名のプレフィックス（「前週」「今週」「来週」など）。
        schedule: get_weekly_schedule の返り値。
            {day(0-4): {'am': [{task_name, hours}×5], 'pm': [同]}}
    """
    days = ["月", "火", "水", "木", "金"]
    start = date.fromisoformat(week_start)
    date_strs = [(start + timedelta(days=i)).strftime("%m/%d") for i in range(5)]

    sheet_name = f"{sheet_label}_{week_start}"[:31]
    ws = wb.create_sheet(title=sheet_name)

    # ヘッダー行
    ws.append(["区分", "枠"] + [f"{days[i]}({date_strs[i]})" for i in range(5)])
    for col in range(1, 8):
        cell = ws.cell(1, col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # std_hours が設定されていればそれを優先、なければ am+pm の合計にフォールバック
    std_hours: float = user.get("std_hours") or (user.get("std_hours_am", 4.0) + user.get("std_hours_pm", 4.0))
    std_am: float = std_hours / 2
    std_pm: float = std_hours / 2

    am_fill = PatternFill("solid", fgColor="E8F4FB")
    pm_fill = PatternFill("solid", fgColor="FDF8F0")
    total_fill_am = PatternFill("solid", fgColor="C0D8EC")
    total_fill_pm = PatternFill("solid", fgColor="ECE0C0")
    red_font = Font(color="DC3545", bold=True)
    black_bold = Font(bold=True)

    def write_section(
        section_label: str,
        slot_key: str,
        std_h: float,
        fill: PatternFill,
        total_fill: PatternFill,
    ) -> None:
        """午前または午後の1セクション分を書き出す内部関数。

        Args:
            section_label: セクション名（「午前」または「午後」）。
            slot_key: scheduleのキー（'am' または 'pm'）。
            std_h: そのセクションの基本勤務時間。
            fill: データ行の背景色。
            total_fill: 小計行の背景色。
        """
        for i in range(5):
            row_data = [section_label if i == 0 else "", str(i + 1)]
            for day in range(5):
                slot = schedule.get(day, {}).get(slot_key, [{}] * 5)
                entry = slot[i] if i < len(slot) else {}
                t: str = entry.get("task_name", "")
                h: float = entry.get("hours", 0.0)
                row_data.append(f"{t} ({h}h)" if t else (f"{h}h" if h else ""))
            ws.append(row_data)
            for col in range(1, 8):
                ws.cell(ws.max_row, col).fill = fill

        # 小計行
        totals = ["", "計"]
        for day in range(5):
            slot = schedule.get(day, {}).get(slot_key, [{}] * 5)
            total = sum(e.get("hours", 0.0) for e in slot)
            totals.append(f"{total:.2f}h / {std_h:.1f}h")
        ws.append(totals)

        subtotal_row = ws.max_row
        for col in range(1, 8):
            cell = ws.cell(subtotal_row, col)
            cell.fill = total_fill
            cell.font = black_bold

        # 小計セルの文字色（不一致→赤）
        for day in range(5):
            slot = schedule.get(day, {}).get(slot_key, [{}] * 5)
            total = sum(e.get("hours", 0.0) for e in slot)
            cell = ws.cell(subtotal_row, day + 3)
            cell.font = red_font if abs(total - std_h) > 0.001 else black_bold

    write_section("午前", "am", std_am, am_fill, total_fill_am)
    write_section("午後", "pm", std_pm, pm_fill, total_fill_pm)

    # 1日合計行
    day_total_fill = PatternFill("solid", fgColor="D8ECD8")
    total_row = ["", "1日合計"]
    for day in range(5):
        day_total = sum(
            e.get("hours", 0.0)
            for slot_key in ["am", "pm"]
            for e in schedule.get(day, {}).get(slot_key, [])
        )
        total_row.append(f"{day_total:.2f}h / {std_hours:.1f}h")
    ws.append(total_row)
    day_total_row = ws.max_row
    for col in range(1, 8):
        cell = ws.cell(day_total_row, col)
        cell.fill = day_total_fill
        cell.font = black_bold
    for day in range(5):
        day_total = sum(
            e.get("hours", 0.0)
            for slot_key in ["am", "pm"]
            for e in schedule.get(day, {}).get(slot_key, [])
        )
        cell = ws.cell(day_total_row, day + 3)
        cell.font = red_font if abs(day_total - std_hours) > 0.001 else black_bold

    # 列幅
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 6
    for col in range(3, 8):
        ws.column_dimensions[get_column_letter(col)].width = 24


def _build_team_week_excel(users: list[dict], week_start: str) -> io.BytesIO:
    """全メンバーの今週予定を1シートに縦積みしたExcelを生成する。

    ユーザー管理の登録順にユーザーブロックを縦積みする。
    A列: 氏名、B列: 区分、C列: 枠番号、D-H列: 月〜金のデータ。

    Args:
        users: ユーザー情報のリスト（id, name, std_hours を含む）。
        week_start: 週開始日（'YYYY-MM-DD' 形式）。

    Returns:
        io.BytesIO: 生成されたExcelファイルのバイトストリーム（先頭にシーク済み）。
    """
    days = ["月", "火", "水", "木", "金"]
    start = date.fromisoformat(week_start)
    date_strs = [(start + timedelta(days=i)).strftime("%m/%d") for i in range(5)]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"今週_{week_start}"[:31]

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    name_fill = PatternFill("solid", fgColor="D9E2F3")
    name_font = Font(bold=True)
    am_fill = PatternFill("solid", fgColor="E8F4FB")
    pm_fill = PatternFill("solid", fgColor="FDF8F0")
    subtotal_am_fill = PatternFill("solid", fgColor="C0D8EC")
    subtotal_pm_fill = PatternFill("solid", fgColor="ECE0C0")
    daytotal_fill = PatternFill("solid", fgColor="D8ECD8")
    red_font = Font(color="DC3545", bold=True)
    black_bold = Font(bold=True)
    center_align = Alignment(horizontal="center")

    # ヘッダー行
    header = ["氏名", "区分", "枠"] + [f"{days[i]}({date_strs[i]})" for i in range(5)]
    ws.append(header)
    for col in range(1, len(header) + 1):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for user in users:
        schedule = get_weekly_schedule(user["id"], week_start)
        std_hours: float = user.get("std_hours") or (
            user.get("std_hours_am", 4.0) + user.get("std_hours_pm", 4.0)
        )
        std_am = std_hours / 2
        std_pm = std_hours / 2

        # ユーザー名行
        name_row_idx = ws.max_row + 1
        ws.append([user["name"], "", ""] + [""] * 5)
        for col in range(1, 9):
            cell = ws.cell(name_row_idx, col)
            cell.fill = name_fill
            cell.font = name_font

        def _write_slot_section(
            slot_key: str,
            label: str,
            std_h: float,
            row_fill: PatternFill,
            sub_fill: PatternFill,
        ) -> None:
            """午前または午後の5行＋小計行を書き出す内部関数。"""
            for i in range(5):
                row_data = ["", label if i == 0 else "", str(i + 1)]
                for day in range(5):
                    slot = schedule.get(day, {}).get(slot_key, [{}] * 5)
                    entry = slot[i] if i < len(slot) else {}
                    t: str = entry.get("task_name", "")
                    h: float = entry.get("hours", 0.0)
                    row_data.append(f"{t} ({h}h)" if t else (f"{h}h" if h else ""))
                ws.append(row_data)
                for col in range(1, 9):
                    ws.cell(ws.max_row, col).fill = row_fill

            totals = ["", "", "計"]
            for day in range(5):
                slot = schedule.get(day, {}).get(slot_key, [{}] * 5)
                total = sum(e.get("hours", 0.0) for e in slot)
                totals.append(f"{total:.2f}h / {std_h:.1f}h")
            ws.append(totals)
            sub_row = ws.max_row
            for col in range(1, 9):
                ws.cell(sub_row, col).fill = sub_fill
            for day in range(5):
                slot = schedule.get(day, {}).get(slot_key, [{}] * 5)
                total = sum(e.get("hours", 0.0) for e in slot)
                ws.cell(sub_row, day + 4).font = (
                    red_font if abs(total - std_h) > 0.001 else black_bold
                )

        _write_slot_section("am", "午前", std_am, am_fill, subtotal_am_fill)
        _write_slot_section("pm", "午後", std_pm, pm_fill, subtotal_pm_fill)

        # 1日合計行
        total_row: list = ["", "", "1日合計"]
        for day in range(5):
            day_total = sum(
                e.get("hours", 0.0)
                for slot_key in ["am", "pm"]
                for e in schedule.get(day, {}).get(slot_key, [])
            )
            total_row.append(f"{day_total:.2f}h / {std_hours:.1f}h")
        ws.append(total_row)
        dt_row = ws.max_row
        for col in range(1, 9):
            ws.cell(dt_row, col).fill = daytotal_fill
        for day in range(5):
            day_total = sum(
                e.get("hours", 0.0)
                for slot_key in ["am", "pm"]
                for e in schedule.get(day, {}).get(slot_key, [])
            )
            ws.cell(dt_row, day + 4).font = (
                red_font if abs(day_total - std_hours) > 0.001 else black_bold
            )

        # ユーザー間の空行
        ws.append([""] * 8)

    # 列幅
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 4
    for col_idx in range(4, 9):
        ws.column_dimensions[get_column_letter(col_idx)].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_team_week_excel_v2(
    users: list[dict], week_start: str, dept_name: str
) -> io.BytesIO:
    """サンプルフォーマットに準拠した全メンバー週間スケジュール表を生成する。

    1シート（部署名）に全ユーザーを縦積みし、各ユーザーブロックは
    AM5行 + PM5行 + 合計稼働時間行（SUM式）の計11行で構成する。

    列構造:
        A=氏名, B=予定/実績, C=午前/午後
        曜日ごとに D/G/J/M/P=タスク名, E/H/K/N/Q=日名, F/I/L/O/R=時間(h)

    Args:
        users: ユーザー情報リスト（id, name を含む）。
        week_start: 週開始日（'YYYY-MM-DD' 形式）。
        dept_name: 部署名（シート名に使用）。

    Returns:
        io.BytesIO: 生成されたExcelファイルのバイトストリーム（先頭にシーク済み）。
    """
    start = date.fromisoformat(week_start)
    day_labels = ["月", "火", "水", "木", "金"]
    date_strs = [(start + timedelta(days=i)).strftime("%m/%d") for i in range(5)]
    year_month = start.strftime("%Y年%m月")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = dept_name[:31]

    # スタイル定義
    title_font = Font(bold=True, size=12)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center")
    am_fill = PatternFill("solid", fgColor="E8F4FB")
    pm_fill = PatternFill("solid", fgColor="FDF8F0")
    sum_fill = PatternFill("solid", fgColor="D8ECD8")
    name_font = Font(bold=True)

    def _task_col(day: int) -> int:
        """曜日インデックス(0-4)からタスク名列番号を返す。"""
        return 4 + day * 3

    def _time_col(day: int) -> int:
        """曜日インデックス(0-4)から時間列番号を返す。"""
        return 6 + day * 3

    # 行1: タイトル
    ws.cell(1, 4, year_month).font = title_font
    ws.cell(1, 7, "週間スケジュール表").font = title_font

    # 行2: ヘッダー
    ws.cell(2, 1, "氏名")
    for day in range(5):
        ws.cell(2, _task_col(day), date_strs[day])
        ws.cell(2, _task_col(day) + 1, f"日（{day_labels[day]}）")
        ws.cell(2, _time_col(day), "時間(ｈ)")
    for col in range(1, 19):
        cell = ws.cell(2, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # 行3: 空行（スペーサー）
    current_row = 3

    for user in users:
        schedule = get_weekly_schedule(user["id"], week_start)
        leave_data = get_weekly_leave(user["id"], week_start)
        data_start_row = current_row + 1  # AM1行目の行番号

        for slot_idx, (slot_key, slot_label) in enumerate([("am", "午前"), ("pm", "午後")]):
            fill = am_fill if slot_key == "am" else pm_fill
            for row_i in range(5):
                current_row += 1
                # A・B・C 列
                if slot_idx == 0 and row_i == 0:
                    ws.cell(current_row, 1, user["name"]).font = name_font
                    ws.cell(current_row, 2, "予定/実績")
                if row_i == 0:
                    ws.cell(current_row, 3, slot_label)
                # 各曜日のタスク・時間
                for day in range(5):
                    leave = leave_data.get(day, "")
                    if leave:
                        # 休暇は1行目のタスク列に種別を表示
                        if row_i == 0:
                            ws.cell(current_row, _task_col(day), leave)
                    else:
                        slot_data = schedule.get(day, {}).get(slot_key, [])
                        entry = slot_data[row_i] if row_i < len(slot_data) else {}
                        task: str = entry.get("task_name") or ""
                        hours: float = entry.get("hours") or 0.0
                        if task:
                            ws.cell(current_row, _task_col(day), task)
                        if hours:
                            ws.cell(current_row, _time_col(day), hours)
                # 背景色
                for col in range(1, 19):
                    ws.cell(current_row, col).fill = fill

        # 合計稼働時間行（SUM式）
        current_row += 1
        for day in range(5):
            tc = _time_col(day)
            col_letter = get_column_letter(tc)
            ws.cell(current_row, _task_col(day), "合計稼働時間").fill = sum_fill
            ws.cell(
                current_row,
                tc,
                f"=SUM({col_letter}{data_start_row}:{col_letter}{current_row - 1})",
            ).fill = sum_fill

    # 列幅
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 5
    for day in range(5):
        ws.column_dimensions[get_column_letter(_task_col(day))].width = 22
        ws.column_dimensions[get_column_letter(_task_col(day) + 1)].width = 9
        ws.column_dimensions[get_column_letter(_time_col(day))].width = 8

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_team_week_from_schedule_tpl(
    users: list[dict], week_start: str
) -> io.BytesIO:
    """週間予定表テンプレートを使用して全メンバーの週間予定を1シートに生成する。

    reports/tpl/週間予定表_テンプレート.xlsx の行4-13をユーザー数分複製し、
    各ユーザーの予定データを書き込む。

    Args:
        users: ユーザー情報リスト（id, name を含む）。
        week_start: 週開始日（'YYYY-MM-DD' 形式）。

    Returns:
        io.BytesIO: 生成されたExcelファイルのバイトストリーム（先頭にシーク済み）。

    Raises:
        FileNotFoundError: テンプレートファイルが存在しない場合。
    """
    from copy import copy
    from openpyxl.styles import Font as _Fnt

    wb = openpyxl.load_workbook(str(_SCHEDULE_TPL_PATH))
    ws = wb["ベース"]

    start = date.fromisoformat(week_start)
    BLOCK = 10  # 1ユーザー = 10行（AM5行 + PM5行）
    MAX_COL = 20

    # --- テンプレートブロック（行4-13）の書式を記録 ---
    tpl_styles: dict[tuple[int, int], dict] = {}
    for r in range(4, 14):
        for c in range(1, MAX_COL + 1):
            src = ws.cell(r, c)
            tpl_styles[(r - 4, c)] = {
                "font": copy(src.font),
                "fill": copy(src.fill),
                "border": copy(src.border),
                "alignment": copy(src.alignment),
                "number_format": src.number_format,
            }

    tpl_heights: dict[int, float | None] = {}
    for r in range(4, 14):
        tpl_heights[r - 4] = ws.row_dimensions[r].height

    # テンプレート内のセル結合を記録（行4-13範囲）
    tpl_merges: list[tuple[int, int, int, int]] = []
    for mc in list(ws.merged_cells.ranges):
        if mc.min_row >= 4 and mc.max_row <= 13:
            tpl_merges.append((mc.min_row - 4, mc.max_row - 4, mc.min_col, mc.max_col))
            ws.unmerge_cells(str(mc))

    # 行14以降に残る結合もクリア
    for mc in list(ws.merged_cells.ranges):
        if mc.min_row >= 14:
            ws.unmerge_cells(str(mc))

    # --- ヘッダーの日付を更新 ---
    ws.cell(1, 4).value = start.strftime("%Y年%m月")
    date_col_letters = ["D", "G", "J", "M", "P"]
    for i, col_l in enumerate(date_col_letters):
        d = start + timedelta(days=i)
        ws[f"{col_l}2"] = d.strftime("%m/%d")

    # --- テンプレート行4以降をクリア ---
    for r in range(4, max(ws.max_row + 1, 17)):
        for c in range(1, MAX_COL + 1):
            ws.cell(r, c).value = None

    # --- 休暇表示用フォント ---
    red_font = _Fnt(color="FF0000", bold=True)
    FULL_DAY_LEAVES = {"1日有休", "特休", "祝日", "その他休み"}

    # --- ユーザーごとにブロックを作成 ---
    for idx, user in enumerate(users):
        base_row = 4 + idx * BLOCK

        # 書式を適用
        for r_off in range(BLOCK):
            row = base_row + r_off
            h = tpl_heights.get(r_off)
            if h is not None:
                ws.row_dimensions[row].height = h
            for c in range(1, MAX_COL + 1):
                cell = ws.cell(row, c)
                style = tpl_styles.get((r_off, c))
                if style:
                    cell.font = style["font"]
                    cell.fill = style["fill"]
                    cell.border = style["border"]
                    cell.alignment = style["alignment"]
                    cell.number_format = style["number_format"]

        # セル結合を適用
        for rs, re, cs, ce in tpl_merges:
            ws.merge_cells(
                start_row=base_row + rs,
                end_row=base_row + re,
                start_column=cs,
                end_column=ce,
            )

        # 固定ラベルを書き込み
        ws.cell(base_row, 1).value = user["name"]
        ws.cell(base_row, 2).value = "予定/実績"
        ws.cell(base_row, 3).value = "午前"
        ws.cell(base_row + 5, 3).value = "午後"

        # スケジュールデータを書き込み
        schedule = get_weekly_schedule(user["id"], week_start)
        leave_data = get_weekly_leave(user["id"], week_start)

        for day, (task_col, hours_col) in enumerate(_DAY_COLS):
            leave = leave_data.get(day, "")

            # ---- 午前セクション（ブロック行0-4） ----
            am_blocked = leave in FULL_DAY_LEAVES or leave == "AM半休"
            if leave in FULL_DAY_LEAVES:
                c = ws.cell(base_row, task_col)
                c.value = leave
                c.font = red_font
            elif leave == "AM半休":
                c = ws.cell(base_row, task_col)
                c.value = "AM有休"
                c.font = red_font

            if not am_blocked:
                plan_am = schedule.get(day, {}).get("am", [])
                for i in range(5):
                    entry = plan_am[i] if i < len(plan_am) else {}
                    task_name = entry.get("task_name") or ""
                    hours = entry.get("hours") or 0
                    if task_name:
                        ws.cell(base_row + i, task_col).value = task_name
                    if hours:
                        ws.cell(base_row + i, hours_col).value = float(hours)

            # ---- 午後セクション（ブロック行5-9） ----
            pm_blocked = leave in FULL_DAY_LEAVES or leave == "PM半休"
            if leave == "PM半休":
                c = ws.cell(base_row + 5, task_col)
                c.value = "PM有休"
                c.font = red_font

            if not pm_blocked:
                plan_pm = schedule.get(day, {}).get("pm", [])
                for i in range(5):
                    entry = plan_pm[i] if i < len(plan_pm) else {}
                    task_name = entry.get("task_name") or ""
                    hours = entry.get("hours") or 0
                    if task_name:
                        ws.cell(base_row + 5 + i, task_col).value = task_name
                    if hours:
                        ws.cell(base_row + 5 + i, hours_col).value = float(hours)

    ws.title = "週間予定"

    # 印刷範囲を最終データ行・最終列まで自動設定
    last_row = 3 + len(users) * BLOCK
    last_col_letter = get_column_letter(ws.max_column)
    ws.print_area = f"A1:{last_col_letter}{last_row}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_schedule_excel(user: dict, week_start: str, schedule: dict) -> io.BytesIO:
    """週間予定をExcelファイルとして生成しBytesIOで返す（後方互換用）。

    内部で _append_schedule_sheet を呼び出し、単一シートのワークブックを生成する。

    Args:
        user: ユーザー情報（name, std_hours を含む）。
        week_start: 週開始日（'YYYY-MM-DD' 形式）。
        schedule: get_weekly_schedule の返り値。

    Returns:
        io.BytesIO: 生成されたExcelファイルのバイトストリーム（先頭にシーク済み）。
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # デフォルトシートを削除
    _append_schedule_sheet(wb, user, week_start, "週間予定", schedule)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# テンプレートファイルのパス
_TPL_PATH = Path(__file__).parents[2] / "reports" / "tpl" / "週間予定実績表 .xlsx"
_SCHEDULE_TPL_PATH = Path(__file__).parents[2] / "reports" / "tpl" / "週間予定表_テンプレート.xlsx"

# 日別の（タスク列番号, 時間列番号）マッピング
# 月: D(4)merged, F(6) / 火: G(7)merged, I(9) / 水: J(10)merged, L(12)
# 木: M(13)merged, O(15) / 金: P(16)merged, R(18)
_DAY_COLS: list[tuple[int, int]] = [(4, 6), (7, 9), (10, 12), (13, 15), (16, 18)]


def _fill_schedule_tpl_sheet(
    ws,
    week_start: str,
    schedule: dict,
    leave_data: dict | None = None,
    results: dict | None = None,
    comments: dict | None = None,
) -> None:
    """テンプレートのコピーシートに週間予定・実績・休暇・コメントを書き込む。

    Args:
        ws: 書き込み先のワークシート（ベースシートのコピー）。
        week_start: 週開始日（'YYYY-MM-DD' 形式）。
        schedule: get_weekly_schedule の返り値。
        leave_data: {day(0-4): leave_type_str} の休暇設定。
        results: {day(0-4): {'am': [...], 'pm': [...], 'has_result': bool}}。
        comments: {day(0-4): {'reflection': str, 'action': str}}。
    """
    from openpyxl.styles import PatternFill as _PF, Font as _Fnt

    start = date.fromisoformat(week_start)
    leave_data = leave_data or {}
    results = results or {}
    comments = comments or {}

    # 実績背景色（実績入力がある日）
    result_fill = _PF("solid", fgColor="ECF6F6")
    # 赤文字フォント（休暇名表示）
    red_font = _Fnt(color="FF0000", bold=True)

    # 全日有休扱い（AM+PM両方を空にする）
    FULL_DAY_LEAVES = {"1日有休", "特休", "祝日", "その他休み"}

    # D1 に YYYY年MM月を記載（月曜日の年月）
    ws["D1"] = start.strftime("%Y年%m月")

    # 日付セルを mm/dd 形式で上書き（D2/G2/J2/M2/P2）
    date_col_letters = ["D", "G", "J", "M", "P"]
    for i, col_letter in enumerate(date_col_letters):
        d = start + timedelta(days=i)
        ws[f"{col_letter}2"] = d.strftime("%m/%d")

    for day, (task_col, hours_col) in enumerate(_DAY_COLS):
        leave = leave_data.get(day, "")
        has_result = results.get(day, {}).get("has_result", False)
        day_result = results.get(day, {})

        # 実績がある場合はタスク行(4-18)の背景色を変更
        if has_result:
            for row in range(4, 19):
                ws.cell(row=row, column=task_col).fill = result_fill
                ws.cell(row=row, column=hours_col).fill = result_fill

        # ---- タスク分類: リスケ・繰越 ----
        reschedule_tasks: list[tuple[str, float]] = []  # (task_name, negative_hours)
        carryover_tasks: list[tuple[str, float]] = []   # (task_name, hours)

        # ---- 午前セクション (rows 4-8) ----
        am_blocked = leave in FULL_DAY_LEAVES or leave == "AM半休"
        if leave in FULL_DAY_LEAVES:
            c = ws.cell(row=4, column=task_col)
            c.value = leave
            c.font = red_font
        elif leave == "AM半休":
            c = ws.cell(row=4, column=task_col)
            c.value = "AM有休"
            c.font = red_font

        if not am_blocked:
            plan_am = schedule.get(day, {}).get("am", [])
            result_am = day_result.get("am", []) if has_result else []
            # 予定タスク名セット（スロット位置に依存しない突発判定用）
            plan_am_tasks: set[str] = {
                (e.get("task_name") or "").strip()
                for e in plan_am if (e.get("task_name") or "").strip()
            }
            row_offset = 0
            for i in range(5):
                p = plan_am[i] if i < len(plan_am) else {}
                r = result_am[i] if i < len(result_am) else {}
                plan_task = p.get("task_name") or ""
                plan_h = p.get("hours") or 0
                result_task = r.get("task_name") or ""
                result_h = r.get("hours") or 0
                result_defer = r.get("defer_date") or ""
                if has_result:
                    if result_defer and result_task:
                        # リスケ: rows 14-18 へ（負数時間）
                        neg_h = result_h if result_h <= 0 else -(plan_h or result_h)
                        reschedule_tasks.append((result_task, neg_h))
                    elif result_task and result_task.strip() not in plan_am_tasks:
                        # 突発: AM全体の予定に存在しないタスク
                        row = 4 + row_offset
                        ws.cell(row=row, column=task_col).value = f"【突発】{result_task}"
                        ws.cell(row=row, column=hours_col).value = float(result_h or 0)
                        row_offset += 1
                    elif plan_task and not result_task and not result_h:
                        # 繰越
                        carryover_tasks.append((plan_task, float(plan_h or 0)))
                    elif result_task and float(result_h or 0) > 0:
                        # 通常実績 → rows 4-8 に詰めて表示
                        row = 4 + row_offset
                        ws.cell(row=row, column=task_col).value = result_task
                        ws.cell(row=row, column=hours_col).value = float(result_h)
                        row_offset += 1
                else:
                    if plan_task:
                        ws.cell(row=4 + i, column=task_col).value = plan_task
                        if plan_h:
                            ws.cell(row=4 + i, column=hours_col).value = float(plan_h)

        # ---- 午後セクション (rows 9-13) ----
        pm_blocked = leave in FULL_DAY_LEAVES or leave == "PM半休"
        if leave == "PM半休":
            c = ws.cell(row=9, column=task_col)
            c.value = "PM有休"
            c.font = red_font

        if not pm_blocked:
            plan_pm = schedule.get(day, {}).get("pm", [])
            result_pm = day_result.get("pm", []) if has_result else []
            # 予定タスク名セット（スロット位置に依存しない突発判定用）
            plan_pm_tasks: set[str] = {
                (e.get("task_name") or "").strip()
                for e in plan_pm if (e.get("task_name") or "").strip()
            }
            row_offset = 0
            for i in range(5):
                p = plan_pm[i] if i < len(plan_pm) else {}
                r = result_pm[i] if i < len(result_pm) else {}
                plan_task = p.get("task_name") or ""
                plan_h = p.get("hours") or 0
                result_task = r.get("task_name") or ""
                result_h = r.get("hours") or 0
                result_defer = r.get("defer_date") or ""
                if has_result:
                    if result_defer and result_task:
                        neg_h = result_h if result_h <= 0 else -(plan_h or result_h)
                        reschedule_tasks.append((result_task, neg_h))
                    elif result_task and result_task.strip() not in plan_pm_tasks:
                        # 突発: PM全体の予定に存在しないタスク
                        row = 9 + row_offset
                        ws.cell(row=row, column=task_col).value = f"【突発】{result_task}"
                        ws.cell(row=row, column=hours_col).value = float(result_h or 0)
                        row_offset += 1
                    elif plan_task and not result_task and not result_h:
                        carryover_tasks.append((plan_task, float(plan_h or 0)))
                    elif result_task and float(result_h or 0) > 0:
                        row = 9 + row_offset
                        ws.cell(row=row, column=task_col).value = result_task
                        ws.cell(row=row, column=hours_col).value = float(result_h)
                        row_offset += 1
                else:
                    if plan_task:
                        ws.cell(row=9 + i, column=task_col).value = plan_task
                        if plan_h:
                            ws.cell(row=9 + i, column=hours_col).value = float(plan_h)

        # ---- リスケ (rows 14-18) ----
        for j, (t, h) in enumerate(reschedule_tasks[:5]):
            row = 14 + j
            ws.cell(row=row, column=task_col).value = f"リスケ: {t}"
            if h:
                ws.cell(row=row, column=hours_col).value = h

        # ---- 繰越業務 (rows 20-24) ----
        for j, (t_name, t_hours) in enumerate(carryover_tasks[:5]):
            row = 20 + j
            ws.cell(row=row, column=task_col).value = t_name
            if t_hours:
                ws.cell(row=row, column=hours_col).value = t_hours

        # ---- コメント ----
        comment = comments.get(day, {})
        reflection = comment.get("reflection") or ""
        action = comment.get("action") or ""
        if reflection:
            ws.cell(row=25, column=task_col).value = reflection
        if action:
            ws.cell(row=29, column=task_col).value = action


def _build_schedule_excel_from_tpl(user: dict, week_start: str, mode: str = "result") -> io.BytesIO:
    """テンプレートを使用して前週・今週・翌週の3シートExcelを生成する。

    reports/tpl/週間予定実績表 .xlsx の「ベース」シートをコピーして
    前週・今週・翌週の3シートを作成し、各週の予定データを書き込む。

    Args:
        user: ユーザー情報（id, name を含む）。
        week_start: 今週月曜日の日付文字列（'YYYY-MM-DD' 形式）。
        mode: 出力モード。'result' なら実績データを使用、'plan' なら予定のみ出力。

    Returns:
        io.BytesIO: 生成されたExcelファイルのバイトストリーム。
    """
    wb = openpyxl.load_workbook(str(_TPL_PATH))
    tpl_ws = wb["ベース"]
    tpl_zoom: int = tpl_ws.sheet_view.zoomScale or 100  # テンプレートの表示倍率

    ws_date = date.fromisoformat(week_start)
    weeks = [
        ("前週", (ws_date - timedelta(weeks=1)).isoformat()),
        ("今週", week_start),
        ("翌週", (ws_date + timedelta(weeks=1)).isoformat()),
    ]

    for sheet_label, wk_start in weeks:
        schedule = get_weekly_schedule(user["id"], wk_start)
        leave_data = get_weekly_leave(user["id"], wk_start)

        # 各曜日の日次実績とコメントを取得
        wk_date = date.fromisoformat(wk_start)
        day_results: dict[int, dict] = {}
        day_comments: dict[int, dict] = {}
        for day_idx in range(5):
            d_str = (wk_date + timedelta(days=day_idx)).isoformat()
            dr = get_daily_result(user["id"], d_str)
            has_result = any(
                entry.get("task_name") or (entry.get("hours") or 0) > 0
                for slot in dr.values()
                for entry in slot
            )
            # modeが'plan'の場合は実績データを使わない
            day_results[day_idx] = {**dr, "has_result": has_result if mode == "result" else False}
            day_comments[day_idx] = get_daily_comment(user["id"], d_str) or {}

        ws = wb.copy_worksheet(tpl_ws)
        ws.title = sheet_label
        ws.sheet_view.zoomScale = tpl_zoom  # テンプレートの表示倍率を明示的に設定
        _fill_schedule_tpl_sheet(
            ws, wk_start, schedule,
            leave_data=leave_data,
            results=day_results,
            comments=day_comments,
        )

    # ベースシートを削除
    wb.remove(tpl_ws)

    # 「今週」シートを開いたときのアクティブシートに設定
    if "今週" in wb.sheetnames:
        wb.active = wb["今週"]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_team_week_tpl(users: list[dict], week_start: str) -> io.BytesIO:
    """テンプレート形式でメンバー全員の週間予定を複数シートのExcelに生成する。

    個人ダウンロードと同じテンプレート形式で、1ユーザー = 1シート。
    実績・コメントは含まず、予定データ（午前・午後行）のみを出力する。

    Args:
        users: ユーザー情報のリスト（id, name を含む）。
        week_start: 週開始日（'YYYY-MM-DD' 形式）。

    Returns:
        io.BytesIO: 生成されたExcelファイルのバイトストリーム（先頭にシーク済み）。

    Raises:
        FileNotFoundError: テンプレートファイルが存在しない場合。
    """
    wb = openpyxl.load_workbook(str(_TPL_PATH))
    tpl_ws = wb["ベース"]
    tpl_zoom: int = tpl_ws.sheet_view.zoomScale or 100

    used_titles: list[str] = []

    for user in users:
        schedule = get_weekly_schedule(user["id"], week_start)
        leave_data = get_weekly_leave(user["id"], week_start)

        # シート名：重複時は連番付与
        base_title = user["name"][:31]
        sheet_title = base_title
        n = 2
        while sheet_title in used_titles:
            sheet_title = f"{base_title[:28]}({n})"
            n += 1
        used_titles.append(sheet_title)

        ws = wb.copy_worksheet(tpl_ws)
        ws.title = sheet_title
        ws.sheet_view.zoomScale = tpl_zoom
        _fill_schedule_tpl_sheet(
            ws,
            week_start,
            schedule,
            leave_data=leave_data,
            results={},    # 予定のみ（実績なし）
            comments={},
        )

    wb.remove(tpl_ws)

    if wb.worksheets:
        wb.active = wb.worksheets[0]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _append_daily_sheet(
    wb: openpyxl.Workbook,
    user: dict,
    date_str: str,
    schedule_am: list,
    schedule_pm: list,
    result: dict,
    comment: dict,
    next_am: list | None = None,
    next_pm: list | None = None,
    next_date_str: str = "",
) -> None:
    """Workbookに日報シートを1枚追加する。

    左列に予定、右列に実績を並べたレイアウトで出力する。
    シート末尾に振り返り・対策コメントを付加する。
    next_am/next_pm が指定された場合は翌日予定セクションも追加する。

    Args:
        wb: 追加先のWorkbook。
        user: ユーザー情報（name, std_hours を含む）。
        date_str: 日付（'YYYY-MM-DD' 形式）。
        schedule_am: 予定の午前5枠リスト。
        schedule_pm: 予定の午後5枠リスト。
        result: 日次実績 {'am': [...], 'pm': [...]}。
        comment: 日次コメント {'reflection': str, 'action': str}。
        next_am: 翌日予定の午前5枠（省略可）。
        next_pm: 翌日予定の午後5枠（省略可）。
        next_date_str: 翌日の日付文字列（省略可）。
    """
    sheet_name = f"{user['name']}"[:31]
    # 同名シートが既に存在する場合は連番を付ける
    existing = [s.title for s in wb.worksheets]
    if sheet_name in existing:
        for n in range(2, 100):
            candidate = f"{sheet_name}({n})"[:31]
            if candidate not in existing:
                sheet_name = candidate
                break

    ws = wb.create_sheet(title=sheet_name)

    # スタイル定義
    am_fill = PatternFill("solid", fgColor="E8F4FB")
    pm_fill = PatternFill("solid", fgColor="FDF8F0")
    total_fill = PatternFill("solid", fgColor="D8ECD8")
    comment_fill = PatternFill("solid", fgColor="F0E8F8")
    header_fill = PatternFill("solid", fgColor="4472C4")
    next_fill = PatternFill("solid", fgColor="E8F0E8")
    bold = Font(bold=True)
    white_bold = Font(bold=True, color="FFFFFF")
    red_bold = Font(bold=True, color="DC3545")
    center = Alignment(horizontal="center")
    wrap = Alignment(wrap_text=True, vertical="top")

    std_hours: float = user.get("std_hours") or 8.0

    # タイトル行
    ws.append([f"日報: {date_str}　氏名: {user['name']}"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, size=13)
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=7)

    ws.append([])  # 空行

    # ヘッダー行
    ws.append(["区分", "枠", "予定タスク", "予定(h)", "実績タスク", "実績(h)", "差異(h)"])
    header_row = ws.max_row
    for col in range(1, 8):
        cell = ws.cell(header_row, col)
        cell.fill = header_fill
        cell.font = white_bold
        cell.alignment = center

    # セクション出力内部関数
    def write_rows(slot_label: str, sched_list: list, result_list: list, fill: PatternFill) -> None:
        """午前または午後の5行を書き出す。"""
        plan_total = 0.0
        result_total = 0.0
        for i in range(5):
            sc = sched_list[i] if i < len(sched_list) else {"task_name": "", "hours": 0.0}
            rs = result_list[i] if i < len(result_list) else {"task_name": "", "hours": 0.0}
            ph = sc.get("hours") or 0.0
            rh = rs.get("hours") or 0.0
            diff = rh - ph
            plan_total += ph
            result_total += rh
            row_data = [
                slot_label if i == 0 else "",
                str(i + 1),
                sc.get("task_name", ""),
                ph if ph else "",
                rs.get("task_name", ""),
                rh if rh else "",
                f"{diff:+.2f}" if (ph or rh) else "",
            ]
            ws.append(row_data)
            cur = ws.max_row
            for col in range(1, 8):
                ws.cell(cur, col).fill = fill
        # 小計行
        ws.append(["", "計", "", plan_total, "", result_total, ""])
        sub_row = ws.max_row
        for col in range(1, 8):
            ws.cell(sub_row, col).fill = PatternFill("solid", fgColor="C8D8E8" if fill == am_fill else "E8D8A8")
            ws.cell(sub_row, col).font = bold
        ws.cell(sub_row, 4).font = bold
        ws.cell(sub_row, 6).font = bold

    write_rows("午前", schedule_am, result.get("am", []), am_fill)
    write_rows("午後", schedule_pm, result.get("pm", []), pm_fill)

    # 1日合計行
    plan_day = sum((e.get("hours") or 0.0) for e in schedule_am + schedule_pm)
    result_day = sum(
        (e.get("hours") or 0.0)
        for slot in ("am", "pm")
        for e in result.get(slot, [])
    )
    ws.append(["", "1日合計", "", plan_day, "", result_day, ""])
    total_row = ws.max_row
    for col in range(1, 8):
        ws.cell(total_row, col).fill = total_fill
        ws.cell(total_row, col).font = bold
    if abs(result_day - std_hours) > 0.001:
        ws.cell(total_row, 6).font = red_bold

    ws.append([])  # 空行

    # コメントセクション
    ws.append(["本日の振り返り"])
    ws.cell(ws.max_row, 1).font = bold
    ws.cell(ws.max_row, 1).fill = comment_fill

    reflection_text = comment.get("reflection", "") if comment else ""
    ws.append([reflection_text])
    ref_row = ws.max_row
    ws.cell(ref_row, 1).alignment = wrap
    ws.merge_cells(start_row=ref_row, start_column=1, end_row=ref_row, end_column=7)
    ws.row_dimensions[ref_row].height = max(40, len(reflection_text) // 30 * 15 + 20)

    ws.append(["今後の対策・懸念事項"])
    ws.cell(ws.max_row, 1).font = bold
    ws.cell(ws.max_row, 1).fill = comment_fill

    action_text = comment.get("action", "") if comment else ""
    ws.append([action_text])
    act_row = ws.max_row
    ws.cell(act_row, 1).alignment = wrap
    ws.merge_cells(start_row=act_row, start_column=1, end_row=act_row, end_column=7)
    ws.row_dimensions[act_row].height = max(40, len(action_text) // 30 * 15 + 20)

    # 翌日予定セクション（指定がある場合のみ）
    if next_am is not None and next_pm is not None:
        ws.append([])
        ws.append([f"翌日予定（{next_date_str}）"])
        ws.cell(ws.max_row, 1).font = Font(bold=True, size=11)
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=7)

        ws.append(["区分", "枠", "タスク", "時間(h)", "", "", ""])
        nh_row = ws.max_row
        for col in range(1, 5):
            ws.cell(nh_row, col).font = bold
            ws.cell(nh_row, col).fill = PatternFill("solid", fgColor="4472C4")
            ws.cell(nh_row, col).font = white_bold

        for slot_label, entries in [("午前", next_am), ("午後", next_pm)]:
            for i, entry in enumerate(entries):
                if entry.get("task_name") or (entry.get("hours") or 0) > 0:
                    ws.append([
                        slot_label if i == 0 else "",
                        str(i + 1),
                        entry.get("task_name", ""),
                        entry.get("hours") or "",
                        "", "", "",
                    ])
                    for col in range(1, 5):
                        ws.cell(ws.max_row, col).fill = next_fill

    # 列幅
    col_widths = [8, 5, 24, 8, 24, 8, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_daily_excel(
    user: dict,
    date_str: str,
    schedule_am: list,
    schedule_pm: list,
    result: dict,
    comment: dict,
    next_am: list | None = None,
    next_pm: list | None = None,
    next_date_str: str = "",
) -> io.BytesIO:
    """1ユーザーの日報Excelを1シートで生成しBytesIOで返す。

    Args:
        user: ユーザー情報（name, std_hours を含む）。
        date_str: 日付（'YYYY-MM-DD' 形式）。
        schedule_am: 予定の午前5枠リスト。
        schedule_pm: 予定の午後5枠リスト。
        result: 日次実績 {'am': [...], 'pm': [...]}。
        comment: 日次コメント {'reflection': str, 'action': str}。
        next_am: 翌日予定の午前5枠（省略可）。
        next_pm: 翌日予定の午後5枠（省略可）。
        next_date_str: 翌日の日付文字列（省略可）。

    Returns:
        io.BytesIO: 生成されたExcelファイルのバイトストリーム（先頭にシーク済み）。
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _append_daily_sheet(wb, user, date_str, schedule_am, schedule_pm, result, comment, next_am, next_pm, next_date_str)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@export_bp.route("/daily/<date_str>", endpoint="export_daily")
def export_daily(date_str: str) -> object:
    """ログインユーザー自身の日報をExcelでダウンロードする。

    クエリパラメータ `date` で対象日を指定できる（未指定時は今日）。

    Args:
        date_str: 日報対象日（'YYYY-MM-DD' 形式）。

    Returns:
        object: Excelファイルのダウンロードレスポンス、または未ログイン時はリダイレクト。
    """
    if not session.get("user_id"):
        return redirect(url_for("auth.login"))

    try:
        date_obj = date.fromisoformat(date_str)
    except ValueError:
        date_obj = date.today()
        date_str = date_obj.isoformat()

    user = get_user_by_id(session["user_id"])
    if user is None:
        abort(404)

    # 当日予定取得
    week_start = _get_monday(date_obj).isoformat()
    day_of_week = date_obj.weekday()
    schedule = get_weekly_schedule(user["id"], week_start)
    day_sch = schedule.get(day_of_week, {"am": [], "pm": []})
    schedule_am = day_sch.get("am", [])
    schedule_pm = day_sch.get("pm", [])

    # 実績・コメント取得
    result = get_daily_result(user["id"], date_str)
    comment = get_daily_comment(user["id"], date_str)

    # 翌日予定取得
    next_date = date_obj + timedelta(days=1)
    while next_date.weekday() >= 5:
        next_date += timedelta(days=1)
    next_week_start = _get_monday(next_date).isoformat()
    next_day_of_week = next_date.weekday()
    next_schedule = get_weekly_schedule(user["id"], next_week_start)
    next_day_sch = next_schedule.get(next_day_of_week, {"am": [], "pm": []})
    next_am = next_day_sch.get("am", [])
    next_pm = next_day_sch.get("pm", [])

    try:
        buf = _build_daily_excel(
            user, date_str, schedule_am, schedule_pm, result, comment,
            next_am, next_pm, next_date.isoformat(),
        )
    except Exception:
        logger.exception("日報Excel生成中にエラーが発生しました (user=%s, date=%s)", user.get("name"), date_str)
        flash("Excel生成中にエラーが発生しました", "warning")
        return redirect(url_for("daily_bp.daily_view", date_str=date_str))
    filename = f"daily_{user['name']}_{date_str}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@export_bp.route("/admin_daily/<date_str>", endpoint="export_admin_daily")
def export_admin_daily(date_str: str) -> object:
    """管理者用: 全ユーザーの日報を1つのExcelファイル（シート別）でダウンロードする。

    Args:
        date_str: 日報対象日（'YYYY-MM-DD' 形式）。

    Returns:
        object: Excelファイルのダウンロードレスポンス。管理者以外は403。
    """
    if not session.get("user_id"):
        return redirect(url_for("auth.login"))
    if not is_privileged(session.get("user_role", "")):
        abort(403)

    try:
        date_obj = date.fromisoformat(date_str)
    except ValueError:
        date_obj = date.today()
        date_str = date_obj.isoformat()

    # 管理職の場合は自部署のみ、マスタは全員
    login_role_exp: str = session.get("user_role", "")
    login_dept_exp: str = session.get("user_dept", "")
    if is_master(login_role_exp):
        all_users = get_all_users()
    else:
        all_users = get_all_users(dept_filter=login_dept_exp if login_dept_exp else None)
    if not all_users:
        flash("ユーザーが存在しないためExcelを生成できません", "warning")
        return redirect(url_for("admin_bp.dashboard"))

    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        week_start = _get_monday(date_obj).isoformat()
        day_of_week = date_obj.weekday()
        for user in all_users:
            schedule = get_weekly_schedule(user["id"], week_start)
            day_sch = schedule.get(day_of_week, {"am": [], "pm": []})
            schedule_am = day_sch.get("am", [])
            schedule_pm = day_sch.get("pm", [])
            result = get_daily_result(user["id"], date_str)
            comment = get_daily_comment(user["id"], date_str)
            _append_daily_sheet(wb, user, date_str, schedule_am, schedule_pm, result, comment)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
    except Exception:
        logger.exception("全員日報Excel生成中にエラーが発生しました (date=%s)", date_str)
        flash("Excel生成中にエラーが発生しました", "warning")
        return redirect(url_for("admin_bp.dashboard"))

    admin_user = get_user_by_id(session["user_id"])
    admin_name = admin_user["name"] if admin_user else "admin"
    filename = f"all_daily_{admin_name}_{date_str}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@export_bp.route("/admin_report/<date_str>", endpoint="export_admin_report")
def export_admin_report(date_str: str) -> object:
    """管理者用日次報告をExcelでダウンロードする。

    管理者自身の作業報告 + 部下全員の作業サマリー + 管理者自身の翌日予定
    を1シートに出力する。

    Args:
        date_str: 報告対象日（'YYYY-MM-DD' 形式）。

    Returns:
        object: Excelファイルのダウンロードレスポンス。管理者以外は403。
    """
    if not session.get("user_id"):
        return redirect(url_for("auth.login"))
    if not is_privileged(session.get("user_role", "")):
        abort(403)

    try:
        date_obj = date.fromisoformat(date_str)
    except ValueError:
        date_obj = date.today()
        date_str = date_obj.isoformat()

    admin_user = get_user_by_id(session["user_id"])
    if admin_user is None:
        abort(404)

    week_start = _get_monday(date_obj).isoformat()
    day_of_week = date_obj.weekday()

    # 管理者自身の実績
    admin_schedule = get_weekly_schedule(admin_user["id"], week_start)
    admin_day_sch = admin_schedule.get(day_of_week, {"am": [], "pm": []})
    admin_schedule_am = admin_day_sch.get("am", [])
    admin_schedule_pm = admin_day_sch.get("pm", [])
    admin_result = get_daily_result(admin_user["id"], date_str)
    admin_comment = get_daily_comment(admin_user["id"], date_str)

    # 翌日予定
    next_date = date_obj + timedelta(days=1)
    while next_date.weekday() >= 5:
        next_date += timedelta(days=1)
    next_week_start = _get_monday(next_date).isoformat()
    next_day_of_week = next_date.weekday()
    next_schedule = get_weekly_schedule(admin_user["id"], next_week_start)
    next_day_sch = next_schedule.get(next_day_of_week, {"am": [], "pm": []})
    next_am = next_day_sch.get("am", [])
    next_pm = next_day_sch.get("pm", [])

    # 部下一覧（管理職の場合は自部署のみ、マスタは全員・特権ユーザー自身は除外）
    sub_login_role: str = session.get("user_role", "")
    sub_login_dept: str = session.get("user_dept", "")
    if is_master(sub_login_role):
        all_target_users = get_all_users()
    else:
        all_target_users = get_all_users(dept_filter=sub_login_dept if sub_login_dept else None)
    subordinates = [u for u in all_target_users if not is_privileged(u["role"]) and u["id"] != admin_user["id"]]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "日次報告"

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="4472C4")
    white_bold = Font(bold=True, color="FFFFFF")
    am_fill = PatternFill("solid", fgColor="E8F4FB")
    pm_fill = PatternFill("solid", fgColor="FDF8F0")
    total_fill = PatternFill("solid", fgColor="D8ECD8")
    sub_fill = PatternFill("solid", fgColor="F5F5DC")
    next_fill = PatternFill("solid", fgColor="E8F0E8")
    comment_fill = PatternFill("solid", fgColor="F0E8F8")
    red_bold = Font(bold=True, color="DC3545")
    center = Alignment(horizontal="center")
    wrap = Alignment(wrap_text=True, vertical="top")

    col_widths = [8, 5, 24, 8, 24, 8, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    def add_row(values: list, fill=None, font=None, merge_to: int = 0, height: int = 0) -> None:
        nonlocal row
        for col, val in enumerate(values, 1):
            cell = ws.cell(row, col, val)
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
        if merge_to > 1:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_to)
        if height:
            ws.row_dimensions[row].height = height
        row += 1

    # タイトル
    add_row([f"日次報告：{date_str}　報告者：{admin_user['name']}"], font=Font(bold=True, size=13), merge_to=7)
    add_row([])

    # --- 管理者自身の作業報告 ---
    add_row(["■ 自身の作業報告"], font=Font(bold=True, size=11), merge_to=7)
    add_row(["区分", "枠", "予定タスク", "予定(h)", "実績タスク", "実績(h)", "差異(h)"],
            fill=header_fill, font=white_bold)

    std_hours = admin_user.get("std_hours") or 8.0
    plan_total_all = 0.0
    result_total_all = 0.0
    for slot_label, sched_list, res_list, fill in [
        ("午前", admin_schedule_am, admin_result.get("am", []), am_fill),
        ("午後", admin_schedule_pm, admin_result.get("pm", []), pm_fill),
    ]:
        plan_sub = 0.0
        res_sub = 0.0
        for i in range(5):
            sc = sched_list[i] if i < len(sched_list) else {"task_name": "", "hours": 0.0}
            rs = res_list[i] if i < len(res_list) else {"task_name": "", "hours": 0.0}
            ph = sc.get("hours") or 0.0
            rh = rs.get("hours") or 0.0
            plan_sub += ph
            res_sub += rh
            plan_total_all += ph
            result_total_all += rh
            add_row([
                slot_label if i == 0 else "", str(i + 1),
                sc.get("task_name", ""), ph if ph else "",
                rs.get("task_name", ""), rh if rh else "",
                f"{rh - ph:+.2f}" if (ph or rh) else "",
            ], fill=fill)
        add_row(["", "計", "", plan_sub, "", res_sub, ""],
                fill=PatternFill("solid", fgColor="C8D8E8" if slot_label == "午前" else "E8D8A8"),
                font=bold)

    # 1日合計
    diff_all = result_total_all - plan_total_all
    diff_font = red_bold if abs(diff_all) > 0.001 else bold
    add_row(["", "1日合計", "", plan_total_all, "", result_total_all, f"{diff_all:+.2f}"],
            fill=total_fill, font=bold)

    # コメント
    add_row([])
    add_row(["本日の振り返り"], fill=comment_fill, font=bold, merge_to=7)
    refl = admin_comment.get("reflection", "")
    h = max(40, len(refl) // 30 * 15 + 20)
    add_row([refl], merge_to=7, height=h)
    ws.cell(row - 1, 1).alignment = wrap
    add_row(["今後の対策・懸念事項"], fill=comment_fill, font=bold, merge_to=7)
    act = admin_comment.get("action", "")
    h = max(40, len(act) // 30 * 15 + 20)
    add_row([act], merge_to=7, height=h)
    ws.cell(row - 1, 1).alignment = wrap

    add_row([])

    # --- 部下の作業報告 ---
    add_row(["■ 部下の作業報告"], font=Font(bold=True, size=11), merge_to=7)
    add_row(["氏名", "部署", "実績合計(h)", "主な実績タスク", "", "", ""],
            fill=header_fill, font=white_bold)

    for sub in subordinates:
        sub_result = get_daily_result(sub["id"], date_str)
        sub_total = sum(
            (e.get("hours") or 0.0)
            for s in ("am", "pm")
            for e in sub_result.get(s, [])
        )
        tasks_done = [
            e.get("task_name", "")
            for s in ("am", "pm")
            for e in sub_result.get(s, [])
            if e.get("task_name") and (e.get("hours") or 0) > 0
        ]
        task_str = "、".join(dict.fromkeys(tasks_done))  # 重複除去
        add_row([sub["name"], sub.get("dept", ""), sub_total, task_str, "", "", ""],
                fill=sub_fill)

    add_row([])

    # --- 翌日予定 ---
    add_row([f"■ 翌日予定（{next_date.isoformat()}）"], font=Font(bold=True, size=11), merge_to=7)
    add_row(["区分", "枠", "タスク", "時間(h)", "", "", ""],
            fill=header_fill, font=white_bold)

    for slot_label, entries, fill in [("午前", next_am, am_fill), ("午後", next_pm, pm_fill)]:
        for i, entry in enumerate(entries):
            if entry.get("task_name") or (entry.get("hours") or 0) > 0:
                add_row([
                    slot_label if i == 0 else "", str(i + 1),
                    entry.get("task_name", ""), entry.get("hours") or "", "", "", "",
                ], fill=fill)

    try:
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
    except Exception:
        logger.exception("管理者日次報告Excel生成中にエラーが発生しました (user=%s, date=%s)", admin_user.get("name"), date_str)
        flash("Excel生成中にエラーが発生しました", "warning")
        return redirect(url_for("daily_bp.daily_view", date_str=date_str))

    filename = f"admin_report_{admin_user['name']}_{date_str}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@export_bp.route("/my", endpoint="export_my")
def export_my() -> object:
    """ログインユーザー自身の週間予定をExcelでダウンロードする。

    クエリパラメータ `week` で対象週を指定できる（未指定時は今週月曜）。

    Returns:
        object: Excelファイルのダウンロードレスポンス、または未ログイン時はリダイレクト。
    """
    if not session.get("user_id"):
        return redirect(url_for("auth.login"))

    week_param: str = request.args.get("week", "")
    if week_param:
        try:
            week_date = date.fromisoformat(week_param)
            week_start = _get_monday(week_date).isoformat()
        except ValueError:
            week_start = _get_current_week_start()
    else:
        week_start = _get_current_week_start()

    user = get_user_by_id(session["user_id"])
    if user is None:
        abort(404)

    schedule = get_weekly_schedule(user["id"], week_start)
    buf = _build_schedule_excel(user, week_start, schedule)

    filename = f"schedule_{user['name']}_{date.today().isoformat()}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@export_bp.route("/team_week", endpoint="export_team_week")
def export_team_week() -> object:
    """管理者用: 全メンバーの今週予定を1シートのExcelでダウンロードする。

    ユーザー管理の登録順に全員のデータを1シートに縦積みして出力する。
    管理職は自部署のメンバーのみ、マスタは全員対象。

    Returns:
        object: Excelファイルのダウンロードレスポンス。管理者以外は403。
    """
    if not session.get("user_id"):
        return redirect(url_for("auth.login"))
    if not is_privileged(session.get("user_role", "")):
        abort(403)

    week_start = _get_current_week_start()
    login_role: str = session.get("user_role", "")
    login_dept: str = session.get("user_dept", "")
    users = get_all_users(dept_filter=login_dept if login_dept else None)

    if not users:
        flash("対象ユーザーが存在しないためExcelを生成できません", "warning")
        return redirect(url_for("admin_bp.dashboard"))

    dept_label = login_dept if login_dept else "全員"
    try:
        buf = _build_team_week_from_schedule_tpl(users, week_start)
    except Exception:
        logger.exception("全員週間予定Excel生成中にエラーが発生しました (week=%s)", week_start)
        flash("Excel生成中にエラーが発生しました", "warning")
        return redirect(url_for("admin_bp.dashboard"))

    filename = f"週間予定_{dept_label}_{week_start}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@export_bp.route("/user/<int:user_id>", endpoint="export_user")
def export_user(user_id: int) -> object:
    """管理者が指定ユーザーの週間予定をExcelでダウンロードする。

    週間予定画面のExcelダウンロードと同じテンプレート形式（前週・今週・翌週の3シート）で出力する。
    クエリパラメータ `week` で今週基準日を指定できる（未指定時は今週月曜）。

    Args:
        user_id (int): エクスポート対象のユーザーID。

    Returns:
        object: Excelファイルのダウンロードレスポンス。管理者以外は403。
    """
    if not session.get("user_id"):
        return redirect(url_for("auth.login"))

    if not is_privileged(session.get("user_role", "")):
        abort(403)

    week_param: str = request.args.get("week", "")
    if week_param:
        try:
            week_date = date.fromisoformat(week_param)
            week_start = _get_monday(week_date).isoformat()
        except ValueError:
            week_start = _get_current_week_start()
    else:
        week_start = _get_current_week_start()

    user = get_user_by_id(user_id)
    if user is None:
        abort(404)

    login_user = {"id": int(session.get("user_id", 0)), "role": session.get("user_role", ""), "dept": session.get("user_dept", "")}
    if not can_access_user(login_user, dict(user)):
        abort(403)

    mode: str = request.args.get("mode", "plan")
    if mode not in ("plan", "result"):
        mode = "plan"

    try:
        buf = _build_schedule_excel_from_tpl(user, week_start, mode=mode)
    except Exception:
        logger.exception("テンプレートExcel生成に失敗しました。フォールバック出力を使用します (user_id=%d)", user_id)
        schedule = get_weekly_schedule(user["id"], week_start)
        buf = _build_schedule_excel(user, week_start, schedule)

    filename = f"週間予定実績表_{user['name']}_{date.today().isoformat()}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@export_bp.route("/multi_week", endpoint="export_multi_week")
def export_multi_week() -> object:
    """ログインユーザーの前週・今週・来週の予定を1つのExcelファイルに出力する。

    クエリパラメータ `week` を起点に3シートを生成する。
    データが存在しないシートもシートとして出力するが、空の旨を記載する。

    Returns:
        object: Excelファイルのダウンロードレスポンス、または未ログイン時はリダイレクト。
    """
    if not session.get("user_id"):
        return redirect(url_for("auth.login"))

    week_param: str = request.args.get("week", "")
    if week_param:
        try:
            week_date = date.fromisoformat(week_param)
            week_start = _get_monday(week_date).isoformat()
        except ValueError:
            week_start = _get_current_week_start()
    else:
        week_start = _get_current_week_start()

    mode: str = request.args.get("mode", "result")

    user = get_user_by_id(session["user_id"])
    if user is None:
        abort(404)

    try:
        buf = _build_schedule_excel_from_tpl(user, week_start, mode=mode)
    except Exception:
        logger.exception("テンプレートExcel生成に失敗しました。フォールバック出力を使用します。")
        # テンプレートが使えない場合は従来フォーマットにフォールバック
        ws_date = date.fromisoformat(week_start)
        weeks = [
            ("前週", (ws_date - timedelta(weeks=1)).isoformat()),
            ("今週", week_start),
            ("翌週", (ws_date + timedelta(weeks=1)).isoformat()),
        ]
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for sheet_label, wk_start in weeks:
            schedule = get_weekly_schedule(user["id"], wk_start)
            _append_schedule_sheet(wb, user, wk_start, sheet_label, schedule)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

    filename = f"週間予定実績表_{user['name']}_{date.today().isoformat()}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---------------------------------------------------------------------------
# 日次業務報告テンプレート出力
# ---------------------------------------------------------------------------

#: 日次業務報告テンプレートExcelのパス
REPORT_TPL: Path = (
    Path(__file__).parent.parent.parent / "reports" / "tpl" / "日次業務報告_テンプレート.xlsx"
)


def _fill_report_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    user: dict,
    date_str: str,
    schedule_am: list,
    schedule_pm: list,
    result: dict,
    comment: dict,
    next_am: list,
    next_pm: list,
) -> None:
    """テンプレートシートに日報データを書き込む。

    セルスタイルはテンプレートのものをそのまま保持し、値のみ上書きする。

    Args:
        ws: 書き込み対象のワークシート。
        user: ユーザー情報辞書（'name', 'std_hours' を含む）。
        date_str: 報告日（'YYYY-MM-DD' 形式）。
        schedule_am: 当日AM予定タスクリスト（最大5要素）。
        schedule_pm: 当日PM予定タスクリスト（最大5要素）。
        result: 日次実績 {'am': [...], 'pm': [...]}。
        comment: 日次コメント {'reflection': str, 'action': str}。
        next_am: 翌日AM予定タスクリスト（最大5要素）。
        next_pm: 翌日PM予定タスクリスト（最大5要素）。
    """
    # ---- ヘッダー部 ----
    ws["C5"] = date_str
    ws["C6"] = user.get("name", "")

    # 累積勤務時間：予定（D8）・実績（F8）
    plan_total: float = sum(
        (e.get("hours") or 0.0) for e in schedule_am + schedule_pm
    )
    result_total: float = sum(
        (e.get("hours") or 0.0)
        for slot in ("am", "pm")
        for e in result.get(slot, [])
    )
    ws["D8"] = plan_total if plan_total else None
    ws["F8"] = result_total if result_total else None

    # ---- タスク行（B12〜F21）: AM・PM を区別せず上から詰めて書き込む ----
    # リスケ・繰越の背景色定義
    _fill_riske: PatternFill = PatternFill("solid", fgColor="FFC000")   # 黄色（リスケ）
    _fill_carry: PatternFill = PatternFill("solid", fgColor="BDD7EE")   # 水色（繰越）

    # AM・PM 全スロットを順番にフラットに収集（データあり行を詰める）
    flat_entries: list[dict] = []
    for slot in ("am", "pm"):
        sched_list: list = schedule_am if slot == "am" else schedule_pm
        result_list: list = result.get(slot, [])
        for idx in range(5):
            sched_entry: dict = sched_list[idx] if idx < len(sched_list) else {}
            result_entry: dict = result_list[idx] if idx < len(result_list) else {}
            planned_task: str = sched_entry.get("task_name", "") or ""
            result_task: str = result_entry.get("task_name", "") or ""
            hours: float = result_entry.get("hours") or 0.0
            defer_date: str = result_entry.get("defer_date", "") or ""
            is_carryover: int = result_entry.get("is_carryover", 0) or 0
            subcategory: str = result_entry.get("subcategory_name", "") or ""
            # いずれかデータがある行のみ収集
            if planned_task or result_task or hours > 0:
                flat_entries.append({
                    "subcategory": subcategory,
                    "planned_task": planned_task,
                    "result_task": result_task,
                    "hours": hours,
                    "defer_date": defer_date,
                    "is_carryover": is_carryover,
                })

    # 行12〜21 に上から詰めて書き込み（最大10行）
    for row_offset in range(10):
        row: int = 12 + row_offset
        if row_offset < len(flat_entries):
            entry: dict = flat_entries[row_offset]
            # 状態判定
            if entry["result_task"]:
                if entry["defer_date"]:
                    status: str = "リスケ"
                elif entry["is_carryover"]:
                    status = "着手"
                elif entry["hours"] > 0:
                    status = "完了"
                else:
                    status = ""
            else:
                status = ""

            ws[f"B{row}"] = entry["subcategory"] or None
            ws[f"C{row}"] = entry["planned_task"] or None
            ws[f"D{row}"] = entry["result_task"] or None
            ws[f"E{row}"] = status or None
            ws[f"F{row}"] = entry["hours"] if entry["hours"] > 0 else None

            # リスケ・繰越の背景色を適用
            if entry["defer_date"]:
                for col in ("B", "C", "D", "E", "F"):
                    ws[f"{col}{row}"].fill = _fill_riske
            elif entry["is_carryover"]:
                for col in ("B", "C", "D", "E", "F"):
                    ws[f"{col}{row}"].fill = _fill_carry
        else:
            # データなし行はプレースホルダーをクリア
            for col in ("B", "C", "D", "E", "F"):
                ws[f"{col}{row}"] = None

    # ---- コメント ----
    ws["A25"] = comment.get("reflection", "") or None
    ws["A28"] = comment.get("action", "") or None

    # ---- 翌日予定 ----
    next_am_task: str = next_am[0].get("task_name", "") if next_am else ""
    next_pm_task: str = next_pm[0].get("task_name", "") if next_pm else ""
    ws["C31"] = next_am_task if next_am_task else None
    ws["C32"] = next_pm_task if next_pm_task else None


def _build_report_excel(
    user: dict,
    date_str: str,
    schedule_am: list,
    schedule_pm: list,
    result: dict,
    comment: dict,
    next_am: list,
    next_pm: list,
) -> io.BytesIO:
    """テンプレートをコピーして1ユーザー分の日報Excelを生成しBytesIOで返す。

    Args:
        user: ユーザー情報辞書。
        date_str: 報告日（'YYYY-MM-DD' 形式）。
        schedule_am: 当日AM予定タスクリスト。
        schedule_pm: 当日PM予定タスクリスト。
        result: 日次実績辞書。
        comment: 日次コメント辞書。
        next_am: 翌日AM予定タスクリスト。
        next_pm: 翌日PM予定タスクリスト。

    Returns:
        io.BytesIO: Excelファイルのバイトストリーム。
    """
    wb = openpyxl.load_workbook(REPORT_TPL)
    ws = wb["日次業務報告"]
    _fill_report_sheet(ws, user, date_str, schedule_am, schedule_pm, result, comment, next_am, next_pm)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@export_bp.route("/report/download", endpoint="download_report")
def download_report() -> object:
    """ログインユーザー本人の日報をテンプレートExcelでダウンロードする。

    クエリパラメータ ``date`` で対象日を指定する（未指定時は今日）。

    Returns:
        object: Excelファイルのダウンロードレスポンス、または未ログイン時はリダイレクト。
    """
    if not session.get("user_id"):
        return redirect(url_for("auth.login"))

    raw_date: str = request.args.get("date", "").strip()
    try:
        date_obj = date.fromisoformat(raw_date)
        date_str: str = date_obj.isoformat()
    except ValueError:
        date_obj = date.today()
        date_str = date_obj.isoformat()

    user = get_user_by_id(int(session["user_id"]))
    if user is None:
        abort(404)

    week_start: str = _get_monday(date_obj).isoformat()
    day_of_week: int = date_obj.weekday()
    schedule: dict = get_weekly_schedule(user["id"], week_start)
    day_sch: dict = schedule.get(day_of_week, {"am": [], "pm": []})
    schedule_am: list = day_sch.get("am", [])
    schedule_pm: list = day_sch.get("pm", [])

    result: dict = get_daily_result(user["id"], date_str)
    comment: dict = get_daily_comment(user["id"], date_str)

    # 翌日予定
    next_date = date_obj + timedelta(days=1)
    while next_date.weekday() >= 5:
        next_date += timedelta(days=1)
    next_week_start: str = _get_monday(next_date).isoformat()
    next_dow: int = next_date.weekday()
    next_schedule: dict = get_weekly_schedule(user["id"], next_week_start)
    next_day_sch: dict = next_schedule.get(next_dow, {"am": [], "pm": []})
    next_am: list = next_day_sch.get("am", [])
    next_pm: list = next_day_sch.get("pm", [])

    try:
        buf = _build_report_excel(
            user, date_str, schedule_am, schedule_pm, result, comment, next_am, next_pm
        )
    except Exception:
        logger.exception(
            "日次業務報告Excel生成中にエラーが発生しました (user=%s, date=%s)",
            user.get("name"),
            date_str,
        )
        flash("日報Excel生成中にエラーが発生しました", "warning")
        return redirect(url_for("daily_bp.daily_view", date_str=date_str))

    filename: str = f"日次業務報告_{user['name']}_{date_str}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@export_bp.route("/report/team", endpoint="download_team_report")
def download_team_report() -> object:
    """管理職・マスタ用: 担当メンバーの日報をシート別にまとめたExcelをダウンロードする。

    クエリパラメータ ``date`` で対象日を指定する（未指定時は今日）。
    管理職・マスタ以外からのアクセスは403を返す。

    Returns:
        object: Excelファイルのダウンロードレスポンス、または未ログイン時はリダイレクト。
    """
    if not session.get("user_id"):
        return redirect(url_for("auth.login"))
    if not is_privileged(session.get("user_role", "")):
        abort(403)

    raw_date: str = request.args.get("date", "").strip()
    try:
        date_obj = date.fromisoformat(raw_date)
        date_str: str = date_obj.isoformat()
    except ValueError:
        date_obj = date.today()
        date_str = date_obj.isoformat()

    login_user = get_user_by_id(int(session["user_id"]))
    if login_user is None:
        abort(404)

    # 担当メンバー取得（自分自身を先頭に含める）
    login_role: str = session.get("user_role", "")
    login_dept: str = session.get("user_dept", "")
    members: list[dict] = get_accessible_users(int(session["user_id"]), login_role, login_dept)
    # 自分自身が含まれていない場合は先頭に追加
    member_ids: set[int] = {m["id"] for m in members}
    if login_user["id"] not in member_ids:
        members = [login_user] + members

    if not members:
        flash("対象メンバーが存在しないためExcelを生成できません", "warning")
        return redirect(url_for("daily_bp.daily_view", date_str=date_str))

    week_start: str = _get_monday(date_obj).isoformat()
    day_of_week: int = date_obj.weekday()

    # 翌日算出
    next_date = date_obj + timedelta(days=1)
    while next_date.weekday() >= 5:
        next_date += timedelta(days=1)
    next_week_start: str = _get_monday(next_date).isoformat()
    next_dow: int = next_date.weekday()

    try:
        # テンプレートを直接操作し、同一ワークブック内でシートをコピーする
        wb = openpyxl.load_workbook(str(REPORT_TPL))
        ws_tpl = wb["日次業務報告"]

        for idx, member in enumerate(members):
            schedule: dict = get_weekly_schedule(member["id"], week_start)
            day_sch_m: dict = schedule.get(day_of_week, {"am": [], "pm": []})
            schedule_am: list = day_sch_m.get("am", [])
            schedule_pm: list = day_sch_m.get("pm", [])

            result: dict = get_daily_result(member["id"], date_str)
            comment: dict = get_daily_comment(member["id"], date_str)

            next_schedule: dict = get_weekly_schedule(member["id"], next_week_start)
            next_day_sch_m: dict = next_schedule.get(next_dow, {"am": [], "pm": []})
            next_am: list = next_day_sch_m.get("am", [])
            next_pm: list = next_day_sch_m.get("pm", [])

            if idx == 0:
                # 最初のメンバーはテンプレートシートをそのまま使用
                ws_new = ws_tpl
            else:
                # 同一ワークブック内でコピー
                ws_new = wb.copy_worksheet(ws_tpl)

            sheet_title: str = member["name"][:10]
            # 同名シート回避
            existing_titles: list[str] = [s.title for s in wb.worksheets]
            if sheet_title in existing_titles:
                for n in range(2, 100):
                    candidate: str = f"{sheet_title}({n})"[:31]
                    if candidate not in existing_titles:
                        sheet_title = candidate
                        break
            ws_new.title = sheet_title

            _fill_report_sheet(
                ws_new, member, date_str,
                schedule_am, schedule_pm,
                result, comment,
                next_am, next_pm,
            )

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
    except Exception:
        logger.exception(
            "チーム日次業務報告Excel生成中にエラーが発生しました (date=%s)",
            date_str,
        )
        flash("チーム日報Excel生成中にエラーが発生しました", "warning")
        return redirect(url_for("daily_bp.daily_view", date_str=date_str))

    filename: str = f"日次業務報告_チーム_{date_str}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ============================================================
# インポート機能（マスタ権限のみ）
# ============================================================

def _import_task_col(day: int) -> int:
    """曜日インデックス(0-4)からタスク名列番号を返す。

    Args:
        day: 曜日インデックス（0=月〜4=金）

    Returns:
        int: タスク名の列番号
    """
    return 4 + day * 3


def _import_time_col(day: int) -> int:
    """曜日インデックス(0-4)から時間列番号を返す。

    Args:
        day: 曜日インデックス（0=月〜4=金）

    Returns:
        int: 時間の列番号
    """
    return 6 + day * 3


def _parse_schedule_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, week_monday: date) -> dict:
    """Excelシートから週間予定データを解析する。

    テンプレート構造:
      Row 1 C4: 月曜日の日にち
      Row 2 C1: 氏名
      Row 4-8: 午前スロット1-5（各日: task_col=タスク名, time_col=時間）
      Row 9-13: 午後スロット1-5

    4行目（AM1）にタスク名がない日はスキップする。

    Args:
        ws: openpyxlのワークシート
        week_monday: 週の月曜日

    Returns:
        dict: {
            'user_name': str,
            'schedule': {day(0-4): {'am': [{task_name, hours}×5], 'pm': [同]}},
            'skipped_days': list[int],
            'imported_days': list[int],
        }
    """
    # 氏名取得（B2:C3 マージセル → A2 or B2 に値がある場合も）
    user_name: str = ""
    for col in (2, 3, 1):
        v = ws.cell(row=2, column=col).value
        if v and str(v).strip():
            user_name = str(v).strip()
            break
    # A1 にも氏名が入る場合がある
    if not user_name:
        v = ws.cell(row=1, column=1).value
        if v and str(v).strip() and str(v).strip() != "氏名":
            user_name = str(v).strip()

    schedule: dict = {}
    skipped_days: list[int] = []
    imported_days: list[int] = []

    for day in range(5):
        task_col = _import_task_col(day)
        time_col = _import_time_col(day)

        # 4行目（AM1）に値がなければスキップ
        am1_task = ws.cell(row=4, column=task_col).value
        if not am1_task or not str(am1_task).strip():
            skipped_days.append(day)
            continue

        imported_days.append(day)
        am_slots: list[dict] = []
        pm_slots: list[dict] = []

        # 午前: Row 4-8
        for i in range(5):
            row = 4 + i
            task = ws.cell(row=row, column=task_col).value
            hours = ws.cell(row=row, column=time_col).value
            am_slots.append({
                "task_name": str(task).strip() if task else "",
                "hours": float(hours) if hours and str(hours).strip() else 0.0,
            })

        # 午後: Row 9-13
        for i in range(5):
            row = 9 + i
            task = ws.cell(row=row, column=task_col).value
            hours = ws.cell(row=row, column=time_col).value
            pm_slots.append({
                "task_name": str(task).strip() if task else "",
                "hours": float(hours) if hours and str(hours).strip() else 0.0,
            })

        schedule[day] = {"am": am_slots, "pm": pm_slots}

    return {
        "user_name": user_name,
        "schedule": schedule,
        "skipped_days": skipped_days,
        "imported_days": imported_days,
    }


@export_bp.route("/import", methods=["GET", "POST"])
def import_schedule():
    """週間予定Excelインポート画面・処理（マスタ権限のみ）。

    GET: インポート画面を表示
    POST: アップロードされたExcelファイルから週間予定を取り込む

    Returns:
        str: インポート画面またはリダイレクト
    """
    from flask import render_template

    if "user_id" not in session:
        return redirect(url_for("auth.login"))
    if not is_master(session.get("user_role", "")):
        abort(403)

    if request.method == "GET":
        all_users = get_all_users()
        return render_template(
            "import_schedule.html",
            all_users=all_users,
            csrf_token=session.get("csrf_token", ""),
        )

    # POST: CSRF検証
    import secrets
    token = request.form.get("csrf_token", "")
    if not secrets.compare_digest(token, session.get("csrf_token", "")):
        abort(400)

    # ファイル取得
    file = request.files.get("excel_file")
    if not file or not file.filename:
        flash("ファイルを選択してください", "warning")
        return redirect(url_for("export_bp.import_schedule"))

    if not file.filename.endswith(".xlsx"):
        flash(".xlsx形式のファイルを選択してください", "warning")
        return redirect(url_for("export_bp.import_schedule"))

    # 対象ユーザーID
    target_user_id_str = request.form.get("target_user_id", "").strip()
    if not target_user_id_str:
        flash("対象ユーザーを選択してください", "warning")
        return redirect(url_for("export_bp.import_schedule"))

    try:
        target_user_id = int(target_user_id_str)
    except ValueError:
        flash("ユーザーIDが不正です", "warning")
        return redirect(url_for("export_bp.import_schedule"))

    target_user = get_user_by_id(target_user_id)
    if target_user is None:
        flash("対象ユーザーが見つかりません", "warning")
        return redirect(url_for("export_bp.import_schedule"))

    # 週開始日（月曜日）
    week_start_str = request.form.get("week_start", "").strip()
    if not week_start_str:
        flash("週開始日を入力してください", "warning")
        return redirect(url_for("export_bp.import_schedule"))

    try:
        week_date = date.fromisoformat(week_start_str)
    except ValueError:
        flash("日付形式が不正です（YYYY-MM-DD）", "warning")
        return redirect(url_for("export_bp.import_schedule"))

    # 月曜日に丸める
    week_monday = _get_monday(week_date)
    week_start = week_monday.isoformat()

    # Excel解析
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb[wb.sheetnames[0]]
        parsed = _parse_schedule_sheet(ws, week_monday)
    except Exception as e:
        logger.exception("Excelインポート解析エラー")
        flash(f"Excelファイルの解析に失敗しました: {e}", "warning")
        return redirect(url_for("export_bp.import_schedule"))

    if not parsed["imported_days"]:
        flash("取り込み可能なデータがありませんでした（全日の4行目が空です）", "warning")
        return redirect(url_for("export_bp.import_schedule"))

    # 既存データとマージ（スキップ日は既存データを維持）
    existing = get_weekly_schedule(target_user_id, week_start)
    merged_data: dict = {}
    for day in range(5):
        if day in parsed["schedule"]:
            merged_data[day] = parsed["schedule"][day]
        else:
            merged_data[day] = existing.get(day, {"am": [{"task_name": "", "hours": 0.0}] * 5,
                                                   "pm": [{"task_name": "", "hours": 0.0}] * 5})

    # 保存
    login_user = get_user_by_id(int(session["user_id"]))
    updater_name: str = login_user["name"] if login_user else ""
    save_weekly_schedule(target_user_id, week_start, merged_data, updated_by=updater_name)

    day_names = ["月", "火", "水", "木", "金"]
    imported_str = "・".join(day_names[d] for d in parsed["imported_days"])
    skipped_str = "・".join(day_names[d] for d in parsed["skipped_days"]) if parsed["skipped_days"] else "なし"

    flash(
        f"「{target_user['name']}」の週間予定（{week_start}週）を取り込みました。"
        f" 取込: {imported_str}曜 / スキップ: {skipped_str}",
        "success",
    )
    logger.info(
        "Excelインポート完了: user_id=%d week=%s days=%s by=%s",
        target_user_id, week_start, parsed["imported_days"], updater_name,
    )
    return redirect(url_for("export_bp.import_schedule"))
