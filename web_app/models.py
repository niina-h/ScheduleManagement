"""web_app/models.py - DBアクセス関数群（ドメインモデル層）"""
from __future__ import annotations

import logging
import secrets
import sqlite3
from datetime import date, datetime, timedelta

from werkzeug.security import check_password_hash, generate_password_hash

from .database import get_db

logger = logging.getLogger(__name__)

# 1週間あたりの曜日数（月〜金）
_DAYS = range(5)
# タイムスロット
_SLOTS = ("am", "pm")
# 1スロットあたりの枠数
_SLOT_SIZE = 5


# ---------------------------------------------------------------------------
# ユーザー関連
# ---------------------------------------------------------------------------


def get_all_users(dept_filter: str | None = None) -> list[dict]:
    """全ユーザー一覧を取得する。

    Args:
        dept_filter: 指定時はその部署のユーザーのみ返す。None なら全件。

    Returns:
        list[dict]: ユーザー情報のリスト（id, name, role, dept, std_hours_am, std_hours_pm, std_hours）
    """
    db = get_db()
    if dept_filter is not None:
        rows = db.execute(
            "SELECT id, name, role, dept, std_hours_am, std_hours_pm, std_hours, manager_id, last_name, first_name "
            "FROM users WHERE dept = ? ORDER BY display_order ASC, id ASC",
            (dept_filter,),
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT id, name, role, dept, std_hours_am, std_hours_pm, std_hours, manager_id, last_name, first_name "
            "FROM users ORDER BY display_order ASC, id ASC"
        ).fetchall()
    return [dict(row) for row in rows]


def update_users_order(user_ids: list[int]) -> None:
    """ユーザーの表示順を更新する。

    Args:
        user_ids: 表示したい順番に並べたユーザーIDのリスト。
                  リストのインデックスが display_order の値となる。
    """
    db = get_db()
    for order, uid in enumerate(user_ids):
        db.execute(
            "UPDATE users SET display_order = ? WHERE id = ?",
            (order, uid),
        )
    db.commit()


def get_user_by_id(user_id: int) -> dict | None:
    """
    IDでユーザーを取得する。

    Args:
        user_id: ユーザーID

    Returns:
        dict | None: ユーザー情報。存在しない場合はNone
    """
    db = get_db()
    row = db.execute(
        "SELECT id, name, role, dept, std_hours_am, std_hours_pm, std_hours, manager_id "
        "FROM users WHERE id = ?",
        (user_id,),
    ).fetchone()
    return dict(row) if row else None


def get_user_by_name(name: str) -> dict | None:
    """
    名前でユーザーを取得する。

    Args:
        name: ユーザー名

    Returns:
        dict | None: ユーザー情報。存在しない場合はNone
    """
    db = get_db()
    row = db.execute(
        "SELECT id, name, role, dept, std_hours_am, std_hours_pm, std_hours, manager_id "
        "FROM users WHERE name = ?",
        (name,),
    ).fetchone()
    return dict(row) if row else None


def get_direct_reports(manager_id: int) -> list[dict]:
    """直属の部下ユーザー一覧を取得する（manager_id が一致するユーザー）。

    Args:
        manager_id: 上長のユーザーID。

    Returns:
        list[dict]: 直属部下のユーザー情報リスト。
    """
    db = get_db()
    rows = db.execute(
        "SELECT id, name, role, dept, std_hours_am, std_hours_pm, std_hours, manager_id "
        "FROM users WHERE manager_id = ? ORDER BY display_order ASC, id ASC",
        (manager_id,),
    ).fetchall()
    return [dict(row) for row in rows]


def get_accessible_users(login_id: int, login_role: str, login_dept: str) -> list[dict]:
    """ログインユーザーがアクセスできるユーザーリストを返す。

    管理職は直属部下（manager_id=login_id）が設定されていればそれのみ、
    未設定の場合は同一部署（マスタ除く）を返す。
    マスタは同一部署のマスタ以外全員を返す。

    Args:
        login_id: ログインユーザーのID。
        login_role: ログインユーザーの役職。
        login_dept: ログインユーザーの部署。

    Returns:
        list[dict]: アクセス可能なユーザーリスト。
    """
    if login_role == "マスタ":
        # マスタは同一部署のマスタ以外全員を返す（所属外の部門は除外）
        dept_users = get_all_users(dept_filter=login_dept if login_dept else None)
        others = [u for u in dept_users if u.get("role") != "マスタ"]
        self_user = next((u for u in dept_users if u.get("id") == login_id), None)
        if self_user and not any(u.get("id") == login_id for u in others):
            return [self_user] + others
        return others
    if login_role == "管理職":
        all_dept = get_all_users(dept_filter=login_dept if login_dept else None)
        self_user = next((u for u in all_dept if u.get("id") == login_id), None)
        direct = get_direct_reports(login_id)
        if direct:
            members = direct
        else:
            # 直属部下未設定の場合: 同一部署（マスタ除く）
            members = [u for u in all_dept if u.get("role") != "マスタ"]
        # 自分自身をリスト先頭に追加（未含の場合）
        if self_user and not any(u.get("id") == login_id for u in members):
            return [self_user] + members
        return members
    return []


def save_user_manager(user_id: int, manager_id: int | None) -> None:
    """ユーザーの直属の上長（manager_id）を設定する。

    Args:
        user_id: 対象ユーザーのID。
        manager_id: 上長のユーザーID。None の場合は割り当てを解除する。
    """
    db = get_db()
    db.execute(
        "UPDATE users SET manager_id = ? WHERE id = ?",
        (manager_id, user_id),
    )
    db.commit()


def set_user_password(user_id: int, password: str) -> bool:
    """ユーザーのパスワードハッシュを設定する。

    Args:
        user_id: 対象ユーザーID
        password: 平文パスワード

    Returns:
        bool: 更新成功時 True
    """
    db = get_db()
    try:
        hashed = generate_password_hash(password)
        db.execute(
            "UPDATE users SET password_hash = ? WHERE id = ?",
            (hashed, user_id),
        )
        db.commit()
        return True
    except Exception:
        db.rollback()
        logger.exception("パスワード設定中にエラーが発生しました (user_id=%s)", user_id)
        return False


def clear_user_password(user_id: int) -> bool:
    """ユーザーのパスワードを削除する（パスワード不要に戻す）。

    Args:
        user_id: 対象ユーザーID

    Returns:
        bool: 更新成功時 True
    """
    db = get_db()
    try:
        db.execute(
            "UPDATE users SET password_hash = '' WHERE id = ?",
            (user_id,),
        )
        db.commit()
        return True
    except Exception:
        db.rollback()
        logger.exception("パスワード削除中にエラーが発生しました (user_id=%s)", user_id)
        return False


def check_user_password(user_id: int, password: str) -> bool:
    """ユーザーIDとパスワードを検証する。

    Args:
        user_id: ユーザーID
        password: 平文パスワード

    Returns:
        bool: パスワードが正しい場合 True
    """
    db = get_db()
    row = db.execute(
        "SELECT password_hash FROM users WHERE id = ?",
        (user_id,),
    ).fetchone()
    if row is None:
        return False
    hashed: str = row["password_hash"] or ""
    if not hashed:
        return False
    return check_password_hash(hashed, password)


def user_has_password(user_id: int) -> bool:
    """ユーザーにパスワードが設定されているか確認する。

    Args:
        user_id: ユーザーID

    Returns:
        bool: パスワードが設定されている場合 True
    """
    db = get_db()
    row = db.execute(
        "SELECT password_hash FROM users WHERE id = ?",
        (user_id,),
    ).fetchone()
    if row is None:
        return False
    return bool(row["password_hash"])


def add_user(
    name: str,
    role: str,
    dept: str,
    std_hours: float,
    last_name: str = "",
    first_name: str = "",
) -> bool:
    """
    ユーザーを追加する。

    std_hours_am / std_hours_pm は互換性のため std_hours / 2 を設定する。

    Args:
        name: ユーザー表示名（姓＋名）
        role: ロール
        dept: 部署
        std_hours: 1日あたりの標準勤務時間
        last_name: 姓
        first_name: 名

    Returns:
        bool: 追加成功時True、重複時False
    """
    db = get_db()
    if not last_name:
        last_name = name
    try:
        max_order = db.execute(
            "SELECT COALESCE(MAX(display_order), 0) FROM users"
        ).fetchone()[0]
        db.execute(
            "INSERT INTO users (name, role, dept, std_hours_am, std_hours_pm, std_hours, "
            " display_order, last_name, first_name) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (name, role, dept, std_hours / 2, std_hours / 2, std_hours,
             max_order + 1, last_name, first_name),
        )
        db.commit()
        return True
    except Exception:
        db.rollback()
        return False


def update_user(
    user_id: int,
    name: str,
    role: str,
    dept: str,
    std_hours: float,
    last_name: str = "",
    first_name: str = "",
) -> bool:
    """ユーザーの全情報（名前・役職・部署・基本勤務時間・姓名）を更新する。

    名前が他ユーザーと重複する場合は False を返す。

    Args:
        user_id: 更新対象のユーザーID
        name: 新しいユーザー名
        role: 新しい役職
        dept: 新しい部署
        std_hours: 新しい1日あたりの基本勤務時間
        last_name: 姓
        first_name: 名

    Returns:
        bool: 更新成功時 True、名前重複などエラー時 False
    """
    db = get_db()
    if not last_name:
        last_name = name
    try:
        # 既存のAM/PM比率を維持
        row = db.execute(
            "SELECT std_hours, std_hours_am, std_hours_pm FROM users WHERE id = ?",
            (user_id,),
        ).fetchone()
        if row and row["std_hours"] and row["std_hours"] > 0:
            ratio_am = row["std_hours_am"] / row["std_hours"]
            new_am = round(std_hours * ratio_am, 3)
            new_pm = round(std_hours - new_am, 3)
        else:
            new_am = std_hours / 2
            new_pm = std_hours / 2
        db.execute(
            "UPDATE users SET name = ?, role = ?, dept = ?, "
            "std_hours = ?, std_hours_am = ?, std_hours_pm = ?, "
            "last_name = ?, first_name = ? WHERE id = ?",
            (name, role, dept, std_hours, new_am, new_pm,
             last_name, first_name, user_id),
        )
        db.commit()
        return True
    except Exception:
        db.rollback()
        return False


def delete_user(user_id: int) -> None:
    """
    ユーザーを削除する（関連データはCASCADE削除）。

    Args:
        user_id: 削除対象のユーザーID
    """
    db = get_db()
    db.execute("DELETE FROM users WHERE id = ?", (user_id,))
    db.commit()


def update_user_std_hours(
    user_id: int,
    std_hours: float,
) -> None:
    """
    ユーザーの標準勤務時間を更新する。

    既存のAM/PM比率を維持したまま合計を変更する。
    AM/PMが未設定（両方0）の場合のみ均等割りにフォールバックする。

    Args:
        user_id: ユーザーID
        std_hours: 1日あたりの標準勤務時間
    """
    db = get_db()
    row = db.execute(
        "SELECT std_hours, std_hours_am, std_hours_pm FROM users WHERE id = ?",
        (user_id,),
    ).fetchone()
    if row and row["std_hours"] and row["std_hours"] > 0:
        # 既存比率を維持
        ratio_am = row["std_hours_am"] / row["std_hours"]
        new_am = round(std_hours * ratio_am, 3)
        new_pm = round(std_hours - new_am, 3)
    else:
        new_am = std_hours / 2
        new_pm = std_hours / 2
    db.execute(
        "UPDATE users SET std_hours = ?, std_hours_am = ?, std_hours_pm = ? WHERE id = ?",
        (std_hours, new_am, new_pm, user_id),
    )
    db.commit()


# ---------------------------------------------------------------------------
# タスクマスター関連
# ---------------------------------------------------------------------------


def get_task_master(user_id: int) -> list[dict]:
    """ユーザーのタスクマスターを取得する。

    大区分・中区分名も JOIN して返す。
    display_order昇順、task_name昇順でソートする。

    Args:
        user_id: ユーザーID

    Returns:
        list[dict]: タスク情報のリスト（category_name / subcategory_name を含む）
    """
    db = get_db()
    rows = db.execute(
        "SELECT tm.id, tm.user_id, tm.task_name, tm.display_order, tm.default_hours,"
        " tm.category_id, tc.name AS category_name,"
        " tm.subcategory_id, ts.name AS subcategory_name"
        " FROM task_master tm"
        " LEFT JOIN task_category tc ON tm.category_id = tc.id"
        " LEFT JOIN task_subcategory ts ON tm.subcategory_id = ts.id"
        " WHERE tm.user_id = ?"
        " ORDER BY tm.display_order ASC, tm.task_name ASC",
        (user_id,),
    ).fetchall()
    return [dict(row) for row in rows]


def get_global_task_category_map() -> dict[str, dict[str, str]]:
    """全ユーザー横断でタスク名→区分情報のマップを取得する。

    同名タスクが複数ユーザーに登録されている場合、最初に見つかった
    区分情報を採用する（区分が設定されているもの優先）。

    Returns:
        dict[str, dict[str, str]]: {task_name: {"category_name": ..., "subcategory_name": ...}}
    """
    db = get_db()
    rows = db.execute(
        "SELECT DISTINCT tm.task_name,"
        " tc.name AS category_name,"
        " ts.name AS subcategory_name"
        " FROM task_master tm"
        " LEFT JOIN task_category tc ON tm.category_id = tc.id"
        " LEFT JOIN task_subcategory ts ON tm.subcategory_id = ts.id"
        " WHERE tc.name IS NOT NULL"
        " ORDER BY tm.task_name",
    ).fetchall()
    result: dict[str, dict[str, str]] = {}
    for row in rows:
        name = row["task_name"]
        if name not in result:
            result[name] = {
                "category_name": row["category_name"] or "",
                "subcategory_name": row["subcategory_name"] or "",
            }
    return result


def add_task(
    user_id: int,
    task_name: str,
    default_hours: float = 0.0,
    category_id: int | None = None,
    subcategory_id: int | None = None,
) -> bool:
    """タスクマスターにタスクを追加する。

    Args:
        user_id: ユーザーID
        task_name: タスク名
        default_hours: デフォルト作業時間（省略時は0.0）
        category_id: 大区分ID（省略時はNone）
        subcategory_id: 中区分ID（省略時はNone）

    Returns:
        bool: 追加成功時True、重複時False
    """
    db = get_db()
    try:
        db.execute(
            "INSERT INTO task_master (user_id, task_name, default_hours, category_id, subcategory_id)"
            " VALUES (?, ?, ?, ?, ?)",
            (user_id, task_name, default_hours, category_id or None, subcategory_id or None),
        )
        db.commit()
        return True
    except Exception:
        db.rollback()
        return False


def update_task_order(task_id: int, user_id: int, display_order: int) -> None:
    """タスクの表示順を更新する。

    Args:
        task_id: 更新対象のタスクID
        user_id: 所有者ユーザーID（所有確認用）
        display_order: 新しい表示順
    """
    db = get_db()
    db.execute(
        "UPDATE task_master SET display_order = ? WHERE id = ? AND user_id = ?",
        (display_order, task_id, user_id),
    )
    db.commit()


def delete_task(task_id: int, user_id: int) -> bool:
    """
    タスクマスターからタスクを削除する。

    user_idが一致する場合のみ削除する。

    Args:
        task_id: タスクID
        user_id: ユーザーID（所有確認用）

    Returns:
        bool: 削除成功時True、対象なし時False
    """
    db = get_db()
    cursor = db.execute(
        "DELETE FROM task_master WHERE id = ? AND user_id = ?",
        (task_id, user_id),
    )
    db.commit()
    return cursor.rowcount > 0


# ---------------------------------------------------------------------------
# 週間予定関連
# ---------------------------------------------------------------------------


def _empty_schedule() -> dict:
    """
    空の週間予定構造体を生成する。

    Returns:
        dict: {day(0-4): {'am': [{task_name:'', hours:0.0}×5], 'pm': [同]}}
    """
    return {
        day: {
            slot: [{"task_name": "", "hours": 0.0} for _ in range(_SLOT_SIZE)]
            for slot in _SLOTS
        }
        for day in _DAYS
    }


def get_weekly_schedule(user_id: int, week_start: str) -> dict:
    """
    指定ユーザー・週の週間予定を取得する。

    データが存在しない枠は task_name=''、hours=0.0 で埋める。

    Args:
        user_id: ユーザーID
        week_start: 週開始日（'YYYY-MM-DD'形式）

    Returns:
        dict: {day(0-4): {'am': [{task_name:str, hours:float}×5], 'pm': [同]}}
    """
    db = get_db()
    rows = db.execute(
        "SELECT day_of_week, time_slot, slot_index, task_name, hours, project_task_id "
        "FROM weekly_schedule "
        "WHERE user_id = ? AND week_start = ?",
        (user_id, week_start),
    ).fetchall()

    schedule = _empty_schedule()
    for row in rows:
        day = row["day_of_week"]
        slot = row["time_slot"]
        idx = row["slot_index"]
        if day in _DAYS and slot in _SLOTS and 0 <= idx < _SLOT_SIZE:
            schedule[day][slot][idx] = {
                "task_name": row["task_name"] or "",
                "hours": row["hours"] or 0.0,
                "project_task_id": row["project_task_id"],
            }
    return schedule


def save_weekly_schedule(
    user_id: int,
    week_start: str,
    data: dict,
    updated_by: str = "",
) -> None:
    """
    週間予定を保存する（UPSERT）。

    created_at は既存レコードがあれば維持し、新規なら現在時刻を設定する。
    updated_at は常に現在時刻で上書きする。

    Args:
        user_id: ユーザーID
        week_start: 週開始日（'YYYY-MM-DD'形式）
        data: {day: {'am': [{task_name:str, hours:float}×5], 'pm': [同]}}
        updated_by: 更新者名（省略可）
    """
    db = get_db()
    now = datetime.now().isoformat()

    # 既存レコードのcreated_at・project_task_idをキャッシュ
    existing_rows = db.execute(
        "SELECT day_of_week, time_slot, slot_index, created_at, project_task_id "
        "FROM weekly_schedule "
        "WHERE user_id = ? AND week_start = ?",
        (user_id, week_start),
    ).fetchall()
    created_at_map: dict[tuple[int, str, int], str] = {
        (r["day_of_week"], r["time_slot"], r["slot_index"]): r["created_at"]
        for r in existing_rows
    }
    pt_id_map: dict[tuple[int, str, int], int | None] = {
        (r["day_of_week"], r["time_slot"], r["slot_index"]): r["project_task_id"]
        for r in existing_rows
    }

    for day, slots in data.items():
        day_int = int(day)
        for slot, entries in slots.items():
            for idx, entry in enumerate(entries):
                key = (day_int, slot, idx)
                created_at = created_at_map.get(key) or now
                # project_task_id: task_nameが空なら必ずNULL、それ以外はフォーム値→既存値の順で採用
                if entry.get("task_name", "").strip():
                    pt_id = entry.get("project_task_id") or pt_id_map.get(key)
                else:
                    pt_id = None
                db.execute(
                    "INSERT OR REPLACE INTO weekly_schedule "
                    "(user_id, week_start, day_of_week, time_slot, slot_index, "
                    " task_name, hours, subcategory_name, project_task_id, "
                    " created_at, updated_at, updated_by) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (
                        user_id,
                        week_start,
                        day_int,
                        slot,
                        idx,
                        entry.get("task_name", ""),
                        entry.get("hours", 0.0),
                        entry.get("subcategory_name", ""),
                        pt_id,
                        created_at,
                        now,
                        updated_by,
                    ),
                )
    db.commit()


def get_weekly_schedule_meta(user_id: int, week_start: str) -> dict | None:
    """週間予定の最終更新日時・更新者を取得する。

    複数レコードがある場合は最も新しい updated_at を返す。

    Args:
        user_id: ユーザーID
        week_start: 週開始日（'YYYY-MM-DD'形式）

    Returns:
        dict | None: {'updated_at': str, 'updated_by': str} または None
    """
    db = get_db()
    row = db.execute(
        "SELECT updated_at, updated_by FROM weekly_schedule "
        "WHERE user_id = ? AND week_start = ? "
        "ORDER BY updated_at DESC LIMIT 1",
        (user_id, week_start),
    ).fetchone()
    return dict(row) if row else None


def copy_last_week_schedule(user_id: int, week_start: str) -> bool:
    """1週間前の週間予定を空き枠にのみコピーする。

    コピー先に既にデータがある枠（リスケ済みなど）は上書きしない。

    Args:
        user_id: ユーザーID
        week_start: コピー先の週開始日（'YYYY-MM-DD'形式）

    Returns:
        bool: コピー成功時True、コピー元データなし時False
    """
    from datetime import timedelta

    target_date = date.fromisoformat(week_start)
    last_week_start = (target_date - timedelta(days=7)).isoformat()

    last_week_data = get_weekly_schedule(user_id, last_week_start)

    # 全枠が空かどうか確認
    has_data = any(
        entry["task_name"] or entry["hours"] > 0.0
        for slots in last_week_data.values()
        for entries in slots.values()
        for entry in entries
    )
    if not has_data:
        return False

    # コピー先の既存データを取得し、データがある枠はスキップ
    current_data = get_weekly_schedule(user_id, week_start)
    merged_data: dict = {}
    for day, slots in last_week_data.items():
        merged_data[day] = {}
        for slot, entries in slots.items():
            merged_data[day][slot] = []
            for idx, entry in enumerate(entries):
                cur_entry = current_data[day][slot][idx]
                if cur_entry["task_name"] or cur_entry.get("hours", 0.0) > 0.0:
                    # コピー先に既存データあり → 既存データを維持
                    merged_data[day][slot].append(cur_entry)
                else:
                    # 空き枠 → 前週データをコピー
                    merged_data[day][slot].append(entry)

    save_weekly_schedule(user_id, week_start, merged_data)
    return True


def get_all_users_schedule_status(
    week_start: str, dept_filter: str | None = None
) -> list[dict]:
    """全ユーザーの週間予定の入力状況を取得する。

    Args:
        week_start: 週開始日（'YYYY-MM-DD'形式）
        dept_filter: 部署名でフィルタリングする場合に指定。None の場合は全部署。

    Returns:
        list[dict]: [{id, name, dept, role, std_hours_am, std_hours_pm, std_hours,
                      has_schedule:bool, filled_slots:int}]
                    filled_slots は hours > 0 の枠数
    """
    db = get_db()
    if dept_filter:
        users = db.execute(
            "SELECT id, name, dept, role, std_hours_am, std_hours_pm, std_hours "
            "FROM users WHERE dept = ? ORDER BY display_order ASC, id ASC",
            (dept_filter,),
        ).fetchall()
    else:
        users = db.execute(
            "SELECT id, name, dept, role, std_hours_am, std_hours_pm, std_hours "
            "FROM users ORDER BY display_order ASC, id ASC"
        ).fetchall()

    result: list[dict] = []
    for user in users:
        user_id = user["id"]
        schedule_rows = db.execute(
            "SELECT COUNT(*) AS cnt, "
            "SUM(CASE WHEN hours > 0.0 THEN 1 ELSE 0 END) AS filled "
            "FROM weekly_schedule "
            "WHERE user_id = ? AND week_start = ?",
            (user_id, week_start),
        ).fetchone()

        total = schedule_rows["cnt"] or 0
        filled = schedule_rows["filled"] or 0

        result.append(
            {
                "id": user["id"],
                "name": user["name"],
                "dept": user["dept"],
                "role": user["role"],
                "std_hours_am": user["std_hours_am"],
                "std_hours_pm": user["std_hours_pm"],
                "std_hours": user["std_hours"] or 8.0,
                "has_schedule": total > 0,
                "filled_slots": int(filled),
            }
        )
    return result


# ---------------------------------------------------------------------------
# 日次実績関連
# ---------------------------------------------------------------------------


def _empty_daily_result() -> dict:
    """空の日次実績構造体を生成する。

    Returns:
        dict: {'am': [{task_name:'', hours:0.0}×5], 'pm': [同]}
    """
    return {
        slot: [{"task_name": "", "hours": 0.0, "defer_date": "", "is_carryover": 0} for _ in range(_SLOT_SIZE)]
        for slot in _SLOTS
    }


def get_daily_result(user_id: int, date_str: str) -> dict:
    """指定ユーザー・日付の日次実績を取得する。

    データが存在しない枠は task_name=''、hours=0.0 で埋める。

    Args:
        user_id: ユーザーID
        date_str: 日付（'YYYY-MM-DD'形式）

    Returns:
        dict: {'am': [{task_name:str, hours:float}×5], 'pm': [同]}
    """
    db = get_db()
    rows = db.execute(
        "SELECT time_slot, slot_index, task_name, hours, defer_date, is_carryover, project_task_id "
        "FROM daily_result "
        "WHERE user_id = ? AND date = ?",
        (user_id, date_str),
    ).fetchall()

    result = _empty_daily_result()
    for row in rows:
        slot = row["time_slot"]
        idx = row["slot_index"]
        if slot in _SLOTS and 0 <= idx < _SLOT_SIZE:
            result[slot][idx] = {
                "task_name": row["task_name"] or "",
                "hours": row["hours"] or 0.0,
                "defer_date": row["defer_date"] or "",
                "is_carryover": int(row["is_carryover"] or 0),
                "project_task_id": row["project_task_id"],
            }
    return result


def get_daily_result_meta(user_id: int, date_str: str) -> dict | None:
    """日次実績の最終更新日時・更新者を取得する。

    Args:
        user_id: ユーザーID
        date_str: 日付（'YYYY-MM-DD'形式）

    Returns:
        dict | None: {'updated_at': str, 'updated_by': str} または None
    """
    db = get_db()
    row = db.execute(
        "SELECT updated_at, updated_by FROM daily_result "
        "WHERE user_id = ? AND date = ? "
        "ORDER BY updated_at DESC LIMIT 1",
        (user_id, date_str),
    ).fetchone()
    return dict(row) if row else None


def save_daily_result(
    user_id: int,
    date_str: str,
    data: dict,
    updated_by: str = "",
) -> None:
    """日次実績を保存する（UPSERT）。

    Args:
        user_id: ユーザーID
        date_str: 日付（'YYYY-MM-DD'形式）
        data: {'am': [{task_name:str, hours:float}×5], 'pm': [同]}
        updated_by: 更新者名
    """
    import sqlite3 as _sqlite3

    db = get_db()
    now = datetime.now().isoformat()
    try:
        for slot, entries in data.items():
            for idx, entry in enumerate(entries):
                db.execute(
                    "INSERT OR REPLACE INTO daily_result "
                    "(user_id, date, time_slot, slot_index, task_name, hours, "
                    " subcategory_name, defer_date, is_carryover, project_task_id, "
                    " updated_at, updated_by) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (
                        user_id,
                        date_str,
                        slot,
                        idx,
                        entry.get("task_name", ""),
                        entry.get("hours", 0.0),
                        entry.get("subcategory_name", ""),
                        entry.get("defer_date", ""),
                        int(entry.get("is_carryover", 0)),
                        entry.get("project_task_id"),
                        now,
                        updated_by,
                    ),
                )
        db.commit()
    except _sqlite3.DatabaseError as exc:
        db.rollback()
        logger.error("日次実績の保存に失敗しました: %s", exc)
        raise


def get_daily_comment(user_id: int, date_str: str) -> dict:
    """日次コメント（振り返り・対策）を取得する。

    Args:
        user_id: ユーザーID
        date_str: 日付（'YYYY-MM-DD'形式）

    Returns:
        dict: {'reflection': str, 'action': str, 'updated_at': str, 'updated_by': str}
              レコードが存在しない場合は空文字列を返す。
    """
    db = get_db()
    row = db.execute(
        "SELECT reflection, action, admin_comment, updated_at, updated_by "
        "FROM daily_comment "
        "WHERE user_id = ? AND date = ?",
        (user_id, date_str),
    ).fetchone()
    if row:
        return {
            "reflection": row["reflection"] or "",
            "action": row["action"] or "",
            "admin_comment": row["admin_comment"] or "",
            "updated_at": row["updated_at"] or "",
            "updated_by": row["updated_by"] or "",
        }
    return {"reflection": "", "action": "", "admin_comment": "", "updated_at": "", "updated_by": ""}


def save_daily_comment(
    user_id: int,
    date_str: str,
    reflection: str,
    action: str,
    updated_by: str = "",
) -> None:
    """日次コメント（振り返り・対策）を保存する（UPSERT）。

    admin_comment は上書きしない（別関数で管理）。

    Args:
        user_id: ユーザーID
        date_str: 日付（'YYYY-MM-DD'形式）
        reflection: 本日の振り返り
        action: 今後の対策・懸念事項
        updated_by: 更新者名
    """
    import sqlite3 as _sqlite3

    db = get_db()
    now = datetime.now().isoformat()
    try:
        db.execute(
            "INSERT INTO daily_comment "
            "(user_id, date, reflection, action, updated_at, updated_by) "
            "VALUES (?, ?, ?, ?, ?, ?) "
            "ON CONFLICT(user_id, date) DO UPDATE SET "
            "reflection=excluded.reflection, action=excluded.action, "
            "updated_at=excluded.updated_at, updated_by=excluded.updated_by",
            (user_id, date_str, reflection, action, now, updated_by),
        )
        db.commit()
    except _sqlite3.DatabaseError as exc:
        db.rollback()
        logger.error("日次コメントの保存に失敗しました: %s", exc)
        raise


def save_admin_comment(
    user_id: int,
    date_str: str,
    admin_comment: str,
    updated_by: str = "",
) -> None:
    """管理者（上長）コメントを保存する（UPSERT）。

    ユーザー本人のコメント（reflection/action）は上書きしない。

    Args:
        user_id: コメント対象ユーザーID
        date_str: 日付（'YYYY-MM-DD'形式）
        admin_comment: 管理者コメント
        updated_by: 更新者名（管理者名）
    """
    import sqlite3 as _sqlite3

    db = get_db()
    now = datetime.now().isoformat()
    try:
        db.execute(
            "INSERT INTO daily_comment "
            "(user_id, date, reflection, action, admin_comment, updated_at, updated_by) "
            "VALUES (?, ?, '', '', ?, ?, ?) "
            "ON CONFLICT(user_id, date) DO UPDATE SET "
            "admin_comment=excluded.admin_comment, "
            "updated_at=excluded.updated_at, updated_by=excluded.updated_by",
            (user_id, date_str, admin_comment, now, updated_by),
        )
        db.commit()
    except _sqlite3.DatabaseError as exc:
        db.rollback()
        logger.error("管理者コメントの保存に失敗しました: %s", exc)
        raise


# ---------------------------------------------------------------------------
# 週間休暇設定関連
# ---------------------------------------------------------------------------


def get_weekly_leave(user_id: int, week_start: str) -> dict:
    """週間休暇設定を取得する。

    Args:
        user_id: ユーザーID
        week_start: 週開始日（'YYYY-MM-DD'形式）

    Returns:
        dict[int, str]: {day_of_week(0-4): leave_type}
                        未設定の曜日は '' を返す。
    """
    db = get_db()
    rows = db.execute(
        "SELECT day_of_week, leave_type FROM weekly_leave "
        "WHERE user_id = ? AND week_start = ?",
        (user_id, week_start),
    ).fetchall()
    result: dict[int, str] = {i: "" for i in range(5)}
    for row in rows:
        day = row["day_of_week"]
        if 0 <= day <= 4:
            result[day] = row["leave_type"] or ""
    return result


def save_weekly_leave(user_id: int, week_start: str, leave_data: dict) -> None:
    """週間休暇設定を保存する（UPSERT）。

    leave_type が空文字の場合はレコードを削除する。

    Args:
        user_id: ユーザーID
        week_start: 週開始日（'YYYY-MM-DD'形式）
        leave_data: {day_of_week(int): leave_type(str)}
    """
    import sqlite3 as _sqlite3

    db = get_db()
    try:
        for day, leave_type in leave_data.items():
            day_int = int(day)
            if leave_type:
                db.execute(
                    "INSERT OR REPLACE INTO weekly_leave "
                    "(user_id, week_start, day_of_week, leave_type) "
                    "VALUES (?, ?, ?, ?)",
                    (user_id, week_start, day_int, leave_type),
                )
            else:
                db.execute(
                    "DELETE FROM weekly_leave "
                    "WHERE user_id = ? AND week_start = ? AND day_of_week = ?",
                    (user_id, week_start, day_int),
                )
        db.commit()
    except _sqlite3.DatabaseError as exc:
        db.rollback()
        logger.error("週間休暇設定の保存に失敗しました: %s", exc)
        raise


def remove_rescheduled_task(user_id: int, target_date_str: str, task_name: str) -> None:
    """リスケ解除時に、振り替え先の週間予定からタスクをクリアする。

    Args:
        user_id: ユーザーID
        target_date_str: リスケ先の日付（'YYYY-MM-DD'形式）
        task_name: 削除するタスク名（例: 「【03/20 リスケ】作業名」）
    """
    target_date = date.fromisoformat(target_date_str)
    day_of_week = target_date.weekday()
    week_start = (target_date - timedelta(days=day_of_week)).isoformat()
    db = get_db()
    db.execute(
        "UPDATE weekly_schedule SET task_name = '', hours = 0 "
        "WHERE user_id = ? AND week_start = ? AND day_of_week = ? AND task_name = ?",
        (user_id, week_start, day_of_week, task_name),
    )
    db.commit()
    logger.info("リスケ解除: weekly_schedule から削除 user=%d task=%s date=%s", user_id, task_name, target_date_str)


def remove_rescheduled_daily_result(user_id: int, target_date_str: str, task_name: str) -> None:
    """リスケ解除時に、振り替え先の日次実績からタスクをクリアする。

    Args:
        user_id: ユーザーID
        target_date_str: リスケ先の日付（'YYYY-MM-DD'形式）
        task_name: クリアするタスク名（例: 「【03/20 リスケ】作業名」）
    """
    db = get_db()
    db.execute(
        "UPDATE daily_result SET task_name = '', hours = 0, defer_date = '', is_carryover = 0 "
        "WHERE user_id = ? AND date = ? AND task_name = ?",
        (user_id, target_date_str, task_name),
    )
    db.commit()
    logger.info("リスケ解除: daily_result からクリア user=%d task=%s date=%s", user_id, task_name, target_date_str)


def defer_task_to_weekly_schedule(
    user_id: int,
    target_date_str: str,
    task_name: str,
    hours: float,
    updated_by: str = "",
) -> None:
    """指定日の週間予定に作業を追加する（後日対応用）。

    空きスロット（task_nameが空のスロット）を午前→午後の順で探して挿入する。
    全スロットが埋まっている場合はスキップする。

    Args:
        user_id: ユーザーID
        target_date_str: 後日対応日（'YYYY-MM-DD'形式）
        task_name: 作業名
        hours: 予定時間
        updated_by: 更新者名
    """
    target_date = date.fromisoformat(target_date_str)
    day_of_week = target_date.weekday()  # 0=月〜4=金
    week_start = (target_date - timedelta(days=day_of_week)).isoformat()

    db = get_db()
    now = datetime.now().isoformat()

    # 同名タスクが対象日の週間予定に既に登録済みであればスキップ（重複防止）
    already = db.execute(
        "SELECT id FROM weekly_schedule "
        "WHERE user_id=? AND week_start=? AND day_of_week=? AND task_name=?",
        (user_id, week_start, day_of_week, task_name),
    ).fetchone()
    if already:
        logger.info(
            "後日対応: 既に登録済みのためスキップ user=%d task=%s date=%s",
            user_id, task_name, target_date_str,
        )
        return

    for slot in ("am", "pm"):
        for idx in range(5):
            existing = db.execute(
                "SELECT task_name FROM weekly_schedule "
                "WHERE user_id=? AND week_start=? AND day_of_week=? AND time_slot=? AND slot_index=?",
                (user_id, week_start, day_of_week, slot, idx),
            ).fetchone()
            if existing is None or not existing["task_name"]:
                db.execute(
                    "INSERT OR REPLACE INTO weekly_schedule "
                    "(user_id, week_start, day_of_week, time_slot, slot_index, "
                    " task_name, hours, created_at, updated_at, updated_by) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (user_id, week_start, day_of_week, slot, idx,
                     task_name, hours, now, now, updated_by),
                )
                db.commit()
                logger.info(
                    "後日対応: user=%d task=%s → %s %s[%d]",
                    user_id, task_name, target_date_str, slot, idx,
                )
                return
    logger.warning(
        "後日対応: 空きスロットなし user=%d date=%s task=%s",
        user_id, target_date_str, task_name,
    )


def get_week_daily_results(user_id: int, week_dates: list[str]) -> dict[int, dict]:
    """週の各日の日次実績を一括取得する。

    Args:
        user_id: ユーザーID
        week_dates: 月〜金の日付文字列リスト（長さ5）

    Returns:
        dict: {day_index(0-4): {'am': [...], 'pm': [...], 'has_result': bool}}
    """
    result: dict[int, dict] = {}
    for i, date_str in enumerate(week_dates):
        data = get_daily_result(user_id, date_str)
        has_result = any(
            entry["task_name"] or entry["hours"] > 0
            for slot in data.values()
            for entry in slot
        )
        result[i] = {**data, "has_result": has_result}
    return result


def get_all_users_daily_status(
    date_str: str, dept_filter: str | None = None
) -> list[dict]:
    """全ユーザーの日次実績入力状況を取得する。

    Args:
        date_str: 日付（'YYYY-MM-DD'形式）
        dept_filter: 部署名でフィルタリングする場合に指定。None の場合は全部署。

    Returns:
        list[dict]: [{id, name, dept, role, has_result:bool, filled_slots:int, updated_at:str|None}]
                    filled_slots は hours > 0 の枠数、updated_at は最終更新日時
    """
    db = get_db()
    if dept_filter:
        users = db.execute(
            "SELECT id, name, dept, role FROM users WHERE dept = ? ORDER BY display_order ASC, id ASC",
            (dept_filter,),
        ).fetchall()
    else:
        users = db.execute(
            "SELECT id, name, dept, role FROM users ORDER BY display_order ASC, id ASC"
        ).fetchall()

    result: list[dict] = []
    for user in users:
        row = db.execute(
            "SELECT COUNT(*) AS cnt, "
            "SUM(CASE WHEN hours > 0.0 THEN 1 ELSE 0 END) AS filled, "
            "MAX(updated_at) AS last_updated "
            "FROM daily_result "
            "WHERE user_id = ? AND date = ?",
            (user["id"], date_str),
        ).fetchone()
        total = row["cnt"] or 0
        filled = row["filled"] or 0
        result.append(
            {
                "id": user["id"],
                "name": user["name"],
                "dept": user["dept"],
                "role": user["role"],
                "has_result": total > 0,
                "filled_slots": int(filled),
                "updated_at": row["last_updated"],
            }
        )
    return result


def get_pending_carryovers(user_id: int) -> list[dict]:
    """保留中の繰越タスク一覧を取得する。

    Args:
        user_id: ユーザーID

    Returns:
        list[dict]: [{id, task_name, original_date, planned_hours}] 日付昇順
    """
    db = get_db()
    rows = db.execute(
        "SELECT id, task_name, original_date, planned_hours "
        "FROM carryover "
        "WHERE user_id = ? AND resolved = 0 "
        "ORDER BY original_date ASC, task_name ASC",
        (user_id,),
    ).fetchall()
    return [dict(row) for row in rows]


def add_carryover(
    user_id: int, original_date: str, task_name: str, planned_hours: float
) -> None:
    """繰越タスクを登録する。既存レコードが解決済みの場合は未解決に戻す。

    Args:
        user_id: ユーザーID
        original_date: 予定日（'YYYY-MM-DD'形式）
        task_name: タスク名
        planned_hours: 予定時間
    """
    now = datetime.now().isoformat()
    db = get_db()
    db.execute(
        "INSERT OR IGNORE INTO carryover "
        "(user_id, task_name, original_date, planned_hours, created_at) "
        "VALUES (?, ?, ?, ?, ?)",
        (user_id, task_name, original_date, planned_hours, now),
    )
    # 既存レコードが解決済みだった場合も未解決に戻す
    db.execute(
        "UPDATE carryover SET resolved = 0, planned_hours = ? "
        "WHERE user_id = ? AND task_name = ? AND original_date = ?",
        (planned_hours, user_id, task_name, original_date),
    )
    db.commit()


def resolve_carryover_by_slot(
    user_id: int, original_date: str, time_slot: str, slot_index: int
) -> None:
    """指定スロットの繰越タスクを解決済みにする（daily_result の is_carryover と連動）。

    Args:
        user_id: ユーザーID
        original_date: 対象日付（'YYYY-MM-DD'形式）
        time_slot: スロット ('am' または 'pm')
        slot_index: スロット内インデックス（0〜4）
    """
    db = get_db()
    # 対象スロットのタスク名を取得
    row = db.execute(
        "SELECT task_name FROM daily_result "
        "WHERE user_id=? AND date=? AND time_slot=? AND slot_index=?",
        (user_id, original_date, time_slot, slot_index),
    ).fetchone()
    if not row or not row["task_name"]:
        return
    db.execute(
        "UPDATE carryover SET resolved = 1 "
        "WHERE user_id=? AND task_name=? AND original_date=? AND resolved=0",
        (user_id, row["task_name"], original_date),
    )
    db.commit()


# ---------------------------------------------------------------------------
# 部署マスタ関連
# ---------------------------------------------------------------------------

def get_all_depts() -> list[dict]:
    """部署マスタの全件を display_order 昇順で取得する。

    Returns:
        list[dict]: [{id, dept_name, display_order}]
    """
    db = get_db()
    rows = db.execute(
        "SELECT id, dept_name, display_order FROM dept_master ORDER BY display_order ASC, id ASC"
    ).fetchall()
    return [dict(row) for row in rows]


def add_dept(dept_name: str, display_order: int = 0) -> None:
    """部署マスタに新しい部署を登録する。

    Args:
        dept_name: 部署名
        display_order: 表示順
    """
    db = get_db()
    db.execute(
        "INSERT INTO dept_master (dept_name, display_order) VALUES (?, ?)",
        (dept_name.strip(), display_order),
    )
    db.commit()


def update_dept(dept_id: int, dept_name: str, display_order: int = 0) -> None:
    """部署マスタの部署名・表示順を更新する。

    Args:
        dept_id: 更新対象の部署ID
        dept_name: 新しい部署名
        display_order: 表示順
    """
    db = get_db()
    db.execute(
        "UPDATE dept_master SET dept_name = ?, display_order = ? WHERE id = ?",
        (dept_name.strip(), display_order, dept_id),
    )
    db.commit()


def delete_dept(dept_id: int) -> bool:
    """部署マスタから部署を削除する。所属ユーザーがいれば削除不可。

    Args:
        dept_id: 削除対象の部署ID

    Returns:
        bool: 削除成功なら True、所属ユーザーがいれば False。
    """
    db = get_db()
    row = db.execute("SELECT dept_name FROM dept_master WHERE id = ?", (dept_id,)).fetchone()
    if not row:
        return False
    dept_name = row["dept_name"]
    count = db.execute(
        "SELECT COUNT(*) FROM users WHERE dept = ?", (dept_name,)
    ).fetchone()[0]
    if count > 0:
        return False
    db.execute("DELETE FROM dept_master WHERE id = ?", (dept_id,))
    db.commit()
    return True


def dept_has_users(dept_id: int) -> bool:
    """指定部署にユーザーが所属しているか確認する。

    Args:
        dept_id: 部署ID

    Returns:
        bool: 所属ユーザーがいれば True。
    """
    db = get_db()
    row = db.execute("SELECT dept_name FROM dept_master WHERE id = ?", (dept_id,)).fetchone()
    if not row:
        return False
    count = db.execute(
        "SELECT COUNT(*) FROM users WHERE dept = ?", (row["dept_name"],)
    ).fetchone()[0]
    return count > 0


def resolve_carryover_by_id(user_id: int, carryover_id: int) -> None:
    """指定IDの繰越タスクを解決済みにし、daily_result の is_carryover もリセットする。

    Args:
        user_id: ユーザーID（権限確認用）
        carryover_id: 繰越レコードのID
    """
    db = get_db()
    row = db.execute(
        "SELECT original_date, task_name FROM carryover WHERE id = ? AND user_id = ?",
        (carryover_id, user_id),
    ).fetchone()
    if row:
        # daily_result の対象スロットの is_carryover を 0 にリセット
        db.execute(
            "UPDATE daily_result SET is_carryover = 0 "
            "WHERE user_id = ? AND date = ? AND task_name = ?",
            (user_id, row["original_date"], row["task_name"]),
        )
    db.execute(
        "UPDATE carryover SET resolved = 1 WHERE id = ? AND user_id = ?",
        (carryover_id, user_id),
    )
    db.commit()


def resolve_carryovers_by_task(user_id: int, task_name: str, original_date: str = "") -> None:
    """タスク名（と日付）が一致する保留中の繰越を解決済みにする。

    Args:
        user_id: ユーザーID
        task_name: 解決するタスク名
        original_date: 対象日付（指定時はその日付のものだけ解決、省略時は全日付）
    """
    db = get_db()
    if original_date:
        db.execute(
            "UPDATE carryover SET resolved = 1 "
            "WHERE user_id = ? AND task_name = ? AND original_date = ? AND resolved = 0",
            (user_id, task_name, original_date),
        )
    else:
        db.execute(
            "UPDATE carryover SET resolved = 1 "
            "WHERE user_id = ? AND task_name = ? AND resolved = 0",
            (user_id, task_name),
        )
    db.commit()


def set_remember_token(user_id: int) -> str:
    """ログイン記憶用のセキュアトークンを生成してDBに保存する。

    Args:
        user_id: ユーザーID

    Returns:
        生成したトークン文字列（64文字の16進数）
    """
    token = secrets.token_hex(32)
    expiry = (datetime.now() + timedelta(days=365)).isoformat()
    db = get_db()
    db.execute(
        "UPDATE users SET remember_token = ?, remember_token_expiry = ? WHERE id = ?",
        (token, expiry, user_id),
    )
    db.commit()
    return token


def get_user_by_remember_token(token: str) -> dict | None:
    """記憶トークンからユーザーを取得する。期限切れは無効とする。

    Args:
        token: クッキーから取得したトークン文字列

    Returns:
        有効なトークンに対応するユーザー辞書、無効・期限切れは None
    """
    if not token:
        return None
    db = get_db()
    row = db.execute(
        "SELECT * FROM users WHERE remember_token = ?", (token,)
    ).fetchone()
    if row is None:
        return None
    expiry: str = row["remember_token_expiry"] or ""
    if expiry and datetime.fromisoformat(expiry) < datetime.now():
        return None
    return dict(row)


def clear_remember_token(user_id: int) -> None:
    """ユーザーの記憶トークンを削除する（ログアウト時に呼び出す）。

    Args:
        user_id: ユーザーID
    """
    db = get_db()
    db.execute(
        "UPDATE users SET remember_token = '', remember_token_expiry = '' WHERE id = ?",
        (user_id,),
    )
    db.commit()


def get_operation_logs(
    limit: int = 50,
    offset: int = 0,
    action_type: str = "",
) -> list[dict]:
    """操作ログを取得する（管理者画面用）。

    Args:
        limit: 取得件数。
        offset: 取得開始位置。
        action_type: アクション種別フィルタ（空文字なら全件）。

    Returns:
        list[dict]: ログレコードのリスト。
    """
    db = get_db()
    if action_type:
        rows = db.execute(
            "SELECT id, user_id, user_name, action_type, detail, ip_address, created_at"
            " FROM operation_log WHERE action_type = ?"
            " ORDER BY id DESC LIMIT ? OFFSET ?",
            (action_type, limit, offset),
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT id, user_id, user_name, action_type, detail, ip_address, created_at"
            " FROM operation_log ORDER BY id DESC LIMIT ? OFFSET ?",
            (limit, offset),
        ).fetchall()
    return [dict(r) for r in rows]


def count_operation_logs(action_type: str = "") -> int:
    """操作ログの総件数を返す。

    Args:
        action_type: アクション種別フィルタ（空文字なら全件）。

    Returns:
        int: 件数。
    """
    db = get_db()
    if action_type:
        row = db.execute(
            "SELECT COUNT(*) FROM operation_log WHERE action_type = ?",
            (action_type,),
        ).fetchone()
    else:
        row = db.execute("SELECT COUNT(*) FROM operation_log").fetchone()
    return row[0] if row else 0


# ---------------------------------------------------------------------------
# 作業区分（大区分・中区分）関連
# ---------------------------------------------------------------------------


def get_all_categories() -> list[dict]:
    """全大区分を取得する（display_order昇順）。

    Returns:
        list[dict]: 大区分情報のリスト（id, name, display_order）
    """
    db = get_db()
    rows = db.execute(
        "SELECT id, name, display_order FROM task_category ORDER BY display_order ASC, id ASC"
    ).fetchall()
    return [dict(row) for row in rows]


def get_subcategories(category_id: int) -> list[dict]:
    """指定大区分の中区分一覧を取得する（display_order昇順）。

    Args:
        category_id: 大区分ID

    Returns:
        list[dict]: 中区分情報のリスト（id, category_id, name, display_order）
    """
    db = get_db()
    rows = db.execute(
        "SELECT id, category_id, name, display_order FROM task_subcategory"
        " WHERE category_id = ? ORDER BY display_order ASC, id ASC",
        (category_id,),
    ).fetchall()
    return [dict(row) for row in rows]


def get_all_subcategories() -> list[dict]:
    """全中区分を取得する（category_id, display_order昇順）。カテゴリ名も含む。

    task_category と JOIN し、各行に category_id, category_name, id, name,
    display_order を含む。

    Returns:
        list[dict]: 中区分情報のリスト
    """
    db = get_db()
    rows = db.execute(
        "SELECT s.id, s.category_id, c.name AS category_name,"
        " s.name, s.display_order"
        " FROM task_subcategory s"
        " JOIN task_category c ON s.category_id = c.id"
        " ORDER BY s.category_id ASC, s.display_order ASC, s.id ASC"
    ).fetchall()
    return [dict(row) for row in rows]


def add_category(name: str) -> bool:
    """大区分を追加する。重複時はFalseを返す。

    Args:
        name: 大区分名

    Returns:
        bool: 追加成功時True、重複時False
    """
    db = get_db()
    try:
        db.execute(
            "INSERT INTO task_category (name) VALUES (?)",
            (name,),
        )
        db.commit()
        return True
    except sqlite3.IntegrityError:
        db.rollback()
        return False


def add_subcategory(category_id: int, name: str) -> bool:
    """中区分を追加する。重複時はFalseを返す。

    Args:
        category_id: 所属する大区分ID
        name: 中区分名

    Returns:
        bool: 追加成功時True、重複時False
    """
    db = get_db()
    try:
        db.execute(
            "INSERT INTO task_subcategory (category_id, name) VALUES (?, ?)",
            (category_id, name),
        )
        db.commit()
        return True
    except sqlite3.IntegrityError:
        db.rollback()
        return False


def delete_category(category_id: int) -> None:
    """大区分を削除する（中区分もCASCADE削除）。

    Args:
        category_id: 削除対象の大区分ID
    """
    db = get_db()
    db.execute("DELETE FROM task_category WHERE id = ?", (category_id,))
    db.commit()


def update_category_order(order_list: list[int]) -> None:
    """大区分の表示順を一括更新する。

    Args:
        order_list: 大区分IDのリスト（表示順）
    """
    db = get_db()
    for idx, cat_id in enumerate(order_list):
        db.execute(
            "UPDATE task_category SET display_order = ? WHERE id = ?",
            (idx, cat_id),
        )
    db.commit()


def update_subcategory_order(order_list: list[int]) -> None:
    """中区分の表示順を一括更新する。

    Args:
        order_list: 中区分IDのリスト（表示順）
    """
    db = get_db()
    for idx, sub_id in enumerate(order_list):
        db.execute(
            "UPDATE task_subcategory SET display_order = ? WHERE id = ?",
            (idx, sub_id),
        )
    db.commit()


def delete_subcategory(subcategory_id: int) -> None:
    """中区分を削除する。

    Args:
        subcategory_id: 削除対象の中区分ID
    """
    db = get_db()
    db.execute("DELETE FROM task_subcategory WHERE id = ?", (subcategory_id,))
    db.commit()


def update_category_name(category_id: int, name: str) -> bool:
    """大区分名を更新する。重複時はFalseを返す。

    Args:
        category_id: 更新対象の大区分ID
        name: 新しい大区分名

    Returns:
        bool: 更新成功時True、重複時False
    """
    db = get_db()
    try:
        db.execute(
            "UPDATE task_category SET name = ? WHERE id = ?",
            (name, category_id),
        )
        db.commit()
        return True
    except sqlite3.IntegrityError:
        db.rollback()
        return False


def update_subcategory_name(subcategory_id: int, name: str) -> bool:
    """中区分名を更新する。重複時はFalseを返す。

    Args:
        subcategory_id: 更新対象の中区分ID
        name: 新しい中区分名

    Returns:
        bool: 更新成功時True、重複時False
    """
    db = get_db()
    try:
        db.execute(
            "UPDATE task_subcategory SET name = ? WHERE id = ?",
            (name, subcategory_id),
        )
        db.commit()
        return True
    except sqlite3.IntegrityError:
        db.rollback()
        return False


# ---------------------------------------------------------------------------
# メール設定関連
# ---------------------------------------------------------------------------


def get_mail_setting(role: str) -> dict:
    """指定役職のメール設定を取得する。

    Args:
        role: 役職名（'管理職' または 'マスタ'）

    Returns:
        dict: {role, to_address, cc_address, bcc_address, subject_template, body_template}
              レコードが存在しない場合は空文字列を返す。
    """
    db = get_db()
    row = db.execute(
        "SELECT role, to_address, cc_address, bcc_address, subject_template, body_template"
        " FROM mail_settings WHERE role = ?",
        (role,),
    ).fetchone()
    if row:
        return dict(row)
    return {
        "role": role,
        "to_address": "",
        "cc_address": "",
        "bcc_address": "",
        "subject_template": "",
        "body_template": "",
    }


def save_mail_setting(
    role: str,
    to_address: str,
    cc_address: str,
    subject_template: str,
    body_template: str,
    bcc_address: str = "",
) -> None:
    """指定役職のメール設定を保存する（INSERT OR REPLACE）。

    Args:
        role: 役職名（'管理職' または 'マスタ'）
        to_address: TO宛先（複数はセミコロン区切り）
        cc_address: CC宛先（複数はセミコロン区切り）
        subject_template: 件名テンプレート
        body_template: 本文テンプレート
        bcc_address: BCC宛先（複数はセミコロン区切り）
    """
    db = get_db()
    db.execute(
        "INSERT OR REPLACE INTO mail_settings"
        " (role, to_address, cc_address, bcc_address, subject_template, body_template)"
        " VALUES (?, ?, ?, ?, ?, ?)",
        (role, to_address, cc_address, bcc_address, subject_template, body_template),
    )
    db.commit()


# ---------------------------------------------------------------------------
# プロジェクトタスク管理
# ---------------------------------------------------------------------------

# 状態の選択肢（表示順）
PROJECT_TASK_STATUSES: list[str] = [
    "未着手", "着手", "順調", "遅れ", "完了", "停止",
]


def _normalize_progress(
    status: str,
    progress: int,
    start_date: str | None = None,
    end_date: str | None = None,
) -> int:
    """状態に応じた進捗率の正規化。

    未着手→0、完了→100 のみ強制。
    「順調」の場合は開始日・終了日・今日の日付から自動計算する。
    それ以外はユーザー入力値をそのまま返す。

    Args:
        status: タスク状態
        progress: ユーザーが入力した進捗率
        start_date: 開始日（YYYY-MM-DD）。順調時の自動計算に使用。
        end_date: 終了日（YYYY-MM-DD）。順調時の自動計算に使用。

    Returns:
        int: 正規化された進捗率
    """
    if status == "未着手":
        return 0
    if status == "完了":
        return 100
    if status == "順調" and start_date and end_date:
        return _calc_progress_by_date(start_date, end_date)
    return max(0, progress)


def _calc_progress_by_date(start_date: str, end_date: str) -> int:
    """開始日・終了日・今日の日付から進捗率を自動計算する。

    計算式: (今日 - 開始日) / (終了日 - 開始日) × 100
    開始日前は0%、終了日以降は100%にクランプする。

    Args:
        start_date: 開始日（YYYY-MM-DD）
        end_date: 終了日（YYYY-MM-DD）

    Returns:
        int: 0〜100の進捗率
    """
    try:
        s = date.fromisoformat(start_date)
        e = date.fromisoformat(end_date)
    except (ValueError, TypeError):
        return 0
    today_d = date.today()
    if today_d <= s:
        return 0
    if today_d >= e:
        return 100
    total_days = (e - s).days
    if total_days <= 0:
        return 0
    elapsed = (today_d - s).days
    return max(0, min(100, round(elapsed / total_days * 100)))


def get_all_project_tasks(
    assigned_to: int | None = None,
    user_ids: list[int] | None = None,
) -> list[dict]:
    """プロジェクトタスクを大区分・中区分の表示順で取得する。

    Args:
        assigned_to: 指定時はそのユーザーに割り当てられたタスクのみ返す。
        user_ids: 指定時はそのユーザー群のいずれかに割り当てられたタスクを返す。
                  None の場合は全タスクを返す。

    Returns:
        list[dict]: プロジェクトタスク一覧
    """
    db = get_db()
    query = (
        "SELECT pt.*, "
        "  tc.name AS category_name, tc.display_order AS cat_order, "
        "  ts.name AS subcategory_name, ts.display_order AS subcat_order, "
        "  u1.name AS assigned_name, u1.last_name AS assigned_last_name, "
        "  u2.name AS assigned_name_2, u2.last_name AS assigned_last_name_2 "
        "FROM project_task pt "
        "LEFT JOIN task_category tc ON pt.category_id = tc.id "
        "LEFT JOIN task_subcategory ts ON pt.subcategory_id = ts.id "
        "LEFT JOIN users u1 ON pt.assigned_to = u1.id "
        "LEFT JOIN users u2 ON pt.assigned_to_2 = u2.id "
    )
    params: tuple = ()
    if assigned_to is not None:
        query += "WHERE (pt.assigned_to = ? OR pt.assigned_to_2 = ?) "
        params = (assigned_to, assigned_to)
    elif user_ids is not None and len(user_ids) > 0:
        placeholders = ",".join("?" * len(user_ids))
        query += f"WHERE (pt.assigned_to IN ({placeholders}) OR pt.assigned_to_2 IN ({placeholders})) "
        params = tuple(user_ids) * 2
    query += "ORDER BY tc.display_order, ts.display_order, pt.display_order"
    rows = db.execute(query, params).fetchall()
    return [dict(r) for r in rows]


def get_project_task_by_id(task_id: int) -> dict | None:
    """IDでプロジェクトタスクを取得する。

    Args:
        task_id: タスクID

    Returns:
        dict | None: タスク情報。見つからなければ None
    """
    db = get_db()
    row = db.execute(
        "SELECT pt.*, "
        "  tc.name AS category_name, ts.name AS subcategory_name, "
        "  u1.name AS assigned_name, u1.last_name AS assigned_last_name, "
        "  u2.name AS assigned_name_2, u2.last_name AS assigned_last_name_2 "
        "FROM project_task pt "
        "LEFT JOIN task_category tc ON pt.category_id = tc.id "
        "LEFT JOIN task_subcategory ts ON pt.subcategory_id = ts.id "
        "LEFT JOIN users u1 ON pt.assigned_to = u1.id "
        "LEFT JOIN users u2 ON pt.assigned_to_2 = u2.id "
        "WHERE pt.id = ?",
        (task_id,),
    ).fetchone()
    if not row:
        return None
    return dict(row)


def add_project_task(
    category_id: int | None,
    subcategory_id: int | None,
    task_name: str,
    description: str,
    start_date: str,
    end_date: str,
    status: str,
    progress: int,
    delay_days: int,
    created_by: int,
    updated_by: str,
    assigned_to: int | None = None,
    assigned_to_2: int | None = None,
    is_milestone: int = 0,
    is_event: int = 0,
    event_start_time: str = "",
    event_end_time: str = "",
    planned_hours: float = 0.0,
) -> int:
    """プロジェクトタスクを追加する。

    Args:
        category_id: 大区分ID
        subcategory_id: 中区分ID
        task_name: タスク名
        description: 対応内容
        start_date: 開始日
        end_date: 終了日
        status: 状態
        progress: 進捗率（手動指定時）
        delay_days: 遅延日数
        created_by: 作成者ユーザーID
        updated_by: 更新者名
        assigned_to: 担当者1ユーザーID
        assigned_to_2: 担当者2ユーザーID
        is_milestone: マイルストーンフラグ（1=マイルストーン）
        is_event: イベントフラグ（1=イベント/会議）
        event_start_time: イベント開始時刻（HH:MM形式）
        event_end_time: イベント終了時刻（HH:MM形式）
        planned_hours: 予定工数（時間）

    Returns:
        int: 新規タスクID
    """
    db = get_db()
    calc_progress = _normalize_progress(status, progress, start_date, end_date)
    # マイルストーン・イベント以外で予定工数未設定なら期間から自動計算
    if planned_hours <= 0 and not is_milestone and not is_event:
        planned_hours = calc_planned_hours(start_date, end_date)
    max_order = db.execute("SELECT COALESCE(MAX(display_order), 0) FROM project_task").fetchone()[0]
    cur = db.execute(
        "INSERT INTO project_task "
        "(category_id, subcategory_id, task_name, description, assigned_to, assigned_to_2, "
        " is_milestone, start_date, end_date, status, delay_days, progress, "
        " display_order, created_by, updated_by, "
        " is_event, event_start_time, event_end_time, planned_hours) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (category_id, subcategory_id, task_name, description, assigned_to, assigned_to_2,
         is_milestone, start_date, end_date, status, delay_days, calc_progress,
         max_order + 1, created_by, updated_by,
         is_event, event_start_time, event_end_time, planned_hours),
    )
    db.commit()
    return cur.lastrowid


def update_project_task(
    task_id: int,
    category_id: int | None,
    subcategory_id: int | None,
    task_name: str,
    description: str,
    start_date: str,
    end_date: str,
    status: str,
    progress: int,
    delay_days: int,
    updated_by: str,
    assigned_to: int | None = None,
    assigned_to_2: int | None = None,
    is_milestone: int = 0,
    is_event: int = 0,
    event_start_time: str = "",
    event_end_time: str = "",
    planned_hours: float = 0.0,
) -> None:
    """プロジェクトタスクを更新する。

    Args:
        task_id: タスクID
        category_id: 大区分ID
        subcategory_id: 中区分ID
        task_name: タスク名
        description: 対応内容
        start_date: 開始日
        end_date: 終了日
        status: 状態
        progress: 進捗率（手動指定時）
        delay_days: 遅延日数
        updated_by: 更新者名
        assigned_to: 担当者1ユーザーID
        assigned_to_2: 担当者2ユーザーID
        is_milestone: マイルストーンフラグ（1=マイルストーン）
        is_event: イベントフラグ（1=イベント/会議）
        event_start_time: イベント開始時刻（HH:MM形式）
        event_end_time: イベント終了時刻（HH:MM形式）
        planned_hours: 予定工数（時間）
    """
    db = get_db()
    calc_progress = _normalize_progress(status, progress, start_date, end_date)
    db.execute(
        "UPDATE project_task SET "
        "category_id=?, subcategory_id=?, task_name=?, description=?, "
        "assigned_to=?, assigned_to_2=?, is_milestone=?, start_date=?, end_date=?, "
        "status=?, delay_days=?, progress=?, "
        "is_event=?, event_start_time=?, event_end_time=?, planned_hours=?, "
        "updated_at=datetime('now','localtime'), updated_by=? "
        "WHERE id=?",
        (category_id, subcategory_id, task_name, description,
         assigned_to, assigned_to_2, is_milestone, start_date, end_date,
         status, delay_days, calc_progress,
         is_event, event_start_time, event_end_time, planned_hours,
         updated_by, task_id),
    )
    db.commit()


def delete_project_task(task_id: int) -> None:
    """プロジェクトタスクを削除する。

    Args:
        task_id: タスクID
    """
    db = get_db()
    db.execute("DELETE FROM project_task WHERE id = ?", (task_id,))
    db.commit()


def get_active_tasks_for_user(user_id: int) -> list[dict]:
    """指定ユーザーに割り当てられた未完了の通常タスク（イベント除く）を取得する。

    週間予定へのインポート候補として使用する。

    Args:
        user_id: ユーザーID

    Returns:
        list[dict]: タスク一覧（id, task_name, start_date, end_date, status, progress）
    """
    db = get_db()
    rows = db.execute(
        "SELECT pt.id, pt.task_name, pt.start_date, pt.end_date, "
        "  pt.status, pt.progress, pt.planned_hours, "
        "  tc.name AS category_name, ts.name AS subcategory_name "
        "FROM project_task pt "
        "LEFT JOIN task_category tc ON pt.category_id = tc.id "
        "LEFT JOIN task_subcategory ts ON pt.subcategory_id = ts.id "
        "WHERE (pt.assigned_to = ? OR pt.assigned_to_2 = ?) "
        "  AND pt.status NOT IN ('完了', '停止') "
        "  AND pt.is_event = 0 "
        "ORDER BY tc.display_order, ts.display_order, pt.display_order",
        (user_id, user_id),
    ).fetchall()
    return [dict(r) for r in rows]


def get_events_for_user_date(user_id: int, target_date: str) -> list[dict]:
    """指定ユーザー・日付に該当するイベントを取得する。

    start_date <= target_date <= end_date のイベントを返す。

    Args:
        user_id: ユーザーID
        target_date: 対象日付（YYYY-MM-DD形式）

    Returns:
        list[dict]: イベント一覧
    """
    db = get_db()
    rows = db.execute(
        "SELECT pt.id, pt.task_name, pt.description, pt.start_date, pt.end_date, "
        "  pt.event_start_time, pt.event_end_time, pt.planned_hours, "
        "  pt.status, pt.progress, "
        "  tc.name AS category_name, ts.name AS subcategory_name "
        "FROM project_task pt "
        "LEFT JOIN task_category tc ON pt.category_id = tc.id "
        "LEFT JOIN task_subcategory ts ON pt.subcategory_id = ts.id "
        "WHERE (pt.assigned_to = ? OR pt.assigned_to_2 = ?) "
        "  AND pt.is_event = 1 "
        "  AND pt.start_date <= ? AND pt.end_date >= ? "
        "ORDER BY pt.event_start_time ASC",
        (user_id, user_id, target_date, target_date),
    ).fetchall()
    return [dict(r) for r in rows]


def import_tasks_to_weekly_schedule(
    user_id: int,
    week_start: str,
    updated_by: str = "",
) -> int:
    """ログインユーザーに割り当てられた期間内タスクを週間予定に自動インポートする。

    各曜日（月〜金）について、その日を含む期間のタスクをAM/PMに均等配分して配置する。
    定例予約行はスキップする。

    Args:
        user_id: ユーザーID
        week_start: 週開始日（YYYY-MM-DD形式）
        updated_by: 更新者名

    Returns:
        int: 実際にインポートした件数
    """
    import sqlite3 as _sqlite3
    from datetime import timedelta as _timedelta

    db = get_db()
    now = datetime.now().isoformat()

    ws = datetime.strptime(week_start, "%Y-%m-%d")
    week_end: str = (ws + _timedelta(days=4)).strftime("%Y-%m-%d")

    # 対象週と期間が重なる、ユーザー担当タスクを取得（start_date/end_dateも取得）
    rows = db.execute(
        "SELECT pt.id, pt.task_name, COALESCE(ts.name,'') AS subcategory_name,"
        "       pt.start_date, pt.end_date "
        "FROM project_task pt "
        "LEFT JOIN task_subcategory ts ON ts.id = pt.subcategory_id "
        "WHERE (pt.assigned_to = ? OR pt.assigned_to_2 = ?) "
        "  AND pt.is_event = 0 "
        "  AND pt.status NOT IN ('完了', '停止') "
        "  AND pt.start_date <= ? AND pt.end_date >= ? "
        "ORDER BY pt.start_date, pt.id",
        (user_id, user_id, week_end, week_start),
    ).fetchall()

    if not rows:
        return 0

    # 定例予約行を除外
    reserved = get_reserved_row_numbers(user_id)
    reserved_am = {r - 1 for r in reserved if 1 <= r <= 5}
    reserved_pm = {r - 6 for r in reserved if 6 <= r <= 10}

    imported = 0

    # 曜日ごとに処理（月〜金）
    for day_idx in range(5):
        day_date: str = (ws + _timedelta(days=day_idx)).strftime("%Y-%m-%d")

        # この曜日を含む期間のタスクを抽出
        day_tasks = [r for r in rows if r["start_date"] <= day_date <= r["end_date"]]
        if not day_tasks:
            continue

        # この曜日に既に配置済みの project_task_id
        existing_day: set[int] = {
            r["project_task_id"]
            for r in db.execute(
                "SELECT project_task_id FROM weekly_schedule "
                "WHERE user_id=? AND week_start=? AND day_of_week=? "
                "  AND project_task_id IS NOT NULL",
                (user_id, week_start, day_idx),
            ).fetchall()
        }

        # この曜日の空きスロットを取得
        def _free_slots(slot: str, reserved_set: set[int]) -> list[int]:
            result = []
            for idx in range(5):
                if idx in reserved_set:
                    continue
                r = db.execute(
                    "SELECT task_name FROM weekly_schedule "
                    "WHERE user_id=? AND week_start=? AND day_of_week=? "
                    "  AND time_slot=? AND slot_index=?",
                    (user_id, week_start, day_idx, slot, idx),
                ).fetchone()
                if r is None or not (r["task_name"] or "").strip():
                    result.append(idx)
            return result

        am_free = _free_slots("am", reserved_am)
        pm_free = _free_slots("pm", reserved_pm)
        am_ptr, pm_ptr = 0, 0

        for i, task in enumerate(day_tasks):
            if task["id"] in existing_day:
                continue
            # AM/PM に交互配分
            if i % 2 == 0:
                if am_ptr < len(am_free):
                    slot_name, idx = "am", am_free[am_ptr]; am_ptr += 1
                elif pm_ptr < len(pm_free):
                    slot_name, idx = "pm", pm_free[pm_ptr]; pm_ptr += 1
                else:
                    break
            else:
                if pm_ptr < len(pm_free):
                    slot_name, idx = "pm", pm_free[pm_ptr]; pm_ptr += 1
                elif am_ptr < len(am_free):
                    slot_name, idx = "am", am_free[am_ptr]; am_ptr += 1
                else:
                    break
            try:
                db.execute(
                    "INSERT OR REPLACE INTO weekly_schedule "
                    "(user_id, week_start, day_of_week, time_slot, slot_index, "
                    " task_name, hours, subcategory_name, project_task_id, "
                    " updated_at, updated_by) "
                    "VALUES (?, ?, ?, ?, ?, ?, 0.0, ?, ?, ?, ?)",
                    (user_id, week_start, day_idx, slot_name, idx,
                     task["task_name"], task["subcategory_name"],
                     task["id"], now, updated_by),
                )
                imported += 1
                existing_day.add(task["id"])
            except _sqlite3.DatabaseError:
                pass

    if imported > 0:
        db.commit()
    return imported


def import_events_to_weekly_schedule(
    user_id: int,
    week_start: str,
    updated_by: str = "",
) -> int:
    """イベントを週間予定に自動配置する。

    イベントの開始時刻からAM/PMを判定し、該当曜日の空き枠に配置する。
    12:00未満はAM、12:00以降はPM。

    Args:
        user_id: ユーザーID
        week_start: 週開始日（YYYY-MM-DD形式）
        updated_by: 更新者名

    Returns:
        int: 配置したイベント数
    """
    import sqlite3 as _sqlite3

    db = get_db()
    now = datetime.now().isoformat()
    ws_date = date.fromisoformat(week_start)
    imported = 0

    # 既存の project_task_id を収集（重複防止）
    existing_rows = db.execute(
        "SELECT project_task_id FROM weekly_schedule "
        "WHERE user_id = ? AND week_start = ? AND project_task_id IS NOT NULL",
        (user_id, week_start),
    ).fetchall()
    existing_pt_ids: set[int] = {r["project_task_id"] for r in existing_rows}

    # 当週の月〜金の日付
    week_dates = [(ws_date + timedelta(days=i)).isoformat() for i in range(5)]

    for day_idx, day_date in enumerate(week_dates):
        events = get_events_for_user_date(user_id, day_date)
        for ev in events:
            if ev["id"] in existing_pt_ids:
                continue

            # 開始時刻からAM/PM判定
            start_time = ev.get("event_start_time") or "09:00"
            try:
                hour = int(start_time.split(":")[0])
            except (ValueError, IndexError):
                hour = 9
            time_slot = "pm" if hour >= 12 else "am"

            # 定例予約行をスキップ
            reserved = get_reserved_row_numbers(user_id)
            if time_slot == "am":
                reserved_slots = {r - 1 for r in reserved if 1 <= r <= 5}
            else:
                reserved_slots = {r - 6 for r in reserved if 6 <= r <= 10}

            # 該当スロットの空き枠を探す（定例行は除く）
            for idx in range(5):
                if idx in reserved_slots:
                    continue
                row = db.execute(
                    "SELECT task_name FROM weekly_schedule "
                    "WHERE user_id = ? AND week_start = ? AND day_of_week = ? "
                    "  AND time_slot = ? AND slot_index = ?",
                    (user_id, week_start, day_idx, time_slot, idx),
                ).fetchone()
                if row is None or not row["task_name"]:
                    # イベント名に時間を付与
                    time_label = f"【{start_time}】"
                    event_task_name = f"{time_label}{ev['task_name']}"
                    hours = ev.get("planned_hours") or 0.0
                    try:
                        db.execute(
                            "INSERT OR REPLACE INTO weekly_schedule "
                            "(user_id, week_start, day_of_week, time_slot, slot_index, "
                            " task_name, hours, subcategory_name, project_task_id, "
                            " updated_at, updated_by) "
                            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                            (user_id, week_start, day_idx, time_slot, idx,
                             event_task_name, hours,
                             ev.get("subcategory_name") or "",
                             ev["id"], now, updated_by),
                        )
                        existing_pt_ids.add(ev["id"])
                        imported += 1
                    except _sqlite3.DatabaseError:
                        pass
                    break

    if imported > 0:
        db.commit()
    return imported


def sync_daily_progress_to_task(
    user_id: int,
    date_str: str,
) -> int:
    """日次実績の作業時間をタスク管理の進捗に反映する。

    project_task_id が紐づいている実績スロットの作業時間を集計し、
    タスクの planned_hours に対する実績比率で進捗を更新する。
    ステータスが「完了」「停止」のタスクは更新しない。

    Args:
        user_id: ユーザーID
        date_str: 実績日付（YYYY-MM-DD形式）

    Returns:
        int: 更新したタスク数
    """
    db = get_db()
    updated = 0

    # 全実績から project_task_id 別に時間を集計
    rows = db.execute(
        "SELECT project_task_id, SUM(hours) AS total_hours "
        "FROM daily_result "
        "WHERE user_id = ? AND project_task_id IS NOT NULL "
        "GROUP BY project_task_id",
        (user_id,),
    ).fetchall()

    for row in rows:
        pt_id = row["project_task_id"]
        total_hours = row["total_hours"] or 0.0

        task = get_project_task_by_id(pt_id)
        if not task:
            continue
        if task["status"] in ("完了", "停止"):
            continue

        planned = task.get("planned_hours") or 0.0
        if planned <= 0:
            # 予定工数未設定：実績があれば「未着手」→「着手」に昇格
            if total_hours > 0 and task["status"] == "未着手":
                db.execute(
                    "UPDATE project_task SET status = '着手', "
                    "updated_at = datetime('now','localtime') "
                    "WHERE id = ?",
                    (pt_id,),
                )
                updated += 1
            continue

        new_progress = max(0, min(100, round(total_hours / planned * 100)))
        # 進捗100%到達時にステータスを「完了」へ昇格
        new_status = task["status"]
        if new_progress >= 100 and task["status"] not in ("完了", "停止"):
            new_status = "完了"
        if new_progress != (task.get("progress") or 0) or new_status != task["status"]:
            db.execute(
                "UPDATE project_task SET progress = ?, status = ?, "
                "updated_at = datetime('now','localtime') "
                "WHERE id = ?",
                (new_progress, new_status, pt_id),
            )
            updated += 1

    if updated > 0:
        db.commit()
    return updated


# ---------------------------------------------------------------------------
# ブラビオExcelインポート
# ---------------------------------------------------------------------------


def _find_user_by_partial_name(name: str, users: list[dict]) -> int | None:
    """ユーザー名の部分一致で該当ユーザーIDを返す。

    姓・名・フルネームのいずれかに部分一致するユーザーを検索する。
    一致しない場合は None を返す。

    Args:
        name: 検索する名前文字列
        users: get_all_users() の結果リスト

    Returns:
        int | None: 一致したユーザーID、または None
    """
    name = name.strip().replace("\u3000", " ")
    if not name:
        return None
    for u in users:
        full = (u.get("name") or "").replace("\u3000", " ")
        last = (u.get("last_name") or "").replace("\u3000", " ")
        first = (u.get("first_name") or "").replace("\u3000", " ")
        # 部分一致チェック
        if name in full or name in last or name in first:
            return u["id"]
        if full and full in name:
            return u["id"]
        # 姓だけでも一致判定
        if last and last in name:
            return u["id"]
    return None


def _map_brabio_status(status_str: str, progress: int) -> str:
    """ブラビオのステータスをプロジェクトタスクのステータスにマッピングする。

    Args:
        status_str: ブラビオのステータス文字列
        progress: 進捗率

    Returns:
        str: マッピング後のステータス
    """
    s = status_str.strip()
    if s == "完了":
        return "完了"
    if s == "停止":
        return "停止"
    if s == "未着手":
        return "未着手"
    if s in ("作業中", ""):
        if progress >= 100:
            return "完了"
        if progress > 0:
            return "着手"
        return "未着手"
    return "着手"


def import_brabio_excel(
    file_path: str,
    created_by: int,
    updated_by: str,
) -> dict:
    """ブラビオExcelファイルからタスクをインポートする。

    folder行を大区分・中区分、task行をプロジェクトタスクとして取り込む。
    milestone行はマイルストーンとして取り込む。
    既に同名タスクが存在する場合はスキップする。
    担当者はユーザー名の部分一致でマッチングする。

    Args:
        file_path: Excelファイルパス
        created_by: 作成者ユーザーID
        updated_by: 更新者名

    Returns:
        dict: {"imported": int, "skipped": int, "errors": list[str]}
    """
    import openpyxl as _openpyxl

    result = {"imported": 0, "skipped": 0, "errors": []}

    try:
        wb = _openpyxl.load_workbook(file_path, data_only=True)
    except Exception as exc:
        result["errors"].append(f"ファイルを開けませんでした: {exc}")
        return result

    ws = wb.worksheets[0]
    all_users = get_all_users()

    # 既存タスク名→IDのマップ（重複時は更新用）
    existing_tasks = get_all_project_tasks()
    existing_name_to_id: dict[str, int] = {t["task_name"]: t["id"] for t in existing_tasks}
    result["updated"] = 0

    # 大区分・中区分のキャッシュ
    categories = get_all_categories()
    cat_name_map: dict[str, int] = {c["name"]: c["id"] for c in categories}

    subcategories = get_all_subcategories()
    subcat_name_map: dict[str, int] = {s["name"]: s["id"] for s in subcategories}

    # folder 階層を追跡（カテゴリ割り当て用）
    # outline=1 の folder → 大区分候補
    # outline=2 の folder → 中区分候補
    current_cat_name: str = ""
    current_cat_id: int | None = None
    current_subcat_name: str = ""
    current_subcat_id: int | None = None

    def _ensure_category(name: str) -> int | None:
        """大区分をDBに登録し、IDを返す。既存なら既存IDを返す。"""
        if name in cat_name_map:
            return cat_name_map[name]
        add_category(name)
        refreshed = get_all_categories()
        for c in refreshed:
            if c["name"] == name:
                cat_name_map[name] = c["id"]
                return c["id"]
        return None

    def _ensure_subcategory(cat_id: int, name: str) -> int | None:
        """中区分をDBに登録し、IDを返す。既存なら既存IDを返す。"""
        if name in subcat_name_map:
            return subcat_name_map[name]
        if cat_id is None:
            return None
        add_subcategory(cat_id, name)
        refreshed = get_all_subcategories()
        for s in refreshed:
            if s["name"] == name:
                subcat_name_map[name] = s["id"]
                return s["id"]
        return None

    for r in range(5, ws.max_row + 1):
        row_type = str(ws.cell(r, 1).value or "").strip()
        outline = ws.cell(r, 2).value
        try:
            outline_level = int(outline) if outline is not None else 0
        except (ValueError, TypeError):
            outline_level = 0
        title = str(ws.cell(r, 4).value or "").strip()
        start_raw = str(ws.cell(r, 5).value or "").strip()
        end_raw = str(ws.cell(r, 6).value or "").strip()
        status_raw = str(ws.cell(r, 7).value or "").strip()
        progress_raw = ws.cell(r, 8).value
        member_raw = str(ws.cell(r, 11).value or "").strip()

        if not title:
            continue

        # 進捗率
        try:
            progress = max(0, min(int(progress_raw), 100)) if progress_raw else 0
        except (ValueError, TypeError):
            progress = 0

        # 日付変換（YYYY/MM/DD → YYYY-MM-DD）
        start_date = start_raw.replace("/", "-") if start_raw else ""
        end_date = end_raw.replace("/", "-") if end_raw else ""

        # folder 行 → カテゴリ追跡（タスクとしては取り込まない）
        if row_type == "folder":
            if outline_level == 1:
                current_cat_name = title
                current_cat_id = _ensure_category(title)
                current_subcat_name = ""
                current_subcat_id = None
            elif outline_level == 2:
                current_subcat_name = title
                current_subcat_id = _ensure_subcategory(current_cat_id, title)
            continue

        # project 行 → スキップ
        if row_type == "project":
            continue

        # task / milestone 行 → インポート対象
        if row_type not in ("task", "milestone"):
            continue

        # 改行含むタイトルは最初の行のみ
        title = title.split("\n")[0].strip()

        # ステータスマッピング
        status = _map_brabio_status(status_raw, progress)

        # 担当者マッチング（複数名の場合は改行区切り）
        assigned_to: int | None = None
        assigned_to_2: int | None = None
        if member_raw:
            members = [m.strip() for m in member_raw.split("\n") if m.strip()]
            if len(members) >= 1:
                assigned_to = _find_user_by_partial_name(members[0], all_users)
            if len(members) >= 2:
                assigned_to_2 = _find_user_by_partial_name(members[1], all_users)

        is_milestone = 1 if row_type == "milestone" else 0

        # 日付がない場合はデフォルト
        if not start_date:
            start_date = date.today().isoformat()
        if not end_date:
            end_date = start_date

        # 既存タスク → 更新、新規 → 追加
        if title in existing_name_to_id:
            try:
                update_project_task(
                    task_id=existing_name_to_id[title],
                    category_id=current_cat_id,
                    subcategory_id=current_subcat_id,
                    task_name=title,
                    description="",
                    start_date=start_date,
                    end_date=end_date,
                    status=status,
                    progress=progress,
                    delay_days=0,
                    updated_by=updated_by,
                    assigned_to=assigned_to,
                    assigned_to_2=assigned_to_2,
                    is_milestone=is_milestone,
                )
                result["updated"] += 1
            except Exception as exc:
                result["errors"].append(f"行{r}「{title}」更新エラー: {exc}")
            continue

        try:
            add_project_task(
                category_id=current_cat_id,
                subcategory_id=current_subcat_id,
                task_name=title,
                description="",
                start_date=start_date,
                end_date=end_date,
                status=status,
                progress=progress,
                delay_days=0,
                created_by=created_by,
                updated_by=updated_by,
                assigned_to=assigned_to,
                assigned_to_2=assigned_to_2,
                is_milestone=is_milestone,
                planned_hours=0.0,
            )
            existing_name_to_id[title] = -1  # ダミーID（再重複防止）
            result["imported"] += 1
        except Exception as exc:
            result["errors"].append(f"行{r}「{title}」: {exc}")

    return result


# ---------------------------------------------------------------------------
# ダッシュボード用ヘルパー関数
# ---------------------------------------------------------------------------


def get_task_progress_summary(user_id: int | None = None) -> dict:
    """指定ユーザーのプロジェクトタスク進捗サマリーを取得する。

    既存の get_all_project_tasks を利用してタスクを取得し、
    Python 側でステータス別件数・遅延件数・平均進捗率を集計する。

    Args:
        user_id: 集計対象のユーザーID。None の場合は全タスクを集計する。

    Returns:
        dict: 以下のキーを持つ辞書。
            - total_count (int): タスク総数
            - completed_count (int): 完了タスク数
            - delayed_count (int): 遅延タスク数（status=="遅れ" or delay_days>0）
            - avg_progress (float): 平均進捗率（0.0〜100.0）
            - status_breakdown (dict[str, int]): ステータス別件数
            - tasks (list[dict]): 個別タスク情報
    """
    tasks: list[dict] = get_all_project_tasks(assigned_to=user_id)

    # ステータス別件数を全ステータスキーで初期化
    status_breakdown: dict[str, int] = {s: 0 for s in PROJECT_TASK_STATUSES}

    completed_count: int = 0
    delayed_count: int = 0
    total_progress: float = 0.0

    for task in tasks:
        status: str = task.get("status", "")
        if status in status_breakdown:
            status_breakdown[status] += 1

        if status == "完了":
            completed_count += 1

        # 遅延判定: status が「遅れ」、または delay_days > 0
        delay_days: int = task.get("delay_days", 0) or 0
        if status == "遅れ" or delay_days > 0:
            delayed_count += 1

        total_progress += float(task.get("progress", 0) or 0)

    total_count: int = len(tasks)
    avg_progress: float = round(total_progress / total_count, 1) if total_count > 0 else 0.0

    return {
        "total_count": total_count,
        "completed_count": completed_count,
        "delayed_count": delayed_count,
        "avg_progress": avg_progress,
        "status_breakdown": status_breakdown,
        "tasks": tasks,
    }


def get_task_overview_summary() -> dict:
    """管理者向けタスク全体俯瞰サマリーを取得する。

    全タスクを対象に、ステータス別件数・カテゴリ別集計・担当者別集計を生成する。
    管理職・マスタ向けダッシュボードで使用。

    Returns:
        dict: 以下のキーを持つ辞書。
            - total_count (int): タスク総数
            - completed_count (int): 完了タスク数
            - delayed_count (int): 遅延タスク数
            - avg_progress (float): 平均進捗率
            - status_breakdown (dict[str, int]): ステータス別件数
            - category_summary (list[dict]): カテゴリ別の件数・平均進捗
            - user_summary (list[dict]): 担当者別の件数・平均進捗
            - tasks (list[dict]): 全タスク情報
    """
    tasks: list[dict] = get_all_project_tasks()

    status_breakdown: dict[str, int] = {s: 0 for s in PROJECT_TASK_STATUSES}
    completed_count: int = 0
    delayed_count: int = 0
    total_progress: float = 0.0

    # カテゴリ別集計用
    cat_data: dict[str, dict] = {}
    # 担当者別集計用
    user_data: dict[str, dict] = {}

    for task in tasks:
        status: str = task.get("status", "")
        if status in status_breakdown:
            status_breakdown[status] += 1
        if status == "完了":
            completed_count += 1
        delay_days: int = task.get("delay_days", 0) or 0
        if status == "遅れ" or delay_days > 0:
            delayed_count += 1
        prog: float = float(task.get("progress", 0) or 0)
        total_progress += prog

        # カテゴリ別
        cat_name: str = task.get("category_name") or "（未分類）"
        if cat_name not in cat_data:
            cat_data[cat_name] = {"name": cat_name, "count": 0, "progress_sum": 0.0,
                                  "completed": 0, "delayed": 0}
        cat_data[cat_name]["count"] += 1
        cat_data[cat_name]["progress_sum"] += prog
        if status == "完了":
            cat_data[cat_name]["completed"] += 1
        if status == "遅れ" or delay_days > 0:
            cat_data[cat_name]["delayed"] += 1

        # 担当者別
        user_name: str = task.get("assigned_name") or "（未割当）"
        user_id_val: int | None = task.get("assigned_to")
        if user_name not in user_data:
            user_data[user_name] = {"name": user_name, "user_id": user_id_val,
                                    "count": 0, "progress_sum": 0.0,
                                    "completed": 0, "delayed": 0}
        user_data[user_name]["count"] += 1
        user_data[user_name]["progress_sum"] += prog
        if status == "完了":
            user_data[user_name]["completed"] += 1
        if status == "遅れ" or delay_days > 0:
            user_data[user_name]["delayed"] += 1

    total_count: int = len(tasks)
    avg_progress: float = round(total_progress / total_count, 1) if total_count > 0 else 0.0

    # カテゴリ別平均進捗を計算
    category_summary: list[dict] = []
    for cd in cat_data.values():
        cd["avg_progress"] = round(cd["progress_sum"] / cd["count"], 1) if cd["count"] > 0 else 0.0
        category_summary.append(cd)
    category_summary.sort(key=lambda x: x["name"])

    # 担当者別平均進捗を計算
    user_summary: list[dict] = []
    for ud in user_data.values():
        ud["avg_progress"] = round(ud["progress_sum"] / ud["count"], 1) if ud["count"] > 0 else 0.0
        user_summary.append(ud)
    user_summary.sort(key=lambda x: -x["count"])

    return {
        "total_count": total_count,
        "completed_count": completed_count,
        "delayed_count": delayed_count,
        "avg_progress": avg_progress,
        "status_breakdown": status_breakdown,
        "category_summary": category_summary,
        "user_summary": user_summary,
        "tasks": tasks,
    }


def get_accessible_users_for_dashboard(
    login_user_id: int, login_role: str, login_dept: str
) -> list[dict]:
    """ダッシュボードでログインユーザーが閲覧可能なユーザー一覧を返す。

    役職に応じたアクセス範囲:
        - 一般ユーザー: 空リスト（自分のみ閲覧のため一覧不要）
        - 管理職: manager_id が login_user_id のユーザー＋自分自身
        - マスタ: 同一部署の全ユーザー（dept_filter=login_dept）

    Args:
        login_user_id: ログインユーザーのID。
        login_role: ログインユーザーの役職。
        login_dept: ログインユーザーの部署。

    Returns:
        list[dict]: 閲覧可能なユーザー一覧。各要素は {id, name, dept} を含む。
    """
    if login_role == "マスタ":
        dept_f = login_dept if login_dept else None
        all_users = get_all_users(dept_filter=dept_f)
        return [{"id": u["id"], "name": u["name"], "dept": u.get("dept", "")} for u in all_users]

    if login_role == "管理職":
        direct_reports = get_direct_reports(login_user_id)
        result: list[dict] = []
        # 自分自身を先頭に追加
        self_user = get_user_by_id(login_user_id)
        if self_user:
            result.append({"id": self_user["id"], "name": self_user["name"], "dept": self_user.get("dept", "")})
        for u in direct_reports:
            if u["id"] != login_user_id:
                result.append({"id": u["id"], "name": u["name"], "dept": u.get("dept", "")})
        return result

    # 一般ユーザー: 空リスト
    return []


# ---------------------------------------------------------------------------
# 休日管理（土日・祝日・会社休日）
# ---------------------------------------------------------------------------

def _get_jpholidays(year: int) -> set[date]:
    """指定年の日本の祝日をセットで返す。

    jpholiday パッケージが利用可能な場合はそれを使用し、
    利用できない場合は空のセットを返す。

    Args:
        year: 対象年

    Returns:
        set[date]: 祝日の日付セット
    """
    try:
        import jpholiday
        result: set[date] = set()
        d = date(year, 1, 1)
        end = date(year, 12, 31)
        while d <= end:
            if jpholiday.is_holiday(d):
                result.add(d)
            d += timedelta(days=1)
        return result
    except ImportError:
        logger.warning("jpholiday パッケージが見つかりません。祝日は考慮されません。")
        return set()


def get_company_holidays(year: int | None = None) -> list[dict]:
    """会社独自の休日一覧を返す。

    Args:
        year: 対象年。None の場合は全件返す。

    Returns:
        list[dict]: {id, holiday_date, holiday_name, created_at} のリスト
    """
    db = get_db()
    if year is not None:
        rows = db.execute(
            "SELECT id, holiday_date, holiday_name, created_at "
            "FROM company_holiday "
            "WHERE strftime('%Y', holiday_date) = ? "
            "ORDER BY holiday_date",
            (str(year),),
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT id, holiday_date, holiday_name, created_at "
            "FROM company_holiday ORDER BY holiday_date"
        ).fetchall()
    return [dict(r) for r in rows]


def add_company_holiday(
    holiday_date: str,
    holiday_name: str,
    created_by: int,
) -> None:
    """会社休日を登録する（マスタ権限のみ使用）。

    Args:
        holiday_date: 休日日付（YYYY-MM-DD）
        holiday_name: 休日名称
        created_by: 登録者ユーザーID
    """
    db = get_db()
    db.execute(
        "INSERT OR IGNORE INTO company_holiday (holiday_date, holiday_name, created_by) "
        "VALUES (?, ?, ?)",
        (holiday_date, holiday_name, created_by),
    )
    db.commit()


def delete_company_holiday(holiday_id: int) -> None:
    """会社休日を削除する。

    Args:
        holiday_id: 削除する会社休日のID
    """
    db = get_db()
    db.execute("DELETE FROM company_holiday WHERE id = ?", (holiday_id,))
    db.commit()


def get_all_holidays(start: date, end: date) -> set[date]:
    """指定期間の全休日（土日＋祝日＋会社休日）をセットで返す。

    Args:
        start: 開始日
        end: 終了日

    Returns:
        set[date]: 期間内の休日日付セット
    """
    holidays: set[date] = set()

    # 土日
    d = start
    while d <= end:
        if d.weekday() >= 5:  # 5=土, 6=日
            holidays.add(d)
        d += timedelta(days=1)

    # 日本の祝日（対象年をすべてカバー）
    for year in range(start.year, end.year + 1):
        holidays |= _get_jpholidays(year)

    # 会社独自休日
    db = get_db()
    rows = db.execute(
        "SELECT holiday_date FROM company_holiday "
        "WHERE holiday_date BETWEEN ? AND ?",
        (start.isoformat(), end.isoformat()),
    ).fetchall()
    for row in rows:
        try:
            holidays.add(date.fromisoformat(row["holiday_date"]))
        except ValueError:
            pass

    return holidays


def count_business_days(start: date, end: date) -> int:
    """開始日〜終了日（両端含む）の営業日数を返す。

    土日・祝日・会社休日を除いた日数を計算する。

    Args:
        start: 開始日
        end: 終了日

    Returns:
        int: 営業日数（0以上）
    """
    if start > end:
        return 0
    holidays = get_all_holidays(start, end)
    count = 0
    d = start
    while d <= end:
        if d not in holidays:
            count += 1
        d += timedelta(days=1)
    return count


def calc_planned_hours(start_date_str: str, end_date_str: str, std_hours_per_day: float = 8.0) -> float:
    """タスクの期間から予定工数（時間）を自動計算する。

    営業日数（土日・祝日・会社休日除く）× 1日の標準時間で算出する。

    Args:
        start_date_str: 開始日（YYYY-MM-DD）
        end_date_str: 終了日（YYYY-MM-DD）
        std_hours_per_day: 1営業日あたりの標準時間（デフォルト8.0h）

    Returns:
        float: 予定工数（時間）。期間が不正な場合は 0.0 を返す。
    """
    try:
        start = date.fromisoformat(start_date_str)
        end = date.fromisoformat(end_date_str)
    except ValueError:
        return 0.0
    bdays = count_business_days(start, end)
    return round(bdays * std_hours_per_day, 1)


# ─────────────────────────────────────────────
# 定例スケジュール管理
# ─────────────────────────────────────────────

def get_routine_schedules(user_id: int) -> list[dict]:
    """ユーザーの定例スケジュール一覧を取得する。

    Args:
        user_id: ユーザーID

    Returns:
        list[dict]: row_number 昇順の定例スケジュール一覧
    """
    db = get_db()
    rows = db.execute(
        "SELECT id, user_id, task_name, subcategory_name, default_hours, row_number"
        " FROM routine_schedule WHERE user_id = ? ORDER BY row_number ASC",
        (user_id,),
    ).fetchall()
    return [dict(r) for r in rows]


def save_routine_task(
    user_id: int,
    task_name: str,
    subcategory_name: str,
    default_hours: float,
    row_number: int,
) -> bool:
    """定例スケジュールを登録（または上書き）する。

    同一ユーザーの同一行番号がある場合は置き換える。

    Args:
        user_id: ユーザーID
        task_name: 作業名
        subcategory_name: 中区分名
        default_hours: デフォルト工数
        row_number: 行番号（1〜10）

    Returns:
        bool: 成功時 True
    """
    if not 1 <= row_number <= 10:
        return False
    db = get_db()
    try:
        db.execute(
            "INSERT INTO routine_schedule"
            " (user_id, task_name, subcategory_name, default_hours, row_number)"
            " VALUES (?, ?, ?, ?, ?)"
            " ON CONFLICT(user_id, row_number) DO UPDATE SET"
            "   task_name=excluded.task_name,"
            "   subcategory_name=excluded.subcategory_name,"
            "   default_hours=excluded.default_hours",
            (user_id, task_name, subcategory_name, default_hours, row_number),
        )
        db.commit()
        return True
    except Exception:
        return False


def delete_routine_task(routine_id: int, user_id: int) -> None:
    """定例スケジュールを削除する。

    Args:
        routine_id: 定例スケジュールID
        user_id: 所有ユーザーID（権限チェック用）
    """
    db = get_db()
    db.execute(
        "DELETE FROM routine_schedule WHERE id = ? AND user_id = ?",
        (routine_id, user_id),
    )
    db.commit()


def get_reserved_row_numbers(user_id: int) -> set[int]:
    """ユーザーの定例スケジュールで使用中の行番号セットを返す。

    タスク管理反映時に定例行をスキップするために使用する。

    Args:
        user_id: ユーザーID

    Returns:
        set[int]: 使用中の行番号（1〜10）
    """
    db = get_db()
    rows = db.execute(
        "SELECT row_number FROM routine_schedule WHERE user_id = ?",
        (user_id,),
    ).fetchall()
    return {r["row_number"] for r in rows}


def apply_routine_to_week(user_id: int, week_start: str, updated_by: str = "") -> None:
    """定例スケジュールを週間予定の空き行に適用する。

    定例スケジュールの row_number (1〜10) を AM(1-5) / PM(6-10) に変換し、
    当該スロットが空の場合のみ書き込む。

    Args:
        user_id: ユーザーID
        week_start: 週開始日（YYYY-MM-DD）
        updated_by: 更新者名
    """
    routines = get_routine_schedules(user_id)
    if not routines:
        return

    db = get_db()
    now = datetime.now().isoformat()

    for r in routines:
        row_num = r["row_number"]  # 1-10
        if row_num <= 5:
            time_slot = "am"
            slot_index = row_num - 1
        else:
            time_slot = "pm"
            slot_index = row_num - 6

        for day_idx in range(5):
            existing = db.execute(
                "SELECT task_name FROM weekly_schedule"
                " WHERE user_id=? AND week_start=? AND day_of_week=?"
                "   AND time_slot=? AND slot_index=?",
                (user_id, week_start, day_idx, time_slot, slot_index),
            ).fetchone()
            if existing is None or not (existing["task_name"] or "").strip():
                db.execute(
                    "INSERT OR REPLACE INTO weekly_schedule"
                    " (user_id, week_start, day_of_week, time_slot, slot_index,"
                    "  task_name, hours, subcategory_name, updated_at, updated_by)"
                    " VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (user_id, week_start, day_idx, time_slot, slot_index,
                     r["task_name"], r["default_hours"],
                     r["subcategory_name"] or "", now, updated_by),
                )
    db.commit()
